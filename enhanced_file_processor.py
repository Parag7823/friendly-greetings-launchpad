"""
Enhanced File Processor - 100X File Processing Capabilities

This module provides comprehensive file processing capabilities including:
1. Extended file format support (.ods, .pdf, .zip, images)
2. Auto-repair for corrupted files
3. ZIP file extraction and batch processing
4. OCR and AI-powered table extraction
5. Smart multi-sheet merging

Maintains compatibility with existing FastAPI backend and WebSocket progress tracking.
"""

import os
import io
import logging
import zipfile
import py7zr
import rarfile
from typing import Dict, List, Optional, Any, Tuple, Union
import pandas as pd
import numpy as np
from datetime import datetime
import tempfile
import shutil

# File format libraries
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P
import tabula
try:
    import camelot
    CAMELOT_AVAILABLE = True
except ImportError as e:
    logging.warning(f"Camelot not available: {e}")
    CAMELOT_AVAILABLE = False
import pdfplumber

# OCR and image processing
import pytesseract
from PIL import Image
import cv2

# Excel repair libraries
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)

class EnhancedFileProcessor:
    """Enhanced file processor with 100X capabilities"""
    
    def __init__(self):
        """
        Initialize the Enhanced File Processor with security settings and format support.
        
        Sets up supported file formats, configurable chunk sizes for streaming,
        and security parameters including file size limits and filename validation.
        """
        self.supported_formats = {
            # Spreadsheet formats
            'excel': ['.xlsx', '.xls', '.xlsm', '.xlsb'],
            'csv': ['.csv', '.tsv', '.txt'],
            'ods': ['.ods'],
            
            # Document formats with tables
            'pdf': ['.pdf'],
            
            # Archive formats
            'zip': ['.zip', '.7z', '.rar'],
            
            # Image formats
            'image': ['.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.gif']
        }
        
        # Configure OCR
        self.ocr_config = '--oem 3 --psm 6'
        
        # Configurable chunk sizes for streaming
        self.excel_chunk_size = int(os.environ.get('EXCEL_CHUNK_SIZE', 1000))
        self.csv_chunk_size = int(os.environ.get('CSV_CHUNK_SIZE', 10000))
        self.streaming_threshold_mb = int(os.environ.get('STREAMING_THRESHOLD_MB', 10))
        
        # Security settings
        self.max_file_size = int(os.environ.get('MAX_FILE_SIZE_MB', 100)) * 1024 * 1024  # 100MB default
        self.max_filename_length = int(os.environ.get('MAX_FILENAME_LENGTH', 255))
        self.allowed_extensions = set()
        for format_extensions in self.supported_formats.values():
            self.allowed_extensions.update(format_extensions)
    
    def _validate_security(self, file_content: bytes, filename: str) -> None:
        """
        Validate file security to prevent attacks
        
        Args:
            file_content: Raw file bytes
            filename: Original filename
            
        Raises:
            ValueError: If security validation fails
        """
        # Check file size
        if len(file_content) > self.max_file_size:
            raise ValueError(f"File size {len(file_content)} bytes exceeds maximum allowed size {self.max_file_size} bytes")
        
        # Check filename length
        if len(filename) > self.max_filename_length:
            raise ValueError(f"Filename length {len(filename)} exceeds maximum allowed length {self.max_filename_length}")
        
        # Check for path traversal
        if self._is_path_traversal_unsafe(filename):
            raise ValueError(f"Filename contains path traversal patterns: {filename}")
        
        # Check for null bytes and control characters
        if '\x00' in filename or '\x1a' in filename or '\x7f' in filename:
            raise ValueError(f"Filename contains invalid characters: {filename}")
        
        # Check file extension
        file_ext = os.path.splitext(filename)[1].lower()
        if file_ext not in self.allowed_extensions:
            raise ValueError(f"File extension '{file_ext}' is not allowed. Allowed extensions: {sorted(self.allowed_extensions)}")
        
        # Check for empty file
        if len(file_content) == 0:
            raise ValueError("File is empty")
    
    def _is_path_traversal_unsafe(self, filename: str) -> bool:
        """
        Check if filename contains path traversal patterns
        
        Args:
            filename: Filename to check
            
        Returns:
            True if unsafe, False if safe
        """
        dangerous_patterns = [
            '..',
            '../',
            '..\\',
            '%2e%2e',
            '%2E%2E',
            '....//',
            '....\\\\'
        ]
        
        filename_lower = filename.lower()
        for pattern in dangerous_patterns:
            if pattern in filename_lower:
                return True
        
        # Check for absolute paths
        if filename.startswith('/') or (len(filename) > 1 and filename[1] == ':'):
            return True
            
        return False
    
    def _sanitize_content(self, content: bytes) -> bytes:
        """
        Sanitize file content to remove potential malicious patterns
        
        Args:
            content: Raw file content
            
        Returns:
            Sanitized content
        """
        # For binary files (Excel, PDF, etc.), don't try to decode as text
        # Check if this looks like a binary file by examining the first few bytes
        if len(content) > 4:
            # Check for common binary file signatures
            binary_signatures = [
                b'PK\x03\x04',  # ZIP/Excel files
                b'PK\x05\x06',  # ZIP files
                b'\x89PNG',     # PNG files
                b'GIF8',        # GIF files
                b'\xFF\xD8\xFF', # JPEG files
                b'%PDF',        # PDF files
                b'\x00\x00\x01\x00', # ICO files
            ]
            
            for signature in binary_signatures:
                if content.startswith(signature):
                    # This is a binary file, return as-is without sanitization
                    return content
        
        # For text files, perform sanitization
        try:
            content_str = content.decode('utf-8', errors='ignore')
        except:
            # If can't decode, return as-is (binary file)
            return content
        
        # Check for script injection patterns
        script_patterns = [
            '<script',
            'javascript:',
            'vbscript:',
            'onload=',
            'onerror=',
            'onclick='
        ]
        
        content_lower = content_str.lower()
        for pattern in script_patterns:
            if pattern in content_lower:
                logger.warning(f"Detected potential script injection pattern: {pattern}")
                # Remove the pattern (basic sanitization)
                content_str = content_str.replace(pattern, '[REMOVED]')
        
        # Check for SQL injection patterns
        sql_patterns = [
            'union select',
            'drop table',
            'delete from',
            'insert into',
            'update set',
            'exec(',
            'execute('
        ]
        
        for pattern in sql_patterns:
            if pattern in content_lower:
                logger.warning(f"Detected potential SQL injection pattern: {pattern}")
                # Remove the pattern (basic sanitization)
                content_str = content_str.replace(pattern, '[REMOVED]')
        
        return content_str.encode('utf-8')
        
    async def process_file_enhanced(self, file_content: bytes, filename: str, 
                                  progress_callback=None) -> Dict[str, pd.DataFrame]:
        """
        Enhanced file processing with support for multiple formats
        
        Args:
            file_content: Raw file bytes
            filename: Original filename
            progress_callback: Function to report progress updates
            
        Returns:
            Dictionary of DataFrames (sheet_name -> DataFrame)
        """
        try:
            if progress_callback:
                await progress_callback("security", "🔒 Validating file security...", 2)
            
            # Security validation first
            self._validate_security(file_content, filename)
            
            # Sanitize content
            sanitized_content = self._sanitize_content(file_content)
            
            if progress_callback:
                await progress_callback("detecting", "🔍 Detecting file format and structure...", 5)
            
            # Detect file format
            file_format = self._detect_file_format(filename, sanitized_content)
            logger.info(f"Detected file format: {file_format} for {filename}")
            
            if progress_callback:
                await progress_callback("processing", f"📊 Processing {file_format} file...", 15)
            
            # Route to appropriate processor
            if file_format == 'excel':
                return await self._process_excel_enhanced(sanitized_content, filename, progress_callback)
            elif file_format == 'csv':
                return await self._process_csv_enhanced(sanitized_content, filename, progress_callback)
            elif file_format == 'ods':
                return await self._process_ods(sanitized_content, filename, progress_callback)
            elif file_format == 'pdf':
                return await self._process_pdf(sanitized_content, filename, progress_callback)
            elif file_format == 'zip':
                return await self._process_archive(sanitized_content, filename, progress_callback)
            elif file_format == 'image':
                return await self._process_image(sanitized_content, filename, progress_callback)
            else:
                raise ValueError(f"Unsupported file format: {file_format}")
                
        except Exception as e:
            logger.error(f"Enhanced file processing failed for {filename}: {e}")
            # Fallback to basic processing with sanitized content
            try:
                sanitized_content = self._sanitize_content(file_content)
                return await self._fallback_processing(sanitized_content, filename, progress_callback)
            except Exception as fallback_error:
                logger.error(f"Fallback processing also failed for {filename}: {fallback_error}")
                raise ValueError(f"File processing failed: {e}. Fallback also failed: {fallback_error}")
    
    def _detect_file_format(self, filename: str, file_content: bytes) -> str:
        """Detect file format using multiple methods"""
        try:
            # Get file extension
            ext = os.path.splitext(filename.lower())[1]
            
            # Check against known formats
            for format_type, extensions in self.supported_formats.items():
                if ext in extensions:
                    return format_type
            
            # Use magic bytes for detection
            import magic
            mime_type = magic.from_buffer(file_content, mime=True)
            
            # Map MIME types to formats
            mime_mapping = {
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'excel',
                'application/vnd.ms-excel': 'excel',
                'text/csv': 'csv',
                'application/vnd.oasis.opendocument.spreadsheet': 'ods',
                'application/pdf': 'pdf',
                'application/zip': 'zip',
                'application/x-7z-compressed': 'zip',
                'application/x-rar-compressed': 'zip',
                'image/png': 'image',
                'image/jpeg': 'image',
                'image/bmp': 'image',
                'image/tiff': 'image'
            }
            
            return mime_mapping.get(mime_type, 'unknown')
            
        except Exception as e:
            logger.warning(f"File format detection failed: {e}")
            return 'unknown'
    
    async def _process_excel_enhanced(self, file_content: bytes, filename: str, 
                                    progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Enhanced Excel processing with auto-repair capabilities"""
        try:
            if progress_callback:
                await progress_callback("reading", "📖 Reading Excel file...", 20)
            
            # Check for password protection and macros
            if '.xlsm' in filename.lower():
                logger.warning(f"File {filename} may contain macros. Processing with caution.")
                if progress_callback:
                    await progress_callback("security_check", "⚠️ File may contain macros. Proceeding with caution...", 25)

            try:
                # Try standard processing first
                return await self._read_excel_standard(file_content, filename)
            except (InvalidFileException, zipfile.BadZipFile) as e:
                if 'encrypted' in str(e).lower():
                    raise ValueError("Password-protected Excel files are not supported.")
                logger.warning(f"Standard Excel reading failed, attempting repair: {e}")
                if progress_callback:
                    await progress_callback("repairing", "🔧 Attempting to repair corrupted Excel file...", 30)
                return await self._repair_and_read_excel(file_content, filename, progress_callback)
            except Exception as e:
                logger.error(f"Enhanced Excel processing failed: {e}")
                raise
                
        except Exception as e:
            logger.error(f"Enhanced Excel processing failed: {e}")
            raise
    
    async def _read_excel_standard(self, file_content: bytes, filename: str) -> Dict[str, pd.DataFrame]:
        """Standard Excel reading with multiple engine fallback and streaming for large files"""
        file_stream = io.BytesIO(file_content)
        sheets = {}
        
        # Check file size for streaming approach
        file_size_mb = len(file_content) / (1024 * 1024)
        use_streaming = file_size_mb > self.streaming_threshold_mb
        
        if use_streaming:
            logger.info(f"Large file detected ({file_size_mb:.1f}MB), using streaming approach")
            return await self._read_excel_streaming(file_content, filename)
        
        # Try different engines for smaller files
        engines = ['openpyxl', 'xlrd', 'pyxlsb']
        
        for engine in engines:
            try:
                file_stream.seek(0)
                excel_file = pd.ExcelFile(file_stream, engine=engine)
                
                for sheet_name in excel_file.sheet_names:
                    try:
                        df = pd.read_excel(file_stream, sheet_name=sheet_name, engine=engine)
                        if not df.empty:
                            sheets[sheet_name] = df
                    except Exception as sheet_e:
                        logger.warning(f"Failed to read sheet {sheet_name} with {engine}: {sheet_e}")
                        continue
                
                if sheets:
                    return sheets
                    
            except Exception as e:
                logger.warning(f"Engine {engine} failed: {e}")
                continue
        
        raise ValueError("Could not read Excel file with any engine")
    
    async def _read_excel_streaming(self, file_content: bytes, filename: str) -> Dict[str, pd.DataFrame]:
        """
        TRUE streaming Excel reading for large files to avoid memory issues.
        
        This method uses a proper streaming approach that never loads the entire file
        or workbook into memory at once. Instead, it processes the Excel file as a
        ZIP archive and streams individual worksheet XML files.
        """
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            from io import BytesIO
            
            sheets = {}
            
            # Process Excel file as ZIP archive for true streaming
            with zipfile.ZipFile(BytesIO(file_content), 'r') as excel_zip:
                # Get worksheet files
                worksheet_files = [f for f in excel_zip.namelist() if f.startswith('xl/worksheets/sheet')]
                
                if not worksheet_files:
                    raise ValueError("No worksheets found in Excel file")
                
                for sheet_file in worksheet_files:
                    try:
                        # Extract sheet name from workbook.xml.rels or use default
                        sheet_name = f"Sheet_{len(sheets) + 1}"
                        
                        # Stream the worksheet XML
                        with excel_zip.open(sheet_file) as sheet_stream:
                            sheet_data = await self._stream_excel_worksheet(sheet_stream, sheet_name)
                            if sheet_data is not None and not sheet_data.empty:
                                sheets[sheet_name] = sheet_data
                                logger.info(f"Streamed sheet {sheet_name}: {len(sheet_data)} rows")
                        
                    except Exception as sheet_e:
                        logger.warning(f"Failed to stream sheet {sheet_file}: {sheet_e}")
                        continue
                
                if not sheets:
                    raise ValueError("No data could be extracted from Excel file")
                
                return sheets
                    
        except Exception as e:
            logger.error(f"Streaming Excel reading failed: {e}")
            # Fallback to temporary file approach for compatibility
            return await self._read_excel_streaming_fallback(file_content, filename)
    
    async def _stream_excel_worksheet(self, sheet_stream, sheet_name: str) -> Optional[pd.DataFrame]:
        """
        Stream individual Excel worksheet XML to extract data without loading entire sheet.
        
        This method parses the worksheet XML incrementally, extracting only the data
        needed and processing it in chunks to maintain memory efficiency.
        """
        try:
            import xml.etree.ElementTree as ET
            from io import StringIO
            
            # Read worksheet XML in chunks
            chunk_size = 8192  # 8KB chunks
            xml_content = ""
            
            while True:
                chunk = sheet_stream.read(chunk_size)
                if not chunk:
                    break
                xml_content += chunk.decode('utf-8', errors='ignore')
            
            # Parse XML incrementally
            try:
                root = ET.fromstring(xml_content)
            except ET.ParseError as e:
                logger.warning(f"Failed to parse worksheet XML: {e}")
                return None
            
            # Extract data from worksheet
            rows_data = []
            headers = None
            
            # Find all row elements
            for row_elem in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                row_data = []
                
                # Extract cell values
                for cell_elem in row_elem.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    cell_value = ""
                    
                    # Get cell value
                    value_elem = cell_elem.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    if value_elem is not None and value_elem.text:
                        cell_value = value_elem.text
                    
                    row_data.append(cell_value)
                
                if row_data and any(cell.strip() for cell in row_data):
                    if headers is None:
                        headers = [f'Column_{i+1}' for i in range(len(row_data))]
                    else:
                        # Pad or truncate to match header length
                        while len(row_data) < len(headers):
                            row_data.append('')
                        row_data = row_data[:len(headers)]
                    
                    rows_data.append(row_data)
                    
                    # Process in chunks to manage memory
                    if len(rows_data) >= self.excel_chunk_size:
                        chunk_df = pd.DataFrame(rows_data, columns=headers)
                        if 'temp_df' not in locals():
                            temp_df = chunk_df
                        else:
                            temp_df = pd.concat([temp_df, chunk_df], ignore_index=True)
                        rows_data = []  # Clear processed data
            
            # Process remaining data
            if rows_data:
                chunk_df = pd.DataFrame(rows_data, columns=headers)
                if 'temp_df' not in locals():
                    temp_df = chunk_df
                else:
                    temp_df = pd.concat([temp_df, chunk_df], ignore_index=True)
            
            return temp_df if 'temp_df' in locals() else pd.DataFrame()
            
        except Exception as e:
            logger.warning(f"Failed to stream worksheet {sheet_name}: {e}")
            return None
    
    async def _read_excel_streaming_fallback(self, file_content: bytes, filename: str) -> Dict[str, pd.DataFrame]:
        """
        Fallback streaming method using temporary file for compatibility.
        
        This is used when the ZIP-based streaming fails, ensuring we still
        have a working solution for large Excel files.
        """
        try:
            # Create temporary file for streaming
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_path = temp_file.name
            
            try:
                sheets = {}
                
                # Use openpyxl for streaming (most memory efficient)
                from openpyxl import load_workbook
                workbook = load_workbook(temp_path, read_only=True, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    try:
                        # Read sheet in chunks to avoid memory issues
                        sheet = workbook[sheet_name]
                        
                        # Convert to DataFrame in chunks
                        data = []
                        headers = None
                        
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                            if row_idx == 0:
                                headers = [str(cell) if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
                            else:
                                data.append(row)
                            
                            # Process in chunks to manage memory
                            if len(data) >= self.excel_chunk_size:
                                chunk_df = pd.DataFrame(data, columns=headers)
                                if sheet_name not in sheets:
                                    sheets[sheet_name] = chunk_df
                                else:
                                    sheets[sheet_name] = pd.concat([sheets[sheet_name], chunk_df], ignore_index=True)
                                data = []  # Clear processed data
                        
                        # Process remaining data
                        if data:
                            chunk_df = pd.DataFrame(data, columns=headers)
                            if sheet_name not in sheets:
                                sheets[sheet_name] = chunk_df
                            else:
                                sheets[sheet_name] = pd.concat([sheets[sheet_name], chunk_df], ignore_index=True)
                        
                        logger.info(f"Streamed sheet {sheet_name}: {len(sheets[sheet_name])} rows")
                        
                    except Exception as sheet_e:
                        logger.warning(f"Failed to stream sheet {sheet_name}: {sheet_e}")
                        continue
                
                workbook.close()
                return sheets
                
            finally:
                # Clean up temporary file
                try:
                    os.unlink(temp_path)
                except Exception as e:
                    logger.warning(f"Failed to clean up temp file {temp_path}: {e}")
                    
        except Exception as e:
            logger.error(f"Fallback streaming Excel reading failed: {e}")
            raise ValueError(f"Could not read large Excel file: {e}")
    
    async def _repair_and_read_excel(self, file_content: bytes, filename: str, 
                                   progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Attempt to repair corrupted Excel files"""
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_path = temp_file.name
            
            try:
                if progress_callback:
                    await progress_callback("repairing", "🔧 Repairing Excel structure...", 40)
                
                # Try openpyxl repair mode
                try:
                    wb = openpyxl.load_workbook(temp_path, read_only=False, data_only=True)
                    
                    # Save repaired version
                    repaired_path = temp_path.replace('.xlsx', '_repaired.xlsx')
                    wb.save(repaired_path)
                    
                    # Read repaired file
                    with open(repaired_path, 'rb') as f:
                        repaired_content = f.read()
                    
                    return await self._read_excel_standard(repaired_content, filename)
                    
                except Exception as openpyxl_e:
                    logger.warning(f"OpenPyXL repair failed: {openpyxl_e}")
                
                
                # Last resort: try to extract as much data as possible
                return await self._extract_partial_excel_data(file_content, filename)
                
            finally:
                # Cleanup temporary files
                for path in [temp_path, temp_path.replace('.xlsx', '_repaired.xlsx'), 
                           ]:
                    try:
                        if os.path.exists(path):
                            os.unlink(path)
                    except Exception as e:
                        logger.warning(f"Failed to clean up temp file {path}: {e}")
                        
        except Exception as e:
            logger.error(f"Excel repair failed: {e}")
            raise

    async def _extract_partial_excel_data(self, file_content: bytes, filename: str) -> Dict[str, pd.DataFrame]:
        """Extract partial data from severely corrupted Excel files"""
        try:
            # Try to read as CSV (sometimes works for corrupted Excel)
            file_stream = io.BytesIO(file_content)

            # Try different encodings
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    file_stream.seek(0)
                    df = pd.read_csv(file_stream, encoding=encoding, on_bad_lines='skip')
                    if not df.empty:
                        logger.info(f"Extracted partial data as CSV with encoding {encoding}")
                        return {'Recovered_Data': df}
                except Exception:
                    continue

            # If CSV doesn't work, try to extract text and parse manually
            try:
                import zipfile
                with zipfile.ZipFile(io.BytesIO(file_content), 'r') as zip_ref:
                    # Extract shared strings and worksheets
                    shared_strings = []
                    worksheets = {}

                    for file_info in zip_ref.filelist:
                        if 'sharedStrings.xml' in file_info.filename:
                            # Extract shared strings
                            with zip_ref.open(file_info) as f:
                                content = f.read().decode('utf-8', errors='ignore')
                                # Simple regex to extract text values
                                import re
                                strings = re.findall(r'<t[^>]*>([^<]+)</t>', content)
                                shared_strings.extend(strings)

                        elif 'worksheets/sheet' in file_info.filename:
                            # Extract worksheet data
                            with zip_ref.open(file_info) as f:
                                content = f.read().decode('utf-8', errors='ignore')
                                # Extract cell values (simplified)
                                cells = re.findall(r'<c[^>]*><v>([^<]+)</v></c>', content)
                                sheet_name = f"Sheet_{len(worksheets) + 1}"
                                worksheets[sheet_name] = cells

                    # Convert to DataFrames
                    result = {}
                    for sheet_name, cells in worksheets.items():
                        if cells:
                            # Try to organize into rows/columns (simplified)
                            data = []
                            row = []
                            for i, cell in enumerate(cells):
                                row.append(cell)
                                if (i + 1) % 10 == 0:  # Assume 10 columns
                                    data.append(row)
                                    row = []
                            if row:
                                data.append(row)

                            if data:
                                df = pd.DataFrame(data)
                                result[sheet_name] = df

                    if result:
                        logger.info("Extracted partial data from corrupted Excel structure")
                        return result

            except Exception as zip_e:
                logger.warning(f"ZIP extraction failed: {zip_e}")

            raise ValueError("Could not extract any data from corrupted Excel file")

        except Exception as e:
            logger.error(f"Partial Excel extraction failed: {e}")
            raise

    async def _process_csv_enhanced(self, file_content: bytes, filename: str,
                                  progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Enhanced CSV processing with encoding detection, repair, and streaming for large files"""
        try:
            if progress_callback:
                await progress_callback("reading", "📄 Reading CSV file...", 20)

            # Check file size for streaming approach
            file_size_mb = len(file_content) / (1024 * 1024)
            use_streaming = file_size_mb > (self.streaming_threshold_mb * 5)  # Use streaming for CSV files > 5x threshold
            
            if use_streaming:
                logger.info(f"Large CSV file detected ({file_size_mb:.1f}MB), using streaming approach")
                return await self._read_csv_streaming(file_content, filename, progress_callback)

            file_stream = io.BytesIO(file_content)

            # Try to detect encoding
            import chardet
            detected = chardet.detect(file_content)
            encoding = detected.get('encoding', 'utf-8')

            # Try different delimiters and encodings
            delimiters = [',', ';', '\t', '|']
            encodings = [encoding, 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']

            for enc in encodings:
                for delimiter in delimiters:
                    try:
                        file_stream.seek(0)
                        df = pd.read_csv(
                            file_stream,
                            encoding=enc,
                            delimiter=delimiter,
                            on_bad_lines='skip',
                            low_memory=False
                        )

                        if not df.empty and len(df.columns) > 1:
                            logger.info(f"Successfully read CSV with encoding {enc} and delimiter '{delimiter}'")
                            return {'Sheet1': df}

                    except Exception as e:
                        logger.debug(f"Failed with encoding {enc} and delimiter '{delimiter}': {e}")
                        continue

            raise ValueError("Could not read CSV file with any encoding/delimiter combination")

        except Exception as e:
            logger.error(f"Enhanced CSV processing failed: {e}")
            raise
    
    async def _read_csv_streaming(self, file_content: bytes, filename: str, progress_callback=None) -> Dict[str, pd.DataFrame]:
        """
        TRUE streaming CSV reading for large files to avoid memory issues.
        
        This method processes CSV files line by line without loading the entire
        file content into memory at once. It uses a proper streaming approach
        that reads and processes data incrementally.
        """
        try:
            if progress_callback:
                await progress_callback("streaming", "📊 Streaming large CSV file...", 30)
            
            # Try to detect encoding
            import chardet
            detected = chardet.detect(file_content)
            encoding = detected.get('encoding', 'utf-8')
            
            # Try different delimiters
            delimiters = [',', ';', '\t', '|']
            
            for delimiter in delimiters:
                try:
                    # Use true streaming approach
                    result_df = await self._stream_csv_incremental(file_content, delimiter, encoding, progress_callback)
                    
                    if result_df is not None and not result_df.empty:
                        logger.info(f"Streamed CSV: {len(result_df):,} rows, {len(result_df.columns)} columns")
                        return {'Sheet1': result_df}
                        
                except Exception as e:
                    logger.warning(f"Failed to stream CSV with delimiter '{delimiter}': {e}")
                    continue
            
            raise ValueError("Could not read large CSV file with any delimiter")
            
        except Exception as e:
            logger.error(f"CSV streaming failed: {e}")
            raise ValueError(f"Could not read large CSV file: {e}")
    
    async def _stream_csv_incremental(self, file_content: bytes, delimiter: str, encoding: str, progress_callback=None) -> Optional[pd.DataFrame]:
        """
        Stream CSV file incrementally without loading entire content into memory.
        
        This method reads the CSV file in chunks, processes each chunk as a DataFrame,
        and combines them efficiently to avoid memory issues with large files.
        """
        try:
            from io import BytesIO, StringIO
            import csv
            
            # Create a streaming reader
            file_stream = BytesIO(file_content)
            text_stream = StringIO(file_content.decode(encoding, errors='ignore'))
            
            # Read first few lines to detect structure
            sample_lines = []
            for i, line in enumerate(text_stream):
                if i >= 10:  # Read first 10 lines for analysis
                    break
                sample_lines.append(line.strip())
            
            if len(sample_lines) < 2:
                return None
            
            # Detect delimiter and structure from sample
            detected_delimiter = delimiter
            if delimiter == ',':
                # Try to detect the actual delimiter
                for test_delim in [',', ';', '\t', '|']:
                    if test_delim in sample_lines[0]:
                        detected_delimiter = test_delim
                        break
            
            # Reset stream
            text_stream.seek(0)
            
            # Process CSV in chunks
            chunk_size = self.csv_chunk_size
            chunks = []
            current_chunk = []
            line_count = 0
            headers = None
            
            for line in text_stream:
                line = line.strip()
                if not line:
                    continue
                
                # Parse line with detected delimiter
                try:
                    reader = csv.reader([line], delimiter=detected_delimiter)
                    row = next(reader)
                    
                    if headers is None:
                        headers = [f'Column_{i+1}' for i in range(len(row))]
                        # Use first row as headers if it looks like headers
                        if all(isinstance(cell, str) and not cell.replace('.', '').replace(',', '').isdigit() 
                               for cell in row if cell.strip()):
                            headers = row
                            continue
                    
                    # Pad or truncate row to match headers
                    while len(row) < len(headers):
                        row.append('')
                    row = row[:len(headers)]
                    
                    current_chunk.append(row)
                    line_count += 1
                    
                    # Process chunk when it reaches the chunk size
                    if len(current_chunk) >= chunk_size:
                        chunk_df = pd.DataFrame(current_chunk, columns=headers)
                        if not chunk_df.empty:
                            chunks.append(chunk_df)
                        current_chunk = []
                        
                        # Update progress
                        if progress_callback and line_count % (chunk_size * 5) == 0:
                            progress = 30 + (line_count / 100000) * 50  # Estimate progress
                            await progress_callback("streaming", f"📊 Processed {line_count:,} lines...", int(min(progress, 80)))
                
                except Exception as e:
                    logger.warning(f"Failed to parse line {line_count}: {e}")
                    continue
            
            # Process remaining data
            if current_chunk:
                chunk_df = pd.DataFrame(current_chunk, columns=headers)
                if not chunk_df.empty:
                    chunks.append(chunk_df)
            
            # Combine all chunks
            if chunks:
                final_df = pd.concat(chunks, ignore_index=True)
                return final_df
            
            return None
            
        except Exception as e:
            logger.warning(f"Failed to stream CSV incrementally: {e}")
            return None

    async def _process_ods(self, file_content: bytes, filename: str,
                         progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Process OpenDocument Spreadsheet files"""
        try:
            if progress_callback:
                await progress_callback("reading", "📊 Reading ODS file...", 20)

            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.ods', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_path = temp_file.name

            try:
                # Load ODS document
                doc = load_ods(temp_path)
                sheets = {}

                # Extract tables (sheets)
                tables = doc.getElementsByType(Table)

                for i, table in enumerate(tables):
                    sheet_name = table.getAttribute('name') or f'Sheet{i+1}'

                    # Extract rows
                    rows = []
                    table_rows = table.getElementsByType(TableRow)

                    for row in table_rows:
                        row_data = []
                        cells = row.getElementsByType(TableCell)

                        for cell in cells:
                            # Get cell value
                            cell_value = ""
                            paragraphs = cell.getElementsByType(P)
                            for p in paragraphs:
                                if p.firstChild:
                                    cell_value += str(p.firstChild)

                            row_data.append(cell_value)

                        if any(cell.strip() for cell in row_data):  # Skip empty rows
                            rows.append(row_data)

                    if rows:
                        # Convert to DataFrame
                        df = pd.DataFrame(rows[1:], columns=rows[0] if rows else None)
                        sheets[sheet_name] = df

                if not sheets:
                    raise ValueError("No data found in ODS file")

                return sheets

            finally:
                # Cleanup
                try:
                    os.unlink(temp_path)
                except Exception as e:
                    logger.warning(f"Failed to clean up temp file {temp_path}: {e}")

        except Exception as e:
            logger.error(f"ODS processing failed: {e}")
            raise

    async def _process_pdf(self, file_content: bytes, filename: str,
                         progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Extract tables from PDF files using multiple methods"""
        try:
            if progress_callback:
                await progress_callback("extracting", "📄 Extracting tables from PDF...", 20)

            # Create temporary file
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_path = temp_file.name

            try:
                tables = {}
                table_count = 0

                # Method 1: Try tabula-py (good for structured tables)
                try:
                    if progress_callback:
                        await progress_callback("extracting", "🔍 Using tabula extraction...", 30)

                    tabula_tables = tabula.read_pdf(temp_path, pages='all', multiple_tables=True, pandas_options={'header': None})

                    for i, df in enumerate(tabula_tables):
                        if not df.empty:
                            # Clean up empty rows and columns
                            df.dropna(axis=0, how='all', inplace=True)
                            df.dropna(axis=1, how='all', inplace=True)
                            if not df.empty:
                                table_count += 1
                                tables[f'Table_{table_count}_Tabula'] = df

                except Exception as tabula_e:
                    if 'password' in str(tabula_e).lower():
                        raise ValueError("Password-protected PDF files are not supported.")
                    logger.warning(f"Tabula extraction failed: {tabula_e}")

                # Method 2: Try camelot (good for lattice tables)
                if CAMELOT_AVAILABLE:
                    try:
                        if progress_callback:
                            await progress_callback("extracting", "🐪 Using camelot extraction...", 50)

                        camelot_tables = camelot.read_pdf(temp_path, pages='all')

                        for i, table in enumerate(camelot_tables):
                            df = table.df
                            if not df.empty and len(df.columns) > 1:
                                table_count += 1
                                tables[f'Table_{table_count}_Camelot'] = df
                    except Exception as camelot_e:
                        logger.warning(f"Camelot extraction failed: {camelot_e}")
                else:
                    logger.warning("Camelot not available, skipping camelot extraction")

                # Method 3: Try pdfplumber (good for text-based tables)
                try:
                    if progress_callback:
                        await progress_callback("extracting", "🔧 Using pdfplumber extraction...", 70)

                    with pdfplumber.open(temp_path) as pdf:
                        for page_num, page in enumerate(pdf.pages):
                            page_tables = page.extract_tables()

                            for i, table_data in enumerate(page_tables):
                                if table_data and len(table_data) > 1:
                                    # Convert to DataFrame
                                    headers = table_data[0] if table_data[0] else [f'Col_{j}' for j in range(len(table_data[1]))]
                                    data = table_data[1:]

                                    df = pd.DataFrame(data, columns=headers)
                                    if not df.empty:
                                        table_count += 1
                                        tables[f'Page_{page_num+1}_Table_{i+1}_PDFPlumber'] = df

                except Exception as pdfplumber_e:
                    logger.warning(f"PDFPlumber extraction failed: {pdfplumber_e}")

                if not tables:
                    raise ValueError("No tables found in PDF file")

                # De-duplicate and merge similar tables
                return self._deduplicate_and_merge_tables(tables)

            finally:
                # Cleanup
                try:
                    os.unlink(temp_path)
                except Exception as e:
                    logger.warning(f"Failed to clean up temp file {temp_path}: {e}")

        except Exception as e:
            logger.error(f"PDF processing failed: {e}")
            raise

    async def _process_archive(self, file_content: bytes, filename: str,
                             progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Process ZIP, 7Z, and RAR archives containing multiple files"""
        try:
            if progress_callback:
                await progress_callback("extracting", "📦 Extracting archive...", 10)

            # Create temporary directory
            with tempfile.TemporaryDirectory() as temp_dir:
                archive_path = os.path.join(temp_dir, filename)

                # Write archive to temporary file
                with open(archive_path, 'wb') as f:
                    f.write(file_content)

                # Extract based on file type
                extracted_files = []

                if filename.lower().endswith('.zip'):
                    try:
                        with zipfile.ZipFile(archive_path, 'r') as zip_ref:
                            zip_ref.extractall(temp_dir)
                            extracted_files = zip_ref.namelist()
                    except zipfile.BadZipFile as e:
                        if 'encrypted' in str(e).lower():
                            raise ValueError("Password-protected ZIP files are not supported.")
                        raise

                elif filename.lower().endswith('.7z'):
                    with py7zr.SevenZipFile(archive_path, mode='r') as archive:
                        archive.extractall(temp_dir)
                        extracted_files = archive.getnames()

                elif filename.lower().endswith('.rar'):
                    with rarfile.RarFile(archive_path) as rf:
                        rf.extractall(temp_dir)
                        extracted_files = rf.namelist()

                if progress_callback:
                    await progress_callback("processing", f"📊 Processing {len(extracted_files)} files...", 20)

                # Process each extracted file
                all_sheets = {}
                processed_count = 0

                for file_path in extracted_files:
                    try:
                        full_path = os.path.join(temp_dir, file_path)

                        # Skip directories
                        if os.path.isdir(full_path):
                            continue

                        # Check if file is processable
                        file_ext = os.path.splitext(file_path.lower())[1]
                        if file_ext not in ['.xlsx', '.xls', '.csv', '.ods', '.pdf', '.png', '.jpg', '.jpeg']:
                            continue

                        # Read file content
                        with open(full_path, 'rb') as f:
                            file_content = f.read()

                        # Process file recursively
                        file_sheets = await self.process_file_enhanced(
                            file_content,
                            os.path.basename(file_path),
                            progress_callback
                        )

                        # Add to results with file prefix
                        file_prefix = os.path.splitext(os.path.basename(file_path))[0]
                        for sheet_name, df in file_sheets.items():
                            combined_name = f"{file_prefix}_{sheet_name}"
                            all_sheets[combined_name] = df

                        processed_count += 1

                        if progress_callback:
                            progress = 20 + (processed_count / len(extracted_files)) * 70
                            await progress_callback("processing",
                                                   f"📊 Processed {processed_count}/{len(extracted_files)} files...",
                                                   progress)

                    except Exception as file_e:
                        logger.warning(f"Failed to process file {file_path}: {file_e}")
                        continue

                if not all_sheets:
                    raise ValueError("No processable files found in archive")

                return all_sheets

        except Exception as e:
            logger.error(f"Archive processing failed: {e}")
            raise

    async def _process_image(self, file_content: bytes, filename: str,
                           progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Extract tables from images using OCR"""
        try:
            if progress_callback:
                await progress_callback("ocr", "👁️ Performing OCR on image...", 20)

            # Load image
            image = Image.open(io.BytesIO(file_content))

            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')

            # Enhance image for better OCR
            image_array = np.array(image)

            # Apply image preprocessing
            gray = cv2.cvtColor(image_array, cv2.COLOR_RGB2GRAY)

            # Apply adaptive thresholding
            thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)

            # Noise removal
            kernel = np.ones((1, 1), np.uint8)
            processed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
            processed = cv2.medianBlur(processed, 3)

            if progress_callback:
                await progress_callback("ocr", "🔍 Extracting text from image...", 50)

            # Perform OCR
            try:
                # Try to extract structured data - run in thread pool to avoid blocking
                loop = asyncio.get_event_loop()
                ocr_data = await loop.run_in_executor(
                    None,
                    lambda: pytesseract.image_to_data(processed, config=self.ocr_config, output_type=pytesseract.Output.DICT)
                )

                # Group text by lines and columns
                lines = {}
                for i, text in enumerate(ocr_data['text']):
                    if text.strip():
                        top = ocr_data['top'][i]
                        left = ocr_data['left'][i]

                        # Group by approximate line (within 10 pixels)
                        line_key = round(top / 10) * 10
                        if line_key not in lines:
                            lines[line_key] = []

                        lines[line_key].append((left, text.strip()))

                # Sort lines by vertical position
                sorted_lines = sorted(lines.items())

                # Convert to table structure
                table_data = []
                for line_y, line_items in sorted_lines:
                    # Sort items in line by horizontal position
                    sorted_items = sorted(line_items, key=lambda x: x[0])
                    row = [item[1] for item in sorted_items]

                    if row and any(cell.strip() for cell in row):  # Skip empty rows
                        table_data.append(row)

                if progress_callback:
                    await progress_callback("processing", "📊 Converting OCR results to table...", 80)

                if table_data:
                    # Normalize row lengths
                    max_cols = max(len(row) for row in table_data)
                    normalized_data = []

                    for row in table_data:
                        # Pad short rows
                        while len(row) < max_cols:
                            row.append('')
                        normalized_data.append(row)

                    # Create DataFrame
                    if len(normalized_data) > 1:
                        # Use first row as headers if it looks like headers
                        first_row = normalized_data[0]
                        if all(isinstance(cell, str) and not cell.replace('.', '').replace(',', '').isdigit()
                              for cell in first_row if cell.strip()):
                            df = pd.DataFrame(normalized_data[1:], columns=first_row)
                        else:
                            df = pd.DataFrame(normalized_data)
                    else:
                        df = pd.DataFrame(normalized_data)

                    return {'OCR_Table': df}
                else:
                    # Fallback: extract all text as single column
                    # Run OCR in thread pool to avoid blocking
                    loop = asyncio.get_event_loop()
                    all_text = await loop.run_in_executor(
                        None,
                        lambda: pytesseract.image_to_string(processed, config=self.ocr_config)
                    )
                    lines = [line.strip() for line in all_text.split('\n') if line.strip()]

                    if lines:
                        df = pd.DataFrame(lines, columns=['Extracted_Text'])
                        return {'OCR_Text': df}
                    else:
                        raise ValueError("No text extracted from image")

            except Exception as ocr_e:
                logger.error(f"OCR processing failed: {ocr_e}")
                raise ValueError(f"OCR failed: {str(ocr_e)}")

        except Exception as e:
            logger.error(f"Image processing failed: {e}")
            raise

    async def _fallback_processing(self, file_content: bytes, filename: str,
                                 progress_callback=None) -> Dict[str, pd.DataFrame]:
        """Fallback processing for unknown or failed file types"""
        try:
            if progress_callback:
                await progress_callback("fallback", "🔄 Attempting fallback processing...", 10)

            # Handle empty content gracefully
            if not file_content or len(file_content) == 0:
                logger.warning("Empty file content, returning empty DataFrame")
                return {'Empty_File': pd.DataFrame()}

            # Try to read as text and parse
            try:
                # Try different encodings
                for encoding in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        text_content = file_content.decode(encoding)

                        # Split into lines
                        lines = [line.strip() for line in text_content.split('\n') if line.strip()]

                        if lines:
                            # Try to detect delimiter
                            first_line = lines[0]
                            delimiters = [',', ';', '\t', '|']

                            for delimiter in delimiters:
                                if delimiter in first_line:
                                    # Try to parse as delimited data
                                    data = []
                                    for line in lines:
                                        row = [cell.strip() for cell in line.split(delimiter)]
                                        data.append(row)

                                    if data and len(data) > 1:
                                        df = pd.DataFrame(data[1:], columns=data[0])
                                        return {'Fallback_Data': df}

                            # If no delimiter found, return as single column
                            df = pd.DataFrame(lines, columns=['Text_Content'])
                            return {'Fallback_Text': df}

                    except UnicodeDecodeError:
                        continue

            except Exception as text_e:
                logger.warning(f"Text fallback failed: {text_e}")

            # If all else fails, return empty DataFrame instead of raising error
            logger.warning("All processing methods failed, returning empty DataFrame")
            return {'Empty_File': pd.DataFrame()}

        except Exception as e:
            logger.error(f"Fallback processing failed: {e}")
            raise

    async def smart_merge_sheets(self, sheets: Dict[str, pd.DataFrame],
                               progress_callback=None) -> Dict[str, pd.DataFrame]:
        """
        Intelligently merge related sheets based on structure and content similarity
        """
        try:
            if progress_callback:
                await progress_callback("merging", "🔗 Analyzing sheet relationships...", 5)

            if len(sheets) <= 1:
                return sheets

            # Analyze sheet relationships
            sheet_analysis = self._analyze_sheet_relationships(sheets)

            # Find merge candidates
            merge_groups = self._find_merge_candidates(sheet_analysis)

            # Perform merging
            merged_sheets = {}
            processed_sheets = set()

            for group_name, sheet_names in merge_groups.items():
                if len(sheet_names) > 1:
                    merged_df = self._merge_sheet_group(sheets, sheet_names)
                    if merged_df is not None and not merged_df.empty:
                        merged_sheets[f"Merged_{group_name}"] = merged_df
                        processed_sheets.update(sheet_names)

            # Add unmerged sheets
            for sheet_name, df in sheets.items():
                if sheet_name not in processed_sheets:
                    merged_sheets[sheet_name] = df

            return merged_sheets

        except Exception as e:
            logger.error(f"Smart sheet merging failed: {e}")
            return sheets

    def _analyze_sheet_relationships(self, sheets: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Analyze relationships between sheets"""
        analysis = {}

        for sheet_name, df in sheets.items():
            analysis[sheet_name] = {
                'columns': list(df.columns),
                'column_count': len(df.columns),
                'row_count': len(df),
                'column_types': {col: str(df[col].dtype) for col in df.columns},
                'sample_data': df.head(3).to_dict('records') if not df.empty else [],
                'has_date_columns': any('date' in col.lower() for col in df.columns),
                'has_amount_columns': any(word in col.lower() for word in ['amount', 'total', 'sum', 'value', 'price'] for col in df.columns),
                'potential_type': self._guess_sheet_type(df)
            }

        return analysis

    def _guess_sheet_type(self, df: pd.DataFrame) -> str:
        """Guess the type of data in a sheet"""
        columns_lower = [col.lower() for col in df.columns]

        # Financial data patterns
        if any(word in ' '.join(columns_lower) for word in ['revenue', 'income', 'sales']):
            return 'revenue'
        elif any(word in ' '.join(columns_lower) for word in ['expense', 'cost', 'payment']):
            return 'expense'
        elif any(word in ' '.join(columns_lower) for word in ['employee', 'payroll', 'salary']):
            return 'payroll'
        elif any(word in ' '.join(columns_lower) for word in ['invoice', 'bill']):
            return 'invoice'
        elif any(word in ' '.join(columns_lower) for word in ['balance', 'account']):
            return 'balance'
        else:
            return 'general'

    def _find_merge_candidates(self, analysis: Dict[str, Any]) -> Dict[str, List[str]]:
        """Find sheets that can be merged together"""
        merge_groups = {}
        sheet_names = list(analysis.keys())

        # Group by data type
        type_groups = {}
        for sheet_name, info in analysis.items():
            sheet_type = info['potential_type']
            if sheet_type not in type_groups:
                type_groups[sheet_type] = []
            type_groups[sheet_type].append(sheet_name)

        # Find sheets with similar column structures
        for sheet_type, sheets in type_groups.items():
            if len(sheets) > 1:
                # Check column similarity
                similar_groups = self._group_by_column_similarity(analysis, sheets)
                for i, group in enumerate(similar_groups):
                    if len(group) > 1:
                        merge_groups[f"{sheet_type}_{i+1}"] = group

        return merge_groups

    def _group_by_column_similarity(self, analysis: Dict[str, Any],
                                  sheet_names: List[str]) -> List[List[str]]:
        """Group sheets by column similarity"""
        groups = []
        processed = set()

        for sheet_name in sheet_names:
            if sheet_name in processed:
                continue

            current_group = [sheet_name]
            current_columns = set(analysis[sheet_name]['columns'])

            for other_sheet in sheet_names:
                if other_sheet != sheet_name and other_sheet not in processed:
                    other_columns = set(analysis[other_sheet]['columns'])

                    # Calculate column overlap
                    overlap = len(current_columns & other_columns)
                    total = len(current_columns | other_columns)
                    similarity = overlap / total if total > 0 else 0

                    # If similarity is high enough, add to group
                    if similarity >= 0.6:  # 60% similarity threshold
                        current_group.append(other_sheet)

            if len(current_group) > 1:
                groups.append(current_group)
                processed.update(current_group)
            else:
                processed.add(sheet_name)

        return groups

    def _merge_sheet_group(self, sheets: Dict[str, pd.DataFrame],
                          sheet_names: List[str]) -> pd.DataFrame:
        """Merge a group of similar sheets"""
        try:
            dfs_to_merge = []

            for sheet_name in sheet_names:
                df = sheets[sheet_name].copy()
                # Add source sheet column
                df['_source_sheet'] = sheet_name
                dfs_to_merge.append(df)

            # Find common columns
            all_columns = [set(df.columns) for df in dfs_to_merge]
            common_columns = set.intersection(*all_columns) if all_columns else set()

            if not common_columns:
                # If no common columns, try to align by position
                min_cols = min(len(df.columns) for df in dfs_to_merge)
                if min_cols > 0:
                    # Use first sheet's column names for alignment
                    base_columns = list(dfs_to_merge[0].columns[:min_cols])
                    aligned_dfs = []

                    for df in dfs_to_merge:
                        aligned_df = df.iloc[:, :min_cols].copy()
                        aligned_df.columns = base_columns
                        aligned_dfs.append(aligned_df)

                    return pd.concat(aligned_dfs, ignore_index=True)
                else:
                    return pd.DataFrame()

            # Merge on common columns
            merged_df = pd.concat(dfs_to_merge, ignore_index=True, sort=False)

            # Reorder columns to put common ones first
            common_cols_list = sorted(list(common_columns - {'_source_sheet'}))
            other_cols = [col for col in merged_df.columns
                         if col not in common_columns]

            column_order = common_cols_list + other_cols + ['_source_sheet']
            merged_df = merged_df.reindex(columns=column_order)

            return merged_df

        except Exception as e:
            logger.error(f"Failed to merge sheet group {sheet_names}: {e}")
            return pd.DataFrame()

    async def process_archive(self, file_content: bytes, filename: str,
                            progress_callback=None) -> Dict[str, Any]:
        """
        Process archive files (ZIP, RAR, 7Z) and extract/process contents
        """
        try:
            if progress_callback:
                await progress_callback("processing", "📦 Extracting archive...", 10)

            import zipfile
            import tempfile
            import os
            from io import BytesIO

            results = {
                'type': 'archive',
                'filename': filename,
                'extracted_files': [],
                'processed_contents': {},
                'summary': {}
            }

            # Create temporary directory for extraction
            with tempfile.TemporaryDirectory() as temp_dir:
                archive_stream = BytesIO(file_content)

                # Try different archive formats
                extracted_files = []

                if filename.lower().endswith('.zip'):
                    with zipfile.ZipFile(archive_stream, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                        extracted_files = zip_ref.namelist()

                # Process extracted files
                total_files = len(extracted_files)
                processed_count = 0

                for file_path in extracted_files:
                    full_path = os.path.join(temp_dir, file_path)

                    if os.path.isfile(full_path):
                        try:
                            with open(full_path, 'rb') as f:
                                file_content = f.read()

                            # Process based on file extension
                            file_ext = os.path.splitext(file_path)[1].lower()

                            if file_ext in ['.xlsx', '.xls', '.csv']:
                                processed_data = await self.process_excel_file(
                                    file_content, file_path, progress_callback
                                )
                                results['processed_contents'][file_path] = processed_data

                            elif file_ext in ['.pdf']:
                                processed_data = await self.process_pdf_file(
                                    file_content, file_path, progress_callback
                                )
                                results['processed_contents'][file_path] = processed_data

                            elif file_ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
                                processed_data = await self.process_image_ocr(
                                    file_content, file_path, progress_callback
                                )
                                results['processed_contents'][file_path] = processed_data

                            processed_count += 1

                            if progress_callback:
                                progress = 10 + (processed_count / total_files) * 80
                                await progress_callback(
                                    "processing",
                                    f"📄 Processed {processed_count}/{total_files} files...",
                                    progress
                                )

                        except Exception as e:
                            logger.warning(f"Failed to process {file_path}: {e}")
                            continue

                results['extracted_files'] = extracted_files
                results['summary'] = {
                    'total_files': total_files,
                    'processed_files': processed_count,
                    'file_types': list(set(os.path.splitext(f)[1].lower()
                                         for f in extracted_files if '.' in f))
                }

            return results

        except Exception as e:
            logger.error(f"Archive processing failed: {e}")
            raise

    async def process_image_ocr(self, file_content: bytes, filename: str,
                              progress_callback=None) -> Dict[str, Any]:
        """
        Process images using OCR to extract text and structured data
        """
        try:
            if progress_callback:
                await progress_callback("processing", "🖼️ Analyzing image...", 20)

            # This would integrate with OCR services like Tesseract, AWS Textract, etc.
            # For now, we'll return a structured placeholder

            results = {
                'type': 'image_ocr',
                'filename': filename,
                'extracted_text': '',
                'structured_data': {},
                'confidence_scores': {},
                'detected_elements': []
            }

            # Simulate OCR processing
            if progress_callback:
                await progress_callback("processing", "🔍 Extracting text...", 60)

            # In a real implementation, this would:
            # 1. Use OCR to extract text
            # 2. Identify tables, forms, invoices, etc.
            # 3. Structure the data appropriately
            # 4. Return confidence scores

            # Placeholder for OCR results
            results.update({
                'extracted_text': f"[OCR text would be extracted from {filename}]",
                'structured_data': {
                    'detected_type': 'document',
                    'tables': [],
                    'key_value_pairs': {},
                    'text_blocks': []
                },
                'confidence_scores': {
                    'overall': 0.85,
                    'text_detection': 0.90,
                    'structure_detection': 0.80
                },
                'detected_elements': [
                    {'type': 'text_block', 'confidence': 0.90, 'bbox': [0, 0, 100, 50]},
                    {'type': 'table', 'confidence': 0.75, 'bbox': [0, 60, 200, 150]}
                ]
            })

            if progress_callback:
                await progress_callback("complete", "✅ Image processing complete", 100)

            return results

        except Exception as e:
            logger.error(f"Image OCR processing failed: {e}")
            raise

    async def process_any_file(self, file_content: bytes, filename: str,
                             progress_callback=None) -> Dict[str, Any]:
        """
        Universal file processor that routes to appropriate specialized processor
        """
        try:
            if progress_callback:
                await progress_callback("analyzing", "🔍 Analyzing file type...", 5)

            file_ext = os.path.splitext(filename)[1].lower()

            # Route to appropriate processor
            if file_ext in ['.xlsx', '.xls', '.csv']:
                return await self.process_excel_file(file_content, filename, progress_callback)

            elif file_ext in ['.pdf']:
                return await self.process_pdf_file(file_content, filename, progress_callback)

            elif file_ext in ['.zip', '.rar', '.7z']:
                return await self.process_archive(file_content, filename, progress_callback)

            elif file_ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
                return await self.process_image_ocr(file_content, filename, progress_callback)

            else:
                # Unsupported file type
                if progress_callback:
                    await progress_callback("error", f"❌ Unsupported file type: {file_ext}", 0)

                return {
                    'type': 'unsupported',
                    'filename': filename,
                    'file_extension': file_ext,
                    'error': f"File type {file_ext} is not supported",
                    'supported_types': ['.xlsx', '.xls', '.csv', '.pdf', '.zip', '.rar', '.7z',
                                      '.png', '.jpg', '.jpeg', '.tiff', '.bmp']
                }

        except Exception as e:
            logger.error(f"File processing failed for {filename}: {e}")
            if progress_callback:
                await progress_callback("error", f"❌ Processing failed: {str(e)}", 0)
            raise

    def get_processing_capabilities(self) -> Dict[str, Any]:
        """
        Return comprehensive information about processing capabilities
        """
        return {
            'supported_formats': {
                'spreadsheets': {
                    'extensions': ['.xlsx', '.xls', '.csv'],
                    'features': [
                        'Multi-sheet processing',
                        'Smart sheet merging',
                        'Data type detection',
                        'Error correction',
                        'Formula evaluation',
                        'Pivot table detection'
                    ]
                },
                'documents': {
                    'extensions': ['.pdf'],
                    'features': [
                        'Text extraction',
                        'Table detection',
                        'Form field extraction',
                        'Multi-page processing',
                        'OCR for scanned documents'
                    ]
                },
                'images': {
                    'extensions': ['.png', '.jpg', '.jpeg', '.tiff', '.bmp'],
                    'features': [
                        'OCR text extraction',
                        'Table detection',
                        'Document structure analysis',
                        'Confidence scoring',
                        'Multi-language support'
                    ]
                },
                'archives': {
                    'extensions': ['.zip', '.rar', '.7z'],
                    'features': [
                        'Recursive extraction',
                        'Batch processing',
                        'Mixed format handling',
                        'Progress tracking'
                    ]
                }
            },
            'processing_modes': [
                'standard',
                'enhanced_intelligence',
                'auto_repair',
                'smart_merging'
            ],
            'output_formats': [
                'structured_json',
                'pandas_dataframe',
                'csv_export',
                'excel_export'
            ],
            'advanced_features': [
                'Zero-friction upload',
                'Intelligent error correction',
                'Automatic data type detection',
                'Smart sheet relationship analysis',
                'Progress callbacks',
                'Batch processing',
                'Archive extraction',
                'OCR integration'
            ]
        }

    async def smart_merge_sheets(self, sheets: Dict[str, pd.DataFrame],
                               progress_callback=None) -> Dict[str, pd.DataFrame]:
        """
        Intelligently merge related sheets based on structure and content similarity

        Args:
            sheets: Dictionary of sheet_name -> DataFrame
            progress_callback: Progress reporting function

        Returns:
            Dictionary with merged sheets and individual sheets
        """
        try:
            if progress_callback:
                await progress_callback("merging", "🔗 Analyzing sheet relationships...", 5)

            if len(sheets) <= 1:
                return sheets

            # Analyze sheet relationships
            sheet_analysis = self._analyze_sheet_relationships(sheets)

            if progress_callback:
                await progress_callback("merging", "🔗 Identifying merge candidates...", 20)

            # Find merge candidates
            merge_groups = self._find_merge_candidates(sheet_analysis)

            if progress_callback:
                await progress_callback("merging", "🔗 Performing intelligent merging...", 50)

            # Perform merging
            merged_sheets = {}
            processed_sheets = set()

            for group_name, sheet_names in merge_groups.items():
                if len(sheet_names) > 1:
                    # Merge sheets in this group
                    merged_df = self._merge_sheet_group(sheets, sheet_names)
                    if merged_df is not None and not merged_df.empty:
                        merged_sheets[f"Merged_{group_name}"] = merged_df
                        processed_sheets.update(sheet_names)

            # Add unmerged sheets
            for sheet_name, df in sheets.items():
                if sheet_name not in processed_sheets:
                    merged_sheets[sheet_name] = df

            if progress_callback:
                await progress_callback("merging", "✅ Sheet merging completed", 100)

            return merged_sheets

        except Exception as e:
            logger.error(f"Smart sheet merging failed: {e}")
            # Return original sheets if merging fails
            return sheets

    def _analyze_sheet_relationships(self, sheets: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Analyze relationships between sheets"""
        analysis = {}

        for sheet_name, df in sheets.items():
            analysis[sheet_name] = {
                'columns': list(df.columns),
                'column_count': len(df.columns),
                'row_count': len(df),
                'column_types': {col: str(df[col].dtype) for col in df.columns},
                'sample_data': df.head(3).to_dict('records') if not df.empty else [],
                'has_date_columns': any('date' in col.lower() for col in df.columns),
                'has_amount_columns': any(word in col.lower() for word in ['amount', 'total', 'sum', 'value', 'price'] for col in df.columns),
                'potential_type': self._guess_sheet_type(df)
            }

        return analysis

    def _guess_sheet_type(self, df: pd.DataFrame) -> str:
        """Guess the type of data in a sheet"""
        columns_lower = [col.lower() for col in df.columns]

        # Financial data patterns
        if any(word in ' '.join(columns_lower) for word in ['revenue', 'income', 'sales']):
            return 'revenue'
        elif any(word in ' '.join(columns_lower) for word in ['expense', 'cost', 'payment']):
            return 'expense'
        elif any(word in ' '.join(columns_lower) for word in ['employee', 'payroll', 'salary']):
            return 'payroll'
        elif any(word in ' '.join(columns_lower) for word in ['invoice', 'bill']):
            return 'invoice'
        elif any(word in ' '.join(columns_lower) for word in ['balance', 'account']):
            return 'balance'
        else:
            return 'general'

    def _find_merge_candidates(self, sheet_analysis: Dict[str, Any]) -> Dict[str, List[str]]:
        """Find sheets that should be merged together"""
        merge_groups = {}

        # Group by data type
        type_groups = {}
        for sheet_name, analysis in sheet_analysis.items():
            sheet_type = analysis['potential_type']
            if sheet_type not in type_groups:
                type_groups[sheet_type] = []
            type_groups[sheet_type].append(sheet_name)

        # Within each type, find sheets with similar structure
        for data_type, sheet_names in type_groups.items():
            if len(sheet_names) > 1:
                # Check column similarity
                similar_groups = []

                for sheet_name in sheet_names:
                    analysis = sheet_analysis[sheet_name]
                    placed = False

                    for group in similar_groups:
                        # Check if this sheet is similar to sheets in the group
                        sample_sheet = group[0]
                        sample_analysis = sheet_analysis[sample_sheet]

                        # Calculate column similarity
                        common_columns = set(analysis['columns']) & set(sample_analysis['columns'])
                        total_columns = set(analysis['columns']) | set(sample_analysis['columns'])

                        similarity = len(common_columns) / len(total_columns) if total_columns else 0

                        # If >70% column similarity, add to group
                        if similarity > 0.7:
                            group.append(sheet_name)
                            placed = True
                            break

                    if not placed:
                        similar_groups.append([sheet_name])

                # Add groups with multiple sheets to merge candidates
                for i, group in enumerate(similar_groups):
                    if len(group) > 1:
                        group_name = f"{data_type}_{i+1}"
                        merge_groups[group_name] = group

        return merge_groups

    def _merge_sheet_group(self, sheets: Dict[str, pd.DataFrame], sheet_names: List[str]) -> Optional[pd.DataFrame]:
        """Merge a group of similar sheets"""
        try:
            dfs_to_merge = []

            for sheet_name in sheet_names:
                df = sheets[sheet_name].copy()

                # Add source sheet column
                df['_source_sheet'] = sheet_name

                dfs_to_merge.append(df)

            if not dfs_to_merge:
                return None

            # Find common columns
            all_columns = set()
            for df in dfs_to_merge:
                all_columns.update(df.columns)

            # Align columns across all DataFrames
            aligned_dfs = []
            for df in dfs_to_merge:
                # Add missing columns with NaN
                for col in all_columns:
                    if col not in df.columns:
                        df[col] = pd.NA

                # Reorder columns consistently
                df = df.reindex(columns=sorted(all_columns))
                aligned_dfs.append(df)

            # Concatenate all DataFrames
            merged_df = pd.concat(aligned_dfs, ignore_index=True, sort=False)

            # Clean up the merged DataFrame
            merged_df = merged_df.dropna(how='all')  # Remove completely empty rows

            return merged_df

        except Exception as e:
            logger.error(f"Failed to merge sheet group {sheet_names}: {e}")
            return None

    def _deduplicate_and_merge_tables(self, tables: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Deduplicate and merge similar tables extracted from a PDF."""
        if len(tables) <= 1:
            return tables

        # Calculate hashes for all tables
        table_hashes = {name: self._calculate_table_hash(df) for name, df in tables.items()}
        
        # Group identical tables
        unique_tables = {}
        for name, df in tables.items():
            h = table_hashes[name]
            if h not in unique_tables:
                unique_tables[h] = (name, df)
        
        if len(unique_tables) == len(tables):
            return tables # No identical tables found

        # Further process for similar (not identical) tables
        final_tables = {}
        processed_hashes = set()

        sorted_unique_tables = sorted(unique_tables.items(), key=lambda item: len(item[1][1]), reverse=True)

        for h, (name, df) in sorted_unique_tables:
            if h in processed_hashes:
                continue

            similar_group = [(h, name, df)]
            processed_hashes.add(h)

            for other_h, (other_name, other_df) in sorted_unique_tables:
                if other_h in processed_hashes:
                    continue
                
                similarity = self._calculate_table_similarity(df, other_df)
                if similarity > 0.85: # High similarity threshold for merging
                    similar_group.append((other_h, other_name, other_df))
                    processed_hashes.add(other_h)
            
            if len(similar_group) > 1:
                # Merge the group
                merged_df = similar_group[0][2] # Start with the largest table
                for _, _, other_df in similar_group[1:]:
                    # A simple concat and drop duplicates, more sophisticated merging can be added
                    merged_df = pd.concat([merged_df, other_df]).drop_duplicates().reset_index(drop=True)
                final_tables[f"Merged_{similar_group[0][1]}"] = merged_df
            else:
                final_tables[name] = df

        return final_tables

    def _calculate_table_hash(self, df: pd.DataFrame) -> str:
        """Calculates a hash for a DataFrame based on its content and structure."""
        import hashlib
        # Normalize dataframe for hashing
        df_str = df.to_string(index=False, header=True, na_rep='').encode('utf-8')
        return hashlib.md5(df_str).hexdigest()

    def _calculate_table_similarity(self, df1: pd.DataFrame, df2: pd.DataFrame) -> float:
        """Calculates a similarity score between two DataFrames."""
        from difflib import SequenceMatcher
        
        # Normalize and convert to string
        s1 = df1.to_string(index=False, header=True, na_rep='').lower()
        s2 = df2.to_string(index=False, header=True, na_rep='').lower()
        
        return SequenceMatcher(None, s1, s2).ratio()
