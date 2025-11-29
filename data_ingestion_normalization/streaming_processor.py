"""
Streaming File Processor for Finley AI
======================================

Memory-efficient streaming processor that handles large files without loading
them entirely into memory. Prevents OOM crashes and enables processing of
files up to several GB in size.

Author: Principal Engineer
Version: 1.0.0
"""

import asyncio
import structlog
import tempfile
import os
import gc
from typing import Dict, List, Any, Optional, AsyncGenerator, Callable
from dataclasses import dataclass
import pandas as pd
import psutil
from aiolimiter import AsyncLimiter
import io

logger = structlog.get_logger(__name__)

@dataclass
class StreamingConfig:
    """Configuration for streaming operations - all values configurable via environment"""
    chunk_size: int = 1000  # Rows per chunk (env: STREAMING_CHUNK_SIZE)
    # FIX #6: Consolidated memory limit default to 1600MB (matches fastapi_backend_v2.py usage)
    # Previously: 800MB default in streaming_processor.py vs 1600MB in fastapi_backend_v2.py
    # Now: Single source of truth via environment variable STREAMING_MEMORY_LIMIT_MB
    memory_limit_mb: int = 1600  # Memory limit in MB (env: STREAMING_MEMORY_LIMIT_MB)
    max_file_size_gb: int = 5  # Maximum file size in GB (env: STREAMING_MAX_FILE_SIZE_GB)
    temp_dir: Optional[str] = None  # Temp directory (env: STREAMING_TEMP_DIR)
    enable_compression: bool = True  # Enable compression (env: STREAMING_ENABLE_COMPRESSION)
    progress_callback_interval: int = 100  # Chunks between progress updates (env: STREAMING_PROGRESS_INTERVAL)
    
    @staticmethod
    def from_env() -> 'StreamingConfig':
        """Create configuration from environment variables"""
        import os
        return StreamingConfig(
            chunk_size=int(os.getenv('STREAMING_CHUNK_SIZE', '1000')),
            # FIX #6: Updated default from 800 to 1600 for consistency
            memory_limit_mb=int(os.getenv('STREAMING_MEMORY_LIMIT_MB', '1600')),
            max_file_size_gb=int(os.getenv('STREAMING_MAX_FILE_SIZE_GB', '5')),
            temp_dir=os.getenv('STREAMING_TEMP_DIR'),
            enable_compression=os.getenv('STREAMING_ENABLE_COMPRESSION', 'true').lower() == 'true',
            progress_callback_interval=int(os.getenv('STREAMING_PROGRESS_INTERVAL', '100'))
        )

@dataclass
class ProcessingStats:
    """Statistics for streaming processing"""
    total_rows: int = 0
    processed_rows: int = 0
    chunks_processed: int = 0
    memory_usage_mb: float = 0.0
    processing_time_seconds: float = 0.0
    errors: List[str] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []

class MemoryMonitor:
    """Monitors memory usage during processing with auto-scaling for wide data"""
    
    def __init__(self, limit_mb: int = 500, estimated_row_width: int = 1):
        """
        Initialize memory monitor with optional auto-scaling.
        
        Args:
            limit_mb: Base memory limit in MB (default 500)
            estimated_row_width: Estimated number of columns per row (for auto-scaling)
                                 Used to adjust limit for wide Excel sheets
        """
        self.process = psutil.Process()
        
        # Auto-scale memory limit for wide data (100+ columns)
        # Wide sheets need more memory per chunk due to DataFrame overhead
        if estimated_row_width > 50:
            # Scale: 50 cols = 1x, 100 cols = 1.2x, 200 cols = 1.4x
            scale_factor = 1.0 + (min(estimated_row_width - 50, 150) / 500.0)
            self.limit_mb = max(limit_mb, int(limit_mb * scale_factor))
            logger.info(f"memory_limit_auto_scaled", base_limit=limit_mb, 
                       scaled_limit=self.limit_mb, row_width=estimated_row_width)
        else:
            self.limit_mb = limit_mb
    
    def get_memory_usage_mb(self) -> float:
        """Get current memory usage in MB"""
        return self.process.memory_info().rss / 1024 / 1024
    
    def check_memory_limit(self) -> bool:
        """Check if memory usage exceeds limit"""
        current_usage = self.get_memory_usage_mb()
        return current_usage > self.limit_mb
    
    def force_garbage_collection(self):
        """Force garbage collection to free memory"""
        gc.collect()

class StreamingExcelProcessor:
    """Memory-efficient Excel file processor with auto-scaling for wide sheets"""
    
    def __init__(self, config: StreamingConfig):
        self.config = config
        # Initialize with default; will be updated per sheet based on column count
        self.memory_monitor = MemoryMonitor(config.memory_limit_mb, estimated_row_width=1)
        # FIX #15: Use aiolimiter rate limiter instead of ThreadPoolExecutor
        # Replaces ThreadPoolExecutor(max_workers=2) with centralized rate limiting
        self.rate_limiter = AsyncLimiter(max_rate=2, time_period=1)
    
    async def process_excel_stream(self, file_path: str, 
                                 progress_callback: Optional[Callable] = None) -> AsyncGenerator[pd.DataFrame, None]:
        """
        Stream process Excel file in chunks to avoid memory exhaustion.
        
        CRITICAL FIX: Supports both .xlsx (openpyxl) and .xls (xlrd) formats.
        
        Yields DataFrame chunks instead of loading entire file.
        """
        try:
            # CRITICAL FIX: Detect file format and use appropriate library
            file_extension = os.path.splitext(file_path)[1].lower()
            
            if file_extension == '.xls':
                # Use xlrd for old .xls format
                logger.info(f"✅ Processing .xls file with xlrd")
                import xlrd
                
                workbook = xlrd.open_workbook(file_path)
                sheet_names = workbook.sheet_names()
                
                for sheet_name in sheet_names:
                    logger.info(f"Streaming sheet: {sheet_name}")
                    sheet = workbook.sheet_by_name(sheet_name)
                    
                    # Extract headers from first row
                    headers = [str(sheet.cell_value(0, col)) if sheet.cell_value(0, col) else f'Column_{col}' 
                              for col in range(sheet.ncols)]
                    
                    # Auto-scale memory limit based on sheet width
                    self.memory_monitor = MemoryMonitor(self.config.memory_limit_mb, 
                                                       estimated_row_width=sheet.ncols)
                    
                    chunk_data = []
                    row_count = 0
                    
                    for row_idx in range(1, sheet.nrows):  # Skip header row
                        # Check memory usage periodically
                        if row_idx % 100 == 0:
                            if self.memory_monitor.check_memory_limit():
                                logger.warning(f"Memory limit exceeded, forcing garbage collection")
                                self.memory_monitor.force_garbage_collection()
                        
                        row = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                        
                        # Skip empty rows
                        if not any(cell for cell in row):
                            continue
                        
                        chunk_data.append(row)
                        row_count += 1
                        
                        # Yield chunk when size limit reached
                        if len(chunk_data) >= self.config.chunk_size:
                            df_chunk = pd.DataFrame(chunk_data, columns=headers)
                            df_chunk.attrs['sheet_name'] = sheet_name
                            df_chunk.attrs['chunk_start_row'] = row_count - len(chunk_data)
                            df_chunk.attrs['chunk_end_row'] = row_count
                            
                            yield df_chunk
                            
                            # Clear chunk data and force garbage collection
                            del chunk_data
                            gc.collect()
                            chunk_data = []
                            
                            # Progress callback
                            if progress_callback:
                                await progress_callback(f"Processing {sheet_name}", 
                                                       f"Processed {row_count} rows", 
                                                       row_count)
                    
                    # Yield remaining data
                    if chunk_data:
                        df_chunk = pd.DataFrame(chunk_data, columns=headers)
                        df_chunk.attrs['sheet_name'] = sheet_name
                        df_chunk.attrs['chunk_start_row'] = row_count - len(chunk_data)
                        df_chunk.attrs['chunk_end_row'] = row_count
                        yield df_chunk
                
                # xlrd doesn't need explicit close
                
            else:
                # Use openpyxl for .xlsx format
                logger.info(f"✅ Processing .xlsx file with openpyxl")
                from openpyxl import load_workbook
                
                # Load workbook in read-only mode for memory efficiency
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    logger.info(f"Streaming sheet: {sheet_name}")
                    
                    sheet = workbook[sheet_name]
                    headers = None
                    chunk_data = []
                    row_count = 0
                    sheet_width = sheet.max_column  # Get column count for auto-scaling
                    
                    # Auto-scale memory limit based on sheet width
                    self.memory_monitor = MemoryMonitor(self.config.memory_limit_mb, 
                                                       estimated_row_width=sheet_width)
                    
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                        # Check memory usage periodically
                        if row_idx % 100 == 0:
                            if self.memory_monitor.check_memory_limit():
                                logger.warning(f"Memory limit exceeded, forcing garbage collection")
                                self.memory_monitor.force_garbage_collection()
                        
                        if row_idx == 0:
                            # Extract headers
                            headers = [str(cell) if cell is not None else f'Column_{i}' 
                                     for i, cell in enumerate(row)]
                            continue
                        
                        # Skip empty rows
                        if not any(cell is not None for cell in row):
                            continue
                        
                        chunk_data.append(row)
                        row_count += 1
                        
                        # Yield chunk when size limit reached
                        if len(chunk_data) >= self.config.chunk_size:
                            df_chunk = pd.DataFrame(chunk_data, columns=headers)
                            df_chunk.attrs['sheet_name'] = sheet_name
                            df_chunk.attrs['chunk_start_row'] = row_count - len(chunk_data)
                            df_chunk.attrs['chunk_end_row'] = row_count
                            
                            yield df_chunk
                            
                            # Clear chunk data and force garbage collection
                            del chunk_data
                            gc.collect()
                            chunk_data = []
                            
                            # Progress callback
                            if progress_callback:
                                await progress_callback(f"Processing {sheet_name}", 
                                                       f"Processed {row_count} rows", 
                                                       row_count)
                    
                    # Yield remaining data
                    if chunk_data:
                        df_chunk = pd.DataFrame(chunk_data, columns=headers)
                        df_chunk.attrs['sheet_name'] = sheet_name
                        df_chunk.attrs['chunk_start_row'] = row_count - len(chunk_data)
                        df_chunk.attrs['chunk_end_row'] = row_count
                        yield df_chunk
                
                workbook.close()
            
        except Exception as e:
            logger.error(f"Excel streaming failed: {e}")
            raise
        finally:
            # Cleanup
            self.memory_monitor.force_garbage_collection()

class StreamingCSVProcessor:
    """Memory-efficient CSV file processor"""
    
    def __init__(self, config: StreamingConfig):
        self.config = config
        self.memory_monitor = MemoryMonitor(config.memory_limit_mb)
    
    async def process_csv_stream(self, file_path: str,
                               progress_callback: Optional[Callable] = None) -> AsyncGenerator[pd.DataFrame, None]:
        """
        Stream process CSV file in chunks.
        """
        try:
            chunk_count = 0
            total_rows = 0
            
            # Use pandas read_csv with chunksize for memory efficiency
            for chunk in pd.read_csv(file_path, chunksize=self.config.chunk_size):
                chunk_count += 1
                total_rows += len(chunk)
                
                # Add chunk metadata
                chunk.attrs['chunk_number'] = chunk_count
                chunk.attrs['chunk_start_row'] = total_rows - len(chunk)
                chunk.attrs['chunk_end_row'] = total_rows
                
                yield chunk
                
                # Memory management - check after EVERY chunk, not just every 10
                if self.memory_monitor.check_memory_limit():
                    self.memory_monitor.force_garbage_collection()
                
                # Progress callback
                if progress_callback and chunk_count % self.config.progress_callback_interval == 0:
                    await progress_callback("Processing CSV", 
                                           f"Processed {total_rows} rows", 
                                           total_rows)
        
        except Exception as e:
            logger.error(f"CSV streaming failed: {e}")
            raise

class StreamingFileProcessor:
    """Main streaming file processor that handles multiple formats"""
    
    def __init__(self, config: Optional[StreamingConfig] = None):
        self.config = config or StreamingConfig()
        self.excel_processor = StreamingExcelProcessor(self.config)
        self.csv_processor = StreamingCSVProcessor(self.config)
        self.memory_monitor = MemoryMonitor(self.config.memory_limit_mb)
    
    async def process_file_streaming(self, file_content=None, filename: str = None,
                                   progress_callback: Optional[Callable] = None,
                                   streamed_file=None) -> AsyncGenerator[Dict[str, Any], None]:
        """
        Process file in streaming mode to prevent memory exhaustion.
        
        Args:
            file_content: (deprecated) Raw bytes - use streamed_file instead
            filename: Filename for processing
            progress_callback: Optional progress callback
            streamed_file: StreamedFile object (preferred)
        
        Returns:
            AsyncGenerator yielding processed chunks with metadata
        """
        # Handle StreamedFile or fallback to bytes
        if streamed_file is not None:
            from streaming_source import StreamedFile
            if not isinstance(streamed_file, StreamedFile):
                raise TypeError("streamed_file must be a StreamedFile instance")
            
            filename = filename or streamed_file.filename
            temp_path = streamed_file.path
            file_size_gb = streamed_file.size / (1024 * 1024 * 1024)
            cleanup_temp = False
        elif file_content is not None:
            # Legacy bytes path - create temp file
            file_size_gb = len(file_content) / (1024 * 1024 * 1024)
            if file_size_gb > self.config.max_file_size_gb:
                raise ValueError(f"File size {file_size_gb:.2f}GB exceeds limit of {self.config.max_file_size_gb}GB")
            
            temp_dir = self.config.temp_dir or tempfile.gettempdir()
            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1], dir=temp_dir) as temp_file:
                temp_file.write(file_content)
                temp_path = temp_file.name
            cleanup_temp = True
        else:
            raise ValueError("Either streamed_file or file_content must be provided")
        
        # Check file size
        if file_size_gb > self.config.max_file_size_gb:
            raise ValueError(f"File size {file_size_gb:.2f}GB exceeds limit of {self.config.max_file_size_gb}GB")
        
        try:
            file_extension = os.path.splitext(filename)[1].lower()
            
            if file_extension in ['.xlsx', '.xls']:
                async for chunk in self.excel_processor.process_excel_stream(temp_path, progress_callback):
                    yield {
                        'chunk_data': chunk,
                        'file_type': 'excel',
                        'filename': filename,
                        'sheet_name': chunk.attrs.get('sheet_name', 'Sheet1'),
                        'chunk_start_row': chunk.attrs.get('chunk_start_row', 0),
                        'chunk_end_row': chunk.attrs.get('chunk_end_row', len(chunk)),
                        'memory_usage_mb': self.memory_monitor.get_memory_usage_mb()
                    }
            
            elif file_extension == '.csv':
                async for chunk in self.csv_processor.process_csv_stream(temp_path, progress_callback):
                    yield {
                        'chunk_data': chunk,
                        'file_type': 'csv',
                        'filename': filename,
                        'sheet_name': 'CSV',
                        'chunk_start_row': chunk.attrs.get('chunk_start_row', 0),
                        'chunk_end_row': chunk.attrs.get('chunk_end_row', len(chunk)),
                        'memory_usage_mb': self.memory_monitor.get_memory_usage_mb()
                    }
            
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
        
        finally:
            # Cleanup temporary file only if we created it
            if cleanup_temp:
                try:
                    os.unlink(temp_path)
                except OSError:
                    pass
            
            # Force garbage collection
            self.memory_monitor.force_garbage_collection()
    
    def get_processing_stats(self) -> ProcessingStats:
        """Get current processing statistics"""
        return ProcessingStats(
            memory_usage_mb=self.memory_monitor.get_memory_usage_mb()
        )

# Global streaming processor instance
_streaming_processor: Optional[StreamingFileProcessor] = None

def initialize_streaming_processor(config: Optional[StreamingConfig] = None):
    """Initialize the global streaming processor"""
    global _streaming_processor
    _streaming_processor = StreamingFileProcessor(config)
    logger.info("✅ Streaming processor initialized")

def get_streaming_processor() -> StreamingFileProcessor:
    """Get the global streaming processor instance"""
    if _streaming_processor is None:
        raise Exception("Streaming processor not initialized. Call initialize_streaming_processor() first.")
    return _streaming_processor
