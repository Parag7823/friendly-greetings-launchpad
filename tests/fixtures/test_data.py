"""
Test data generation utilities for ingestion tests.
Provides functions to create various test Excel/CSV files.
"""

from pathlib import Path
import random
import string
from datetime import datetime, timedelta
from openpyxl import Workbook


def create_test_excel_file(tmp_path: Path, rows: int = 100, columns: int = 5) -> Path:
    """Create a simple test Excel file with financial data."""
    # Generate realistic financial data
    vendors = ["Amazon Web Services", "Microsoft Azure", "Google Cloud", "Stripe", "Shopify"]
    categories = ["Software", "Infrastructure", "Payment Processing", "E-commerce", "Marketing"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Headers
    headers = ["Date", "Vendor", "Amount", "Category", "Description"]
    for i in range(columns - 5):
        headers.append(f"Column_{i+1}")
    ws.append(headers)
    
    # Data rows
    for i in range(rows):
        row_data = [
            (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d"),
            random.choice(vendors),
            round(random.uniform(100, 10000), 2),
            random.choice(categories),
            f"Transaction {i+1}"
        ]
        # Add extra columns if requested
        for j in range(columns - 5):
            row_data.append(f"Value_{j}")
        ws.append(row_data)
    
    output_path = tmp_path / "test_data.xlsx"
    wb.save(output_path)
    
    return output_path


def create_test_csv_file(tmp_path: Path, rows: int = 100) -> Path:
    """Create a test CSV file."""
    output_path = tmp_path / "test_data.csv"
    
    with open(output_path, 'w') as f:
        # Header
        f.write("Date,Vendor,Amount,Category,Description\n")
        
        # Rows
        vendors = ["AWS", "Azure", "GCP", "Stripe"]
        for i in range(rows):
            date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            vendor = random.choice(vendors)
            amount = round(random.uniform(100, 5000), 2)
            category = random.choice(["Infrastructure", "Payments", "Software"])
            desc = f"Transaction {i+1}"
            
            f.write(f"{date},{vendor},{amount},{category},{desc}\n")
    
    return output_path


def create_malformed_excel(tmp_path: Path) -> Path:
    """Create a corrupted/malformed Excel file."""
    output_path = tmp_path / "malformed.xlsx"
    
    # Write invalid Excel content (random bytes)
    with open(output_path, 'wb') as f:
        f.write(b'\x00' * 100 + b'CORRUPTED_EXCEL_FILE' + b'\xFF' * 100)
    
    return output_path


def create_large_excel_file(tmp_path: Path, rows: int = 100000, columns: int = 10) -> Path:
    """Create a large Excel file for performance testing."""
    wb = Workbook(write_only=True)  # write_only mode for memory efficiency
    ws = wb.create_sheet()
    
    # Headers
    headers = [f"Column_{i}" if i > 0 else "ID" for i in range(columns)]
    headers[1] = "Timestamp"
    headers[2] = "Amount"
    ws.append(headers)
    
    # Data rows
    for i in range(rows):
        row_data = [
            f"ID_{i:06d}",
            (datetime.now() - timedelta(seconds=i)).strftime("%Y-%m-%d %H:%M:%S"),
            round(random.uniform(1, 10000), 2)
        ]
        # Add remaining columns
        for j in range(columns - 3):
            row_data.append(''.join(random.choices(string.ascii_letters + string.digits, k=20)))
        ws.append(row_data)
    
    output_path = tmp_path / "large_data.xlsx"
    wb.save(output_path)
    
    return output_path

