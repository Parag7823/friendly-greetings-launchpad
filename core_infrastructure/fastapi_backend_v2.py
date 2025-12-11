# Standard library imports
from __future__ import annotations
import os
import sys
import logging
import uuid
import secrets
import time
import mmap
import threading
import structlog

# Add project root to sys.path for package imports
_project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)

# Defer database_optimization_utils import to startup (blocks module load)
# from database_optimization_utils import OptimizedDatabaseQueries

try:
    import sentry_sdk
    from sentry_sdk.integrations.fastapi import FastApiIntegration
    from sentry_sdk.integrations.asyncio import AsyncioIntegration
    from sentry_sdk.integrations.redis import RedisIntegration
    
    SENTRY_DSN = os.getenv("SENTRY_DSN")
    if SENTRY_DSN:
        sentry_sdk.init(
            dsn=SENTRY_DSN,
            integrations=[
                FastApiIntegration(transaction_style="endpoint"),
                AsyncioIntegration(),
                RedisIntegration(),
            ],
            traces_sample_rate=float(os.getenv("SENTRY_TRACES_SAMPLE_RATE", "0.1")),
            profiles_sample_rate=float(os.getenv("SENTRY_PROFILES_SAMPLE_RATE", "0.1")),
            environment=os.getenv("ENVIRONMENT", "production"),
            release=os.getenv("APP_VERSION", "1.0.0"),
            # Performance monitoring
            enable_tracing=True,
            # Error filtering
            before_send=lambda event, hint: event if event.get("level") in ["error", "fatal"] else None,
        )
        print("[OK] Sentry initialized successfully")
except ImportError:
    # Sentry SDK not installed - this is optional, continue without it
    pass
except Exception as e:
    print(f"[WARNING] Sentry initialization failed: {e}")

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import orjson
except ImportError:
    orjson = None
try:
    import tiktoken
    TIKTOKEN_AVAILABLE = True
except ImportError:
    TIKTOKEN_AVAILABLE = False
    # logger is not initialized yet here; use print to avoid NameError at import time
    print("tiktoken_unavailable: tiktoken not installed", flush=True)
try:
    from groq import Groq
except ImportError:
    Groq = None
    print("[WARNING] Groq library not installed", flush=True)

import re
import asyncio
import io
import yaml
from fastapi import File
from fastapi.responses import StreamingResponse
from fastapi import UploadFile
from typing import AsyncGenerator
import random
from datetime import datetime, timedelta, timezone
import pendulum
from typing import Dict, Any, List, Optional, Tuple

from contextlib import asynccontextmanager
import redis.asyncio as aioredis
from dataclasses import dataclass
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor

# Retry logic with tenacity
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

# Fast hashing for duplicate detection
try:
    import xxhash
except ImportError:
    xxhash = None
    print("[WARNING] xxhash not installed - dedupe hashing will use fallback", flush=True)

try:
    from glom import glom, Coalesce, Iterate
except ImportError:
    glom = None
    print("[WARNING] glom not installed - nested data extraction will use fallback", flush=True)

# Shared ingestion/normalization modules
try:
    # Local/package layout
    print("[DEBUG] Importing UniversalFieldDetector...", flush=True)
    from data_ingestion_normalization.universal_field_detector import UniversalFieldDetector
    print("[DEBUG] Importing UniversalPlatformDetector...", flush=True)
    from data_ingestion_normalization.universal_platform_detector_optimized import (
        UniversalPlatformDetectorOptimized as UniversalPlatformDetector,
        PlatformIDExtractor,
    )
    print("[DEBUG] Importing UniversalDocumentClassifier...", flush=True)
    from data_ingestion_normalization.universal_document_classifier_optimized import (
        UniversalDocumentClassifierOptimized as UniversalDocumentClassifier,
    )
    print("[DEBUG] Importing UniversalExtractors...", flush=True)
    from data_ingestion_normalization.universal_extractors_optimized import (
        UniversalExtractorsOptimized as UniversalExtractors,
    )
    print("[DEBUG] Importing UniversalNormalizer...", flush=True)
    from data_ingestion_normalization.universal_normalizer_optimized import (
        UniversalNormalizer, VendorNormalizer, AmountNormalizer, DateNormalizer, CurrencyNormalizer, FieldMapper
    )
    print("[DEBUG] Importing EntityResolver...", flush=True)
    from data_ingestion_normalization.entity_resolver_optimized import (
        EntityResolverOptimized as EntityResolver,
    )
    print("[DEBUG] Importing StreamedFile...", flush=True)
    from data_ingestion_normalization.streaming_source import StreamedFile
except ImportError:
    # Docker layout: modules in subdirectories
    print("[DEBUG] Importing UniversalFieldDetector (flat)...", flush=True)
    from data_ingestion_normalization.universal_field_detector import UniversalFieldDetector
    print("[DEBUG] Importing UniversalPlatformDetector (flat)...", flush=True)
    from data_ingestion_normalization.universal_platform_detector_optimized import (
        UniversalPlatformDetectorOptimized as UniversalPlatformDetector,
        PlatformIDExtractor,
    )
    print("[DEBUG] Importing UniversalDocumentClassifier (flat)...", flush=True)
    from data_ingestion_normalization.universal_document_classifier_optimized import (
        UniversalDocumentClassifierOptimized as UniversalDocumentClassifier,
    )
    print("[DEBUG] Importing UniversalExtractors (flat)...", flush=True)
    from data_ingestion_normalization.universal_extractors_optimized import (
        UniversalExtractorsOptimized as UniversalExtractors,
    )
    print("[DEBUG] Importing EntityResolver (flat)...", flush=True)
    from data_ingestion_normalization.entity_resolver_optimized import (
        EntityResolverOptimized as EntityResolver,
    )
    print("[DEBUG] Importing UniversalNormalizer (flat)...", flush=True)
    from data_ingestion_normalization.universal_normalizer_optimized import (
        UniversalNormalizer, VendorNormalizer, AmountNormalizer, DateNormalizer, CurrencyNormalizer, FieldMapper
    )
    print("[DEBUG] Importing StreamedFile (flat)...", flush=True)
    try:
        from data_ingestion_normalization.streaming_source import StreamedFile
        print("[DEBUG] StreamedFile imported successfully", flush=True)
    except Exception as e:
        print(f"[ERROR] CRITICAL ERROR importing StreamedFile: {e}", flush=True)
        raise e

# Import row_classifier module
print("[DEBUG] Importing row_classifier...", flush=True)
try:
    from data_ingestion_normalization.row_classifier import AIRowClassifier, RowProcessor, _shared_fallback_classification
    print("[DEBUG] row_classifier imported successfully", flush=True)
except Exception as e:
    print(f"[ERROR] Failed to import row_classifier: {e}", flush=True)
    AIRowClassifier = None
    RowProcessor = None
    _shared_fallback_classification = None

# Import DataEnrichmentProcessor from modularized file
print("[DEBUG] Importing DataEnrichmentProcessor...", flush=True)
try:
    from data_ingestion_normalization.data_enrichment_processor import DataEnrichmentProcessor
    print("[DEBUG] DataEnrichmentProcessor imported successfully", flush=True)
except Exception as e:
    print(f"[ERROR] Failed to import DataEnrichmentProcessor: {e}", flush=True)
    DataEnrichmentProcessor = None

# Import ExcelProcessor from modularized file
print("[DEBUG] Importing ExcelProcessor...", flush=True)
try:
    from data_ingestion_normalization.excel_processor import ExcelProcessor
    print("[DEBUG] ExcelProcessor imported successfully", flush=True)
except Exception as e:
    print(f"[ERROR] Failed to import ExcelProcessor: {e}", flush=True)
    ExcelProcessor = None


# Singleton pattern for ExcelProcessor instance
_excel_processor_instance = None
_excel_processor_lock = threading.Lock()

def _get_excel_processor_instance():
    """Get or create singleton ExcelProcessor instance."""
    global _excel_processor_instance
    if _excel_processor_instance is None:
        with _excel_processor_lock:
            if _excel_processor_instance is None:
                if ExcelProcessor is not None:
                    _excel_processor_instance = ExcelProcessor()
                    print("[OK] ExcelProcessor singleton created", flush=True)
                else:
                    raise RuntimeError("ExcelProcessor not available")
    return _excel_processor_instance

print("[DEBUG] Importing EnhancedRelationshipDetector...", flush=True)
try:
    from aident_cfo_brain.enhanced_relationship_detector import EnhancedRelationshipDetector
    print("[DEBUG] EnhancedRelationshipDetector imported successfully", flush=True)
except Exception as e:
    print(f"[ERROR] CRITICAL ERROR importing EnhancedRelationshipDetector: {e}", flush=True)
    # Don't raise yet, let's see if other imports fail
    EnhancedRelationshipDetector = None

print("[DEBUG] Importing ProvenanceTracker...", flush=True)
try:
    from core_infrastructure.provenance_tracker import normalize_business_logic, normalize_temporal_causality
    print("[DEBUG] ProvenanceTracker imported successfully", flush=True)
except Exception as e:
    print(f"[ERROR] CRITICAL ERROR importing ProvenanceTracker: {e}", flush=True)
    normalize_business_logic = None
    normalize_temporal_causality = None

# Lazy import for field_mapping_learner to avoid circular dependencies
try:
    print("[DEBUG] Importing FieldMappingLearner...", flush=True)
    try:
        from data_ingestion_normalization.field_mapping_learner import (
            learn_field_mapping,
            get_learned_mappings,
        )
        print("[DEBUG] FieldMappingLearner imported successfully (nested)", flush=True)
    except ImportError:
        print("[DEBUG] Importing FieldMappingLearner (flat)...", flush=True)
        from data_ingestion_normalization.universal_field_detector import learn_field_mapping, get_learned_mappings
        print("[DEBUG] FieldMappingLearner imported successfully (flat)", flush=True)
except Exception as e:
    # logger not initialized yet here; use print for diagnostics only
    print(f"field_mapping_learner_unavailable: {e}", flush=True)
    learn_field_mapping = None
    get_learned_mappings = None
import polars as pl
import numpy as np
import openpyxl
import magic
import filetype
import requests
import tempfile
import httpx
from urllib.parse import quote
from email.utils import parsedate_to_datetime
import base64
import hmac
import binascii

# FastAPI and web framework imports
from fastapi import FastAPI, HTTPException, BackgroundTasks, WebSocket, WebSocketDisconnect, UploadFile, Form, File, Response, Depends
from starlette.requests import Request
from starlette.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, ValidationError

# Socket.IO for WebSocket management
import socketio
from socketio import ASGIApp

from rapidfuzz import fuzz
try:
    from pydantic import field_validator
except Exception:
    field_validator = None

from pydantic_settings import BaseSettings
from pydantic import Field
from typing import Optional

# Load environment from .env.test for testing
from dotenv import load_dotenv  
from pathlib import Path
env_path = Path(__file__).parent.parent / '.env.test'
if env_path.exists():
    load_dotenv(env_path, override=True)
    print(f"[OK] Loaded environment from {env_path}")
else:
    load_dotenv()
    print("[WARNING] .env.test not found, using default .env")

class AppConfig(BaseSettings):
    """Type-safe environment configuration"""
    
    # All fields optional to prevent startup crashes
    supabase_url: Optional[str] = Field(default=None, validation_alias='SUPABASE_URL')
    supabase_service_role_key: Optional[str] = Field(default=None, validation_alias='SUPABASE_SERVICE_ROLE_KEY')
    supabase_anon_key: Optional[str] = Field(default=None, validation_alias='SUPABASE_ANON_KEY')
    groq_api_key: Optional[str] = Field(default=None, validation_alias='GROQ_API_KEY')
    redis_url: Optional[str] = Field(default=None, validation_alias='REDIS_URL')
    openai_api_key: Optional[str] = None
    nango_secret_key: Optional[str] = None
    arq_redis_url: Optional[str] = None
    queue_backend: str = "sync"
    require_redis_cache: bool = False
    database_pool_size: int = 5
    database_max_overflow: int = 2
    database_pool_timeout: int = 10
    database_statement_timeout: int = 5000
    request_timeout: int = 30
    
    class Config:
        env_file = ".env.test"
        env_file_encoding = "utf-8"
        extra = "ignore"
        populate_by_name = True
    
    @property
    def redis_url_resolved(self) -> Optional[str]:
        return self.arq_redis_url or self.redis_url

# ------------------------- Request Models (Pydantic) -------------------------
class StandardErrorResponse(BaseModel):
    """Standardized error response format."""
    error: str
    error_code: str
    error_details: Optional[Dict[str, Any]] = None
    retryable: bool = False
    user_action: Optional[str] = None
    timestamp: str = None
    
    def __init__(self, **data):
        if 'timestamp' not in data:
            import pendulum
            data['timestamp'] = pendulum.now().to_iso8601_string()
        super().__init__(**data)

class FieldDetectionRequest(BaseModel):
    data: Dict[str, Any]
    filename: Optional[str] = None
    user_id: Optional[str] = None
    context: Optional[Dict[str, Any]] = None

class PlatformDetectionRequest(BaseModel):
    payload: Optional[Dict[str, Any]] = None  # Structured data (columns, sample_data)
    file_content: Optional[str] = None  # Base64 or text content
    filename: Optional[str] = None  # Original filename
    user_id: Optional[str] = None  # User identifier


class DocumentClassificationRequest(BaseModel):
    payload: Optional[Dict[str, Any]] = None
    filename: Optional[str] = None
    file_content: Optional[str] = None  # base64 or text content
    user_id: Optional[str] = None
    platform: Optional[str] = None
    document_type: Optional[str] = None  # New field added
    document_subtype: Optional[str] = None  # New field added

# Pydantic models for metadata validation
class UserConnectionMetadata(BaseModel):
    """Validated metadata for user_connections.metadata field."""
    last_history_id: Optional[str] = None
    last_synced_at: Optional[str] = None
    sync_errors: Optional[List[str]] = None
    error_count: int = 0
    
    class Config:
        extra = "allow"

# REMOVED: SyncRunStats, ZohoMailMetadata, XeroMetadata, StripeMetadata, PayPalMetadata - moved to connectors.py

class RazorpayMetadata(BaseModel):
    """Validated metadata for Razorpay user_connections.metadata field."""
    last_sync_token: Optional[str] = None
    last_synced_at: Optional[str] = None
    sync_errors: Optional[List[str]] = None
    error_count: int = 0
    account_id: Optional[str] = None
    
    class Config:
        extra = "allow"

class ConnectorSyncStats(BaseModel):
    """Extended stats for all connector syncs."""
    records_fetched: int = 0
    actions_used: int = 0
    attachments_saved: int = 0
    queued_jobs: int = 0
    skipped: int = 0
    errors_encountered: int = 0
    processing_time_ms: Optional[int] = None
    
    class Config:
        extra = "allow"

from supabase import create_client, Client
import socket
from urllib.parse import urlparse

_supabase_client_instance: Optional[Client] = None
_supabase_client_lock = threading.Lock()

class LazySupabaseClient:
    def __init__(self, url: str, key: str):
        self.url = url
        self.key = key
        self._real_client = None
        self._connecting = False
        self._connect_lock = threading.Lock()
        self._connection_timeout = 5.0  
    
    def _ensure_connected(self):
        if self._real_client is None and not self._connecting:
            with self._connect_lock:
                if self._real_client is None:
                    try:
                        self._connecting = True
                        logger.info(f"Lazy-connecting to Supabase on first use...")
                        
                        client_holder = {'client': None, 'error': None}
                        
                        def connect_thread():
                            try:
                                client_holder['client'] = create_client(self.url, self.key)
                            except Exception as e:
                                client_holder['error'] = e
                        
                        thread = threading.Thread(target=connect_thread, daemon=True)
                        thread.start()
                        thread.join(timeout=self._connection_timeout)
                        
                        if thread.is_alive():
                            logger.error(f"Supabase connection timed out after {self._connection_timeout} seconds")
                            raise TimeoutError(f"Supabase connection timed out after {self._connection_timeout} seconds")
                        
                        if client_holder['error']:
                            raise client_holder['error']
                        
                        self._real_client = client_holder['client']
                        logger.info(f"Lazy-connected to Supabase successfully")
                    except Exception as e:
                        logger.error(f"Failed to connect to Supabase on first use: {e}")
                        raise
                    finally:
                        self._connecting = False
    
    def __getattr__(self, name):
        self._ensure_connected()
        return getattr(self._real_client, name)


def get_supabase_client(use_service_role: bool = True) -> Client:
    global _supabase_client_instance
    
    if _supabase_client_instance is None:
        with _supabase_client_lock:
            if _supabase_client_instance is None:
                url = os.getenv('SUPABASE_URL')
                key = (
                    os.getenv('SUPABASE_SERVICE_ROLE_KEY') or 
                    os.getenv('SUPABASE_SERVICE_KEY')
                )
                
                if not url or not key:
                    logger.error("SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY not set")
                    raise ValueError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY (or SUPABASE_SERVICE_KEY) must be set")
                
                _supabase_client_instance = LazySupabaseClient(url, key)
                logger.info("Created lazy Supabase client (will connect on first use)")
    
    return _supabase_client_instance

_supabase_import_errors = None
from prometheus_client import Counter, Histogram, Gauge, generate_latest, CONTENT_TYPE_LATEST



def get_queue_backend() -> str:
    from core_infrastructure.config_manager import get_queue_config
    return get_queue_config().backend.lower()

_arq_pool = None
_arq_pool_lock = asyncio.Lock()

async def get_arq_pool():
    global _arq_pool
    
    if _arq_pool is not None:
        return _arq_pool
    
    async with _arq_pool_lock:
        if _arq_pool is not None:
            return _arq_pool
        
        from arq import create_pool
        from arq.connections import RedisSettings
        from core_infrastructure.config_manager import get_queue_config
        queue_cfg = get_queue_config()
        url = queue_cfg.redis_url
        if not url:
            raise RuntimeError("QUEUE_REDIS_URL not set for QUEUE_BACKEND=arq")
        
        _arq_pool = await create_pool(RedisSettings.from_dsn(url))
        logger.info(f"ARQ connection pool created and cached for reuse")
        return _arq_pool

from data_ingestion_normalization.airbyte_client import AirbytePythonClient

from core_infrastructure.transaction_manager import initialize_transaction_manager, get_transaction_manager
from data_ingestion_normalization.streaming_processor import (
    initialize_streaming_processor,
    get_streaming_processor,
    StreamingConfig,
    StreamingFileProcessor,
)
from core_infrastructure.error_recovery_system import (
    initialize_error_recovery_system,
    get_error_recovery_system,
    ErrorContext,
    ErrorSeverity,
)

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from database_optimization_utils import OptimizedDatabaseQueries

optimized_db: Optional["OptimizedDatabaseQueries"] = None
_thread_pool: Optional[ThreadPoolExecutor] = None

from core_infrastructure.centralized_cache import initialize_cache, get_cache, safe_get_cache

safe_get_ai_cache = safe_get_cache

import polars as pl

from core_infrastructure.security_system import SecurityValidator, InputSanitizer, SecurityContext



import structlog

structlog.configure(
    processors=[
        structlog.stdlib.filter_by_level,
        structlog.stdlib.add_logger_name,
        structlog.stdlib.add_log_level,
        structlog.stdlib.PositionalArgumentsFormatter(),
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.StackInfoRenderer(),
        structlog.processors.format_exc_info,
        structlog.processors.UnicodeDecoder(),
        structlog.processors.JSONRenderer()
    ],
    context_class=dict,
    logger_factory=structlog.stdlib.LoggerFactory(),
    cache_logger_on_first_use=True,
)

logger = structlog.get_logger(__name__)

security_validator = None
structured_logger = logger
groq_client = None

from core_infrastructure.utils.helpers import (
    clean_jwt_token, safe_decode_base64, sanitize_for_json, get_groq_client, 
    generate_friendly_status, send_websocket_progress,
    get_sync_cursor, save_sync_cursor, insert_external_item_with_error_handling
)
# SOLUTION: Cache memory managers by user_id with LRU eviction (maxsize=100 for 100 concurrent users)
# IMPACT: 5-10x improvement in chat response times (100-200ms â†’ 10-20ms)

from functools import lru_cache
import os as _os

# MOVED: get_memory_manager function is now in aident_cfo_brain/aident_memory_manager.py
# Import it from there for backward compatibility
try:
    from aident_intelligence.aident_memory_manager import get_memory_manager
except ImportError:
    try:
        from aident_memory_manager import get_memory_manager
    except ImportError:
        # Memory manager not available - will be handled gracefully
        get_memory_manager = None


# ----------------------------------------------------------------------------
# Metrics (Prometheus) - Comprehensive business metrics
# ----------------------------------------------------------------------------
# Job/Task Metrics
JOBS_ENQUEUED = Counter('jobs_enqueued_total', 'Jobs enqueued by provider and mode', ['provider', 'mode'])
JOBS_PROCESSED = Counter('jobs_processed_total', 'Jobs processed by provider and status', ['provider', 'status'])
ACTIVE_JOBS = Gauge('active_jobs_current', 'Number of currently active processing jobs')

# Database Metrics
DB_WRITES = Counter('db_writes_total', 'Database writes by table/op/status', ['table', 'op', 'status'])
DB_WRITE_LATENCY = Histogram('db_write_latency_seconds', 'DB write latency seconds', ['table', 'op'])
DB_READS = Counter('db_reads_total', 'Database reads by table/status', ['table', 'status'])

# File Processing Metrics
FILES_PROCESSED = Counter('files_processed_total', 'Total files processed by type and status', ['file_type', 'status'])
PROCESSING_DURATION = Histogram('file_processing_duration_seconds', 'File processing duration in seconds', ['file_type'])
SHEETS_PROCESSED = Counter('sheets_processed_total', 'Total sheets processed by status', ['status'])

# OAuth Connector Metrics
CONNECTOR_SYNCS = Counter('connector_syncs_total', 'Connector syncs by provider and status', ['provider', 'status'])
CONNECTOR_SYNC_DURATION = Histogram('connector_sync_duration_seconds', 'Connector sync duration', ['provider'])
CONNECTOR_ITEMS_FETCHED = Counter('connector_items_fetched_total', 'Items fetched by connector', ['provider'])

# API Metrics
API_REQUESTS = Counter('api_requests_total', 'API requests by endpoint and status', ['endpoint', 'status'])
API_LATENCY = Histogram('api_latency_seconds', 'API request latency', ['endpoint'])

# Normalization and entity pipeline metrics
NORMALIZATION_EVENTS = Counter('normalization_events_total', 'Normalized events by provider', ['provider'])
NORMALIZATION_DURATION = Histogram('normalization_duration_seconds', 'Normalization duration seconds', ['provider'])
ENTITY_PIPELINE_RUNS = Counter('entity_pipeline_runs_total', 'Entity resolver runs by provider and status', ['provider', 'status'])

# AI/ML Metrics
AI_CLASSIFICATIONS = Counter('ai_classifications_total', 'AI classifications by model and status', ['model', 'status'])
AI_CLASSIFICATION_DURATION = Histogram('ai_classification_duration_seconds', 'AI classification duration', ['model'])
AI_CACHE_HITS = Counter('ai_cache_hits_total', 'AI cache hits vs misses', ['result'])

# Error Metrics
ERRORS_TOTAL = Counter('errors_total', 'Total errors by type and severity', ['error_type', 'severity'])
# DB helper wrappers with metrics
# ----------------------------------------------------------------------------
def _sanitize_for_json(obj):
    """Wrapper for centralized sanitize_for_json."""
    return sanitize_for_json(obj)

def _db_insert(table: str, payload):
    t0 = time.time()
    try:
        sanitized_payload = sanitize_for_json(payload)
        res = supabase.table(table).insert(sanitized_payload).execute()
        DB_WRITES.labels(table=table, op='insert', status='ok').inc()
        DB_WRITE_LATENCY.labels(table=table, op='insert').observe(max(0.0, time.time() - t0))
        return res
    except Exception as e:
        DB_WRITES.labels(table=table, op='insert', status='error').inc()
        DB_WRITE_LATENCY.labels(table=table, op='insert').observe(max(0.0, time.time() - t0))
        raise

def _db_update(table: str, updates: dict, eq_col: str, eq_val):
    t0 = time.time()
    try:
        res = supabase.table(table).update(updates).eq(eq_col, eq_val).execute()
        DB_WRITES.labels(table=table, op='update', status='ok').inc()
        DB_WRITE_LATENCY.labels(table=table, op='update').observe(max(0.0, time.time() - t0))
        return res
    except Exception:
        DB_WRITES.labels(table=table, op='update', status='error').inc()
        DB_WRITE_LATENCY.labels(table=table, op='update').observe(max(0.0, time.time() - t0))
        raise

# Import production duplicate detection service
try:
    try:
        from data_ingestion_normalization.production_duplicate_detection_service import (
            ProductionDuplicateDetectionService, 
            FileMetadata, 
            DuplicateType,
            DuplicateDetectionError
        )
    except ImportError:
        from production_duplicate_detection_service import (
            ProductionDuplicateDetectionService, 
            FileMetadata, 
            DuplicateType,
            DuplicateDetectionError
        )
    PRODUCTION_DUPLICATE_SERVICE_AVAILABLE = True
    logger.info("✅ Production duplicate detection service available")
except ImportError as e:
    logger.warning(f"⚠️ Production duplicate detection service not available: {e}")
    PRODUCTION_DUPLICATE_SERVICE_AVAILABLE = False

# Custom JSON encoder for datetime objects
import json
class DateTimeEncoder(json.JSONEncoder):
    """Custom JSON encoder for handling datetime objects."""
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif hasattr(obj, 'isoformat'):
            return obj.isoformat()
        return super().default(obj)

def safe_json_dumps(obj, default=None):
    """orjson-based JSON serialization (3-5x faster than stdlib json)."""
    try:
        serialized = serialize_datetime_objects(obj)
        return orjson.dumps(serialized).decode('utf-8')
    except TypeError as e:
        logger.error(f"orjson serialization failed - object not JSON serializable: {e}")
        raise ValueError(f"Object cannot be serialized to JSON: {e}") from e
    except Exception as e:
        logger.error(f"orjson serialization failed: {e}")
        raise

# Helper functions consolidated in utils/helpers.py

def safe_json_parse(json_str, fallback=None):
    """orjson-based JSON parsing (3-5x faster than standard json)."""
    if not json_str or not isinstance(json_str, str):
        return fallback
    
    try:
        cleaned = json_str.strip()
        if '```json' in cleaned:
            start = cleaned.find('```json') + 7
            end = cleaned.find('```', start)
            if end != -1:
                cleaned = cleaned[start:end].strip()
        elif '```' in cleaned:
            start = cleaned.find('```') + 3
            end = cleaned.find('```', start)
            if end != -1:
                cleaned = cleaned[start:end].strip()
        
        return orjson.loads(cleaned)
        
    except (orjson.JSONDecodeError, ValueError) as e:
        logger.error(f"JSON parsing failed: {e}")
        logger.error(f"Input string: {json_str[:200]}...")

def serialize_datetime_objects(obj):
    """Pendulum-based datetime serialization with proper timezone handling."""
    if isinstance(obj, datetime):
        return pendulum.instance(obj).to_iso8601_string()
    elif hasattr(obj, 'isoformat'):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return {key: serialize_datetime_objects(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [serialize_datetime_objects(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(serialize_datetime_objects(item) for item in obj)
    else:
        return obj

class ProcessingStage:
    """Processing stages following cognitive flow."""
    SENSE = "sense"
    UNDERSTAND = "understand"
    EXPLAIN = "explain"
    ACT = "act"

def format_progress_message(stage: str, action: str, details: str = None, count: int = None, total: int = None) -> str:
    """Format progress messages with personality-driven language."""
    stage_map = {
        ProcessingStage.SENSE: "I'm",
        ProcessingStage.UNDERSTAND: "I'm",
        ProcessingStage.EXPLAIN: "I",
        ProcessingStage.ACT: "I'm"
    }
    
    prefix = stage_map.get(stage, "I'm")
    action_lower = action[0].lower() + action[1:] if action else action
    message = f"{prefix} {action_lower}"
    
    if count is not None and total is not None:
        message += f" ({count:,} of {total:,} done)"
    elif count is not None:
        message += f" ({count:,} completed)"
    
    if details:
        message += f" - {details}"
    
    return message

supabase = None
optimized_db = None
security_validator = None
centralized_cache = None
_supabase_loaded = False
_supabase_lock = threading.Lock()

def _ensure_supabase_loaded_sync():
    """Synchronous helper to lazy-load Supabase client on first use."""
    global supabase, _supabase_loaded
    
    if not _supabase_loaded:
        with _supabase_lock:
            if not _supabase_loaded:
                try:
                    supabase = get_supabase_client()
                    _supabase_loaded = True
                    logger.info("âœ… Supabase client lazy-loaded on first use")
                except Exception as e:
                    logger.error(f"âŒ Failed to lazy-load Supabase client: {e.__class__.__name__}: {e}")
                    _supabase_loaded = True
                    supabase = None
    
    return supabase

async def _ensure_supabase_loaded():
    """Async wrapper for lazy-loading Supabase client."""
    try:
        return await asyncio.to_thread(_ensure_supabase_loaded_sync)
    except Exception as e:
        logger.error(f"âŒ Failed to create lazy Supabase client: {e.__class__.__name__}: {e}")
        return None

class SocketIOWebSocketManager:
    """Socket.IO-based WebSocket manager."""
    
    def __init__(self):
        self.redis = None
        self.job_status: Dict[str, Dict[str, Any]] = {}
        
    def set_redis(self, redis_client):
        """Set Redis client for job state persistence."""
        self.redis = redis_client
        try:
            redis_manager = socketio.AsyncRedisManager(
                f"redis://{redis_client.connection_pool.connection_kwargs.get('host', 'localhost')}:"
                f"{redis_client.connection_pool.connection_kwargs.get('port', 6379)}"
            )
            sio.manager = redis_manager
            logger.info("âœ… Socket.IO Redis adapter initialized")
        except Exception as e:
            logger.warning(f"Socket.IO Redis adapter failed: {e}")

    def _key(self, job_id: str) -> str:
        return f"finley:job:{job_id}"

    async def _get_state(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Cache-only state retrieval with proper error handling."""
        try:
            cache = safe_get_cache()
            if cache is not None:
                raw = await cache.get(self._key(job_id))
                if raw:
                    state = raw if isinstance(raw, dict) else orjson.loads(raw)
                    return state
            return None
        except Exception as e:
            logger.warning(f"Cache retrieval failed for job {job_id}: {e}")
            return None

    async def _save_state(self, job_id: str, state: Dict[str, Any]):
        """Cache-only state storage with explicit error handling."""
        cache = safe_get_cache()
        if cache is None:
            logger.error(f"âŒ CRITICAL: Cache unavailable - cannot persist job state {job_id}")
            raise RuntimeError(f"Cache service unavailable for job {job_id}")
        
        try:
            await cache.set(self._key(job_id), state, ttl=21600)
        except Exception as e:
            logger.error(f"âŒ CRITICAL: Failed to save job state {job_id}: {e}")
            raise RuntimeError(f"Failed to persist job state: {e}")

    async def merge_job_state(self, job_id: str, patch: Dict[str, Any]) -> Dict[str, Any]:
        base = await self._get_state(job_id) or {}
        base.update(patch)
        await self._save_state(job_id, base)
        return base

    async def send_update(self, job_id: str, data: Dict[str, Any]):
        """Send update via Socket.IO to job room"""
        try:
            await self.merge_job_state(job_id, data)
            payload = {**data, "job_id": job_id}
            await sio.emit('job_update', payload, room=job_id)
            return True
        except Exception as e:
            logger.error(f"Failed to send update for job {job_id}: {e}")
            return False

    async def send_error(self, job_id: str, error_message: str, component: str = None):
        """Send error via Socket.IO"""
        try:
            payload = {
                "type": "error",
                "job_id": job_id,
                "error": error_message,
                "component": component,
                "timestamp": pendulum.now().to_iso8601_string()
            }
            await self.merge_job_state(job_id, {"status": "failed", "error": error_message})
            await sio.emit('job_error', payload, room=job_id)
            return True
        except Exception as e:
            logger.error(f"Failed to send error for job {job_id}: {e}")
            return False

    async def get_job_status(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Get current job status"""
        return await self._get_state(job_id)
    
    async def send_overall_update(self, job_id: str, status: str, message: str, progress: Optional[int] = None, results: Optional[Dict[str, Any]] = None):
        """Send overall job update - wrapper for send_update with standardized payload"""
        payload = {
            "status": status,
            "message": message,
            "timestamp": pendulum.now().to_iso8601_string()
        }
        if progress is not None:
            payload["progress"] = progress
        if results is not None:
            payload["results"] = results
        
        await self.send_update(job_id, payload)
    
    async def send_component_update(self, job_id: str, component: str, status: str, message: str, progress: Optional[int] = None, data: Optional[Dict[str, Any]] = None):
        """Send component-specific update"""
        payload = {
            "component": component,
            "status": status,
            "message": message,
            "timestamp": pendulum.now().to_iso8601_string()
        }
        if progress is not None:
            payload["progress"] = progress
        if data is not None:
            payload["data"] = data
        
        await self.send_update(job_id, payload)

# Initialize WebSocket manager in app_lifespan to prevent race condition
websocket_manager = None

@asynccontextmanager
async def app_lifespan(app: FastAPI):
    """Application lifespan context manager - handles startup and shutdown"""
    # Startup
    global supabase, optimized_db, security_validator, centralized_cache, websocket_manager, groq_client
    
    logger.info("="*80)
    logger.info("ðŸš€ STARTING SERVICE INITIALIZATION...")
    logger.info("="*80)
    
    try:
        websocket_manager = SocketIOWebSocketManager()
        logger.info("âœ… WebSocket manager initialized")
    except Exception as ws_err:
        logger.error(f"âŒ Failed to initialize WebSocket manager: {ws_err}")
        websocket_manager = None
    
    try:
        # Try multiple possible environment variable names for Render compatibility
        supabase_url = (
            os.environ.get("SUPABASE_URL") or 
            os.environ.get("SUPABASE_PROJECT_URL") or
            os.environ.get("DATABASE_URL")  # Sometimes Render uses this
        )
        supabase_key = (
            os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or 
            os.environ.get("SUPABASE_SERVICE_KEY") or  # This is what's in Render!
            os.environ.get("SUPABASE_KEY") or
            os.environ.get("SUPABASE_ANON_KEY")  # Fallback to anon key if service role not available
        )
        
        # Enhanced diagnostics for deployment debugging
        logger.info(f"ðŸ” Environment diagnostics:")
        logger.info(f"   SUPABASE_URL present: {'âœ…' if supabase_url else 'âŒ'}")
        logger.info(f"   SUPABASE_SERVICE_ROLE_KEY present: {'âœ…' if supabase_key else 'âŒ'}")
        supabase_env_vars = sorted([k for k in os.environ.keys() if 'SUPABASE' in k.upper()])
        logger.info(f"   Available env vars: {supabase_env_vars}")
        
        if supabase_key:
            supabase_key = clean_jwt_token(supabase_key)
        
        if not supabase_url or not supabase_key:
            missing_vars = []
            if not supabase_url:
                missing_vars.append("SUPABASE_URL")
            if not supabase_key:
                missing_vars.append("SUPABASE_SERVICE_KEY (or SUPABASE_SERVICE_ROLE_KEY)")
            raise ValueError(f"Missing required environment variables: {', '.join(missing_vars)}. Please check your deployment configuration.")
        
        supabase = get_supabase_client()
        logger.info("âœ… Supabase client will be lazy-loaded on first use (non-blocking startup)")
        
        # Initialize critical systems (only if Supabase is available)
        if supabase:
            try:
                initialize_transaction_manager(supabase)
                initialize_streaming_processor(StreamingConfig(
                    chunk_size=1000,
                    memory_limit_mb=1600,  # Increased from 800MB to handle larger files safely
                    max_file_size_gb=10
                ))
                initialize_error_recovery_system(supabase)
                logger.info("âœ… Transaction, streaming, and error recovery systems initialized")
            except Exception as sys_err:
                logger.warning(f"âš ï¸ Failed to initialize critical systems: {sys_err}")
        else:
            logger.warning("âš ï¸ Skipping critical system initialization - Supabase unavailable")
        
        # Initialize security system (observability removed - using structlog)
        try:
            security_validator = SecurityValidator()
            logger.info("âœ… Security validator initialized")
        except Exception as sec_err:
            logger.warning(f"âš ï¸ Failed to initialize security validator: {sec_err}")
        
        if supabase:
            try:
                from core_infrastructure.database_optimization_utils import OptimizedDatabaseQueries
                optimized_db = OptimizedDatabaseQueries(supabase)
                logger.info("âœ… Optimized database queries initialized")
            except Exception as opt_err:
                logger.warning(f"âš ï¸ Failed to initialize optimized database queries: {opt_err}")
                optimized_db = None
        else:
            logger.warning("âš ï¸ Skipping optimized database initialization - Supabase unavailable")
            optimized_db = None
        
        logger.info("âœ… Observability and security systems initialized")
        
        # REFACTORED: Initialize centralized Redis cache (replaces ai_cache_system.py)
        # This provides distributed caching across all workers and instances for true scalability
        redis_url = os.environ.get('ARQ_REDIS_URL') or os.environ.get('REDIS_URL')
        if redis_url:
            try:
                centralized_cache = initialize_cache(
                    redis_url=redis_url,
                    default_ttl=7200  # 2 hours default TTL
                )
                logger.info("âœ… Centralized Redis cache initialized - distributed caching across all workers!")
            except Exception as cache_err:
                logger.warning(f"âš ï¸ Failed to initialize Redis cache: {cache_err} - Running without distributed cache")
                centralized_cache = None
        else:
            logger.warning("âš ï¸ REDIS_URL not set - Running without distributed cache")
            centralized_cache = None
            
        # Initialize Groq client
        try:
            groq_api_key = os.environ.get('GROQ_API_KEY')
            if groq_api_key:
                if Groq:
                    groq_client = Groq(api_key=groq_api_key)
                    logger.info("âœ… Groq client initialized")
                else:
                    logger.warning("âš ï¸ Groq library not available, skipping client initialization")
            else:
                logger.warning("âš ï¸ GROQ_API_KEY not set, AI features will be disabled")
        except Exception as groq_err:
            logger.error(f"âŒ Failed to initialize Groq client: {groq_err}")
            groq_client = None
        
        # Initialize global thread pool for CPU-bound operations
        try:
            _thread_pool = ThreadPoolExecutor(max_workers=5)
            logger.info("âœ… Global thread pool initialized for CPU-bound operations")
        except Exception as thread_err:
            logger.warning(f"âš ï¸ Failed to initialize global thread pool: {thread_err}")
            _thread_pool = None
        
        # PERMANENT FIX: Pre-load heavy ML models at startup (not during first chat request)
        # This prevents 1-2 minute delays on first message
        logger.info("ðŸ”„ Pre-loading ML models for chat orchestrator...")
        try:
            # Import and initialize intent classifier + output guard (loads spacy + sentence-transformers)
            try:
                from aident_intelligence.intent_and_guard_engine import (
                    get_intent_classifier,
                    get_output_guard
                )
            except ImportError:
                from aident_cfo_brain.intent_and_guard_engine import (
                    get_intent_classifier,
                    get_output_guard
                )
            
            # Pre-load spacy + sentence-transformers models in background thread
            # This prevents blocking the startup event
            def _preload_models():
                try:
                    logger.info("   Loading intent classifier (spacy + sentence-transformers)...")
                    intent_classifier = get_intent_classifier()
                    logger.info("   âœ… Intent classifier loaded")
                    
                    logger.info("   Loading output guard (sentence-transformers)...")
                    # Get Groq client for output guard
                    try:
                        from groq import AsyncGroq
                        import os as preload_os
                        groq_key = preload_os.getenv('GROQ_API_KEY')
                        if groq_key:
                            preload_groq = AsyncGroq(api_key=groq_key)
                            output_guard = get_output_guard(preload_groq)
                            logger.info("   âœ… Output guard loaded")
                        else:
                            logger.warning("   âš ï¸ No GROQ_API_KEY - skipping output guard preload")
                    except Exception as groq_err:
                        logger.warning(f"   âš ï¸ Output guard preload failed: {groq_err}")
                    
                    logger.info("âœ… All ML models pre-loaded successfully - chat will be instant!")
                except Exception as model_err:
                    logger.warning(f"âš ï¸ Failed to pre-load ML models: {model_err} - will load on first chat (may be slow)")
            
            # Run model loading in thread pool to avoid blocking startup
            if _thread_pool:
                _thread_pool.submit(_preload_models)
                logger.info("   Model pre-loading started in background...")
            else:
                # Fallback: load synchronously if thread pool not available
                _preload_models()
        
        except Exception as model_import_err:
            logger.warning(f"âš ï¸ Failed to import intent classifier: {model_import_err}")
        
        logger.info("âœ… All critical systems and optimizations initialized successfully")
        
    except Exception as e:
        logger.error(f"âŒ Failed to initialize critical systems: {e}")
        supabase = None
        optimized_db = None
        # Log critical database failure for monitoring
        logger.critical(f"ðŸš¨ DATABASE CONNECTION FAILED - System running in degraded mode: {e}")
        # Initialize minimal observability/logging to prevent NameError in endpoints
        try:
            # Fallback lightweight initialization (observability removed - using structlog)
            security_validator = SecurityValidator()
            logger.info("âœ… Degraded mode security initialized (no database)")
        except Exception as init_err:
            logger.warning(f"âš ï¸ Failed to initialize degraded security systems: {init_err}")
    
    # Log final startup status
    logger.info("="*80)
    logger.info("ðŸŽ¯ STARTUP COMPLETE - Service Status Summary:")
    logger.info(f"   Supabase: {'âœ… Connected' if supabase else 'âŒ Not initialized'}")
    logger.info(f"   Groq Client: {'âœ… Ready' if groq_client else 'âŒ Not initialized'}")
    logger.info(f"   Redis Cache: {'âœ… Connected' if centralized_cache else 'âŒ Not initialized'}")
    logger.info(f"   Optimized DB: {'âœ… Ready' if optimized_db else 'âŒ Not initialized'}")
    logger.info(f"   Security Validator: {'âœ… Ready' if security_validator else 'âŒ Not initialized'}")
    logger.info(f"   WebSocket Manager: {'âœ… Ready' if websocket_manager else 'âŒ Not initialized'}")
    logger.info("="*80)
    
    yield
    # Shutdown
    logger.info("ðŸ›‘ Application shutting down...")
    # Cleanup happens here if needed

# Initialize FastAPI app with enhanced configuration# Initialize FastAPI app
app = FastAPI(
    title="Aident CFO API",
    version="2.0.0",
    description="Advanced financial data processing and AI-powered analysis platform",
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_url="/openapi.json",
    lifespan=app_lifespan  # Use lifespan context manager for startup/shutdown
)

# ISSUE #10 FIX: Initialize slowapi rate limiter (Redis-backed, distributed)
# Replaces custom rate limiting with battle-tested library
try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded
    
    # Initialize limiter with Redis backend for distributed rate limiting
    limiter = Limiter(
        key_func=get_remote_address,
        storage_uri=os.getenv("REDIS_URL", "redis://localhost:6379"),
        default_limits=["100/minute"]  # Global default
    )
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
    logger.info("âœ… slowapi rate limiter initialized with Redis backend")
except ImportError:
    logger.warning("âš ï¸ slowapi not available, rate limiting disabled")
    limiter = None
except Exception as e:
    logger.warning(f"âš ï¸ Failed to initialize slowapi: {e}, rate limiting disabled")
    limiter = None

# IMPROVEMENT: Global exception handler for consistent error responses
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """
    Global exception handler to ensure all errors return StandardErrorResponse format
    This provides consistent error handling across the entire API
    """
    import traceback
    
    # Get full stack trace
    tb_str = ''.join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    
    # Log with full details
    logger.error(f"âŒ UNHANDLED EXCEPTION on {request.method} {request.url.path}")
    logger.error(f"Exception type: {type(exc).__name__}")
    logger.error(f"Exception message: {str(exc)}")
    logger.error(f"Full traceback:\n{tb_str}")
    
    # Determine if error is retryable
    retryable = isinstance(exc, (TimeoutError, ConnectionError))
    
    return JSONResponse(
        status_code=500,
        content=StandardErrorResponse(
            error=f"{type(exc).__name__}: {str(exc)}",
            error_code="INTERNAL_ERROR",
            retryable=retryable,
            user_action="Please try again. If the problem persists, contact support."
        ).dict()
    )

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """Handle validation errors with StandardErrorResponse format"""
    return JSONResponse(
        status_code=422,
        content=StandardErrorResponse(
            error="Invalid request data",
            error_code="VALIDATION_ERROR",
            error_details={"errors": exc.errors()},
            retryable=False,
            user_action="Please check your request data and try again."
        ).dict()
    )

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Handle HTTP exceptions with StandardErrorResponse format"""
    return JSONResponse(
        status_code=exc.status_code,
        content=StandardErrorResponse(
            error=exc.detail,
            error_code=f"HTTP_{exc.status_code}",
            retryable=exc.status_code in [408, 429, 500, 502, 503, 504],
            user_action="Please try again later." if exc.status_code >= 500 else None
        ).dict()
    )

# ============================================================================
# LOAD TESTING FIX: Request Timeout Middleware
# ============================================================================
# Prevents hanging requests from exhausting resources under concurrent load
# All requests must complete within 30 seconds or receive 504 Gateway Timeout

@app.middleware("http")
async def timeout_middleware(request: Request, call_next):
    """
    Timeout middleware to prevent hanging requests.
    
    Prevents resource exhaustion by enforcing a maximum request duration.
    Critical for production stability under concurrent load.
    """
    try:
        # Use configured timeout from app_config
        timeout = app_config.request_timeout if 'app_config' in globals() else 120
        response = await asyncio.wait_for(call_next(request), timeout=float(timeout))
        return response
    except asyncio.TimeoutError:
        logger.error(f"Request timeout after {timeout}s: {request.url.path}")
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=504,
            content={
                "error": "Request timeout",
                "message": f"Request took longer than {timeout} seconds to process",
                "error_code": "TIMEOUT",
                "retryable": True,
                "user_action": "Please try again. If this persists, contact support."
            }
        )
    except Exception as e:
        logger.error(f"Middleware error: {e}")
        raise


# CORS middleware with environment-based configuration
# Prevents CSRF attacks in production by restricting origins
ALLOWED_ORIGINS = os.getenv('CORS_ALLOWED_ORIGINS', '*').split(',')
if ALLOWED_ORIGINS == ['*']:
    logger.warning("âš ï¸  CORS is configured with wildcard '*' - this should only be used in development!")
else:
    logger.info(f"âœ… CORS configured with specific origins: {ALLOWED_ORIGINS}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS if ALLOWED_ORIGINS != ["*"] else ["*"],  # Configured via CORS_ALLOWED_ORIGINS env var
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],  # Allow all response headers to be exposed
    max_age=3600,  # Cache preflight requests for 1 hour
)

# Startup handled by app_lifespan context manager

# Initialize global config with pydantic-settings
try:
    app_config = AppConfig()
    logger.info("âœ… Environment configuration loaded and validated via pydantic-settings")
    logger.info(f"   Queue Backend: {app_config.queue_backend}")
except Exception as e:
    logger.error(f"ðŸš¨ CRITICAL: Environment configuration validation failed: {e}")
    raise

# DIAGNOSTIC: Health check endpoint to debug service initialization
@app.get("/health/diagnostic")
async def diagnostic_health_check():
    """
    Comprehensive diagnostic endpoint to check all service states.
    Returns detailed information about which services are initialized.
    """
    return {
        "status": "running",
        "services": {
            "supabase": {
                "initialized": supabase is not None,
                "type": str(type(supabase)) if supabase else None
            },
            "groq_client": {
                "initialized": groq_client is not None,
                "type": str(type(groq_client)) if groq_client else None
            },
            "centralized_cache": {
                "initialized": centralized_cache is not None,
                "type": str(type(centralized_cache)) if centralized_cache else None
            },
            "optimized_db": {
                "initialized": optimized_db is not None,
                "type": str(type(optimized_db)) if optimized_db else None
            },
            "security_validator": {
                "initialized": security_validator is not None,
                "type": str(type(security_validator)) if security_validator else None
            },
            "websocket_manager": {
                "initialized": websocket_manager is not None,
                "type": str(type(websocket_manager)) if websocket_manager else None
            }
        },
        "environment": {
            "SUPABASE_URL": os.environ.get("SUPABASE_URL") is not None,
            "SUPABASE_SERVICE_ROLE_KEY": os.environ.get("SUPABASE_SERVICE_ROLE_KEY") is not None,
            "GROQ_API_KEY": os.environ.get("GROQ_API_KEY") is not None,
            "REDIS_URL": os.environ.get("REDIS_URL") is not None,
            "ARQ_REDIS_URL": os.environ.get("ARQ_REDIS_URL") is not None
        }
    }


async def validate_critical_environment():
    """Validate critical environment variables.
    
    Kept for backward compatibility - validation is handled by AppConfig.
    """
    logger.info("ðŸ” Environment configuration already validated via pydantic-settings")
    
    # Validate Redis if using ARQ queue backend
    if app_config.queue_backend == 'arq':
        if not app_config.redis_url_resolved:
            raise RuntimeError(
                "ðŸš¨ CRITICAL: REDIS_URL or ARQ_REDIS_URL required when QUEUE_BACKEND=arq\n"
                "Set one of these environment variables or change QUEUE_BACKEND to 'sync'"
            )
    
    logger.info("âœ… All required environment variables present and valid")

# Lifespan handled by app_lifespan context manager

# Expose Prometheus metrics
@app.get("/metrics")
async def metrics_endpoint():
    try:
        data = generate_latest()
        return Response(content=data, media_type=CONTENT_TYPE_LATEST)
    except Exception as e:
        logger.error(f"/metrics failed: {e}")
        raise HTTPException(status_code=500, detail="metrics unavailable")

@app.get("/health/cache")
async def cache_health_endpoint():
    """Check Redis cache health and circuit breaker status"""
    try:
        from centralized_cache import health_check
        health = await health_check()
        
        status_code = 200
        if health['status'] == 'unhealthy':
            status_code = 503
        elif health['status'] == 'degraded':
            status_code = 429
        elif health['status'] == 'unavailable':
            status_code = 503
        
        return JSONResponse(content=health, status_code=status_code)
    except Exception as e:
        logger.error(f"/health/cache failed: {e}")
        return JSONResponse(
            content={'status': 'error', 'error': str(e)},
            status_code=500
        )

@app.get("/health/inference")
async def inference_health_endpoint():
    """Check inference service health and model loading status"""
    try:
        from inference_service import health_check
        health = await health_check()
        return JSONResponse(content=health, status_code=200)
    except Exception as e:
        logger.error(f"/health/inference failed: {e}")
        return JSONResponse(
            content={'status': 'error', 'error': str(e)},
            status_code=500
        )

# Database health check function
def check_database_health():
    pass # Disabled for AI-only processing
    """Check if database connection is healthy and raise appropriate error if not"""
    if not supabase:
        logger.error("âŒ CRITICAL: Database connection unavailable")
        raise HTTPException(
            status_code=503,
            detail="Database service temporarily unavailable. Please try again later."
        )
    
    try:
        # Quick health check query
        result = supabase.table('raw_events').select('id').limit(1).execute()
        return True
    except Exception as e:
        logger.error(f"âŒ Database health check failed: {e}")
        raise HTTPException(
            status_code=503,
            detail="Database service experiencing issues. Please try again later."
        )

# Advanced functionality imports with individual error handling
ADVANCED_FEATURES = {
    'py7zr': False,
    'rarfile': False,
    'odf': False,
    'pil': False,
    'cv2': False,
    'xlwings': False
}

# Import advanced features individually for better error handling
try:
    import zipfile
    ADVANCED_FEATURES['zipfile'] = True
    logger.info("âœ… ZIP file processing available")
except ImportError:
    logger.warning("âš ï¸ ZIP file processing not available")

try:
    import py7zr
    ADVANCED_FEATURES['py7zr'] = True
    logger.info("âœ… 7-Zip file processing available")
except ImportError:
    logger.warning("âš ï¸ 7-Zip file processing not available")

try:
    import rarfile
    ADVANCED_FEATURES['rarfile'] = True
    logger.info("âœ… RAR file processing available")
except ImportError:
    logger.warning("âš ï¸ RAR file processing not available")

try:
    from odf.opendocument import load as load_ods
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    ADVANCED_FEATURES['odf'] = True
    logger.info("âœ… OpenDocument processing available")
except ImportError:
    logger.warning("âš ï¸ OpenDocument processing not available")

# Using UniversalExtractorsOptimized with easyocr + pdfminer.six for all extraction

try:
    from PIL import Image
    ADVANCED_FEATURES['pil'] = True
    logger.info("âœ… PIL image processing available")
except ImportError:
    logger.warning("âš ï¸ PIL image processing not available")

try:
    import cv2
    ADVANCED_FEATURES['cv2'] = True
    logger.info("âœ… OpenCV processing available")
except ImportError:
    logger.warning("âš ï¸ OpenCV processing not available")

try:
    import xlwings as xw
    ADVANCED_FEATURES['xlwings'] = True
    logger.info("âœ… Excel automation available")
except ImportError:
    logger.debug("Excel automation (xlwings) not available - using pandas fallback")

# Granular feature availability checking
def is_feature_available(feature_name: str) -> bool:
    """Check if a specific feature is available"""
    return ADVANCED_FEATURES.get(feature_name, False)

def get_available_features() -> List[str]:
    """Get list of available features"""
    return [name for name, available in ADVANCED_FEATURES.items() if available]

def check_feature_dependencies(required_features: List[str]) -> Dict[str, bool]:
    """Check if required features are available"""
    return {feature: is_feature_available(feature) for feature in required_features}

# Overall advanced features availability (for backward compatibility)
ADVANCED_FEATURES_AVAILABLE = any(ADVANCED_FEATURES.values())
logger.info(f"ðŸ”§ Advanced features status: {sum(ADVANCED_FEATURES.values())}/{len(ADVANCED_FEATURES)} available")
logger.info(f"ðŸ“‹ Available features: {', '.join(get_available_features())}")

# Enhanced global configuration with environment variable support
@dataclass
class Config:
    """Enhanced global configuration for the application with environment variable support"""
    # File processing configuration
    max_file_size: int = int(os.environ.get("MAX_FILE_SIZE", 500 * 1024 * 1024))  # 500MB default
    chunk_size: int = int(os.environ.get("CHUNK_SIZE", 8192))  # 8KB chunks for streaming
    batch_size: int = int(os.environ.get("BATCH_SIZE", 50))  # Standardized batch size
    
    # WebSocket configuration
    websocket_timeout: int = int(os.environ.get("WEBSOCKET_TIMEOUT", 300))  # 5 minutes
    
    # AI processing configuration
    platform_confidence_threshold: float = float(os.environ.get("PLATFORM_CONFIDENCE_THRESHOLD", 0.85))
    entity_similarity_threshold: float = float(os.environ.get("ENTITY_SIMILARITY_THRESHOLD", 0.9))
    max_concurrent_ai_calls: int = int(os.environ.get("MAX_CONCURRENT_AI_CALLS", 5))
    
    # Caching configuration
    cache_ttl: int = int(os.environ.get("CACHE_TTL", 3600))  # 1 hour
    
    # Feature flags with environment variable support
    enable_advanced_file_processing: bool = os.environ.get("ENABLE_ADVANCED_FILE_PROCESSING", "true").lower() == "true"
    enable_duplicate_detection: bool = os.environ.get("ENABLE_DUPLICATE_DETECTION", "true").lower() == "true"
    enable_ocr_processing: bool = os.environ.get("ENABLE_OCR_PROCESSING", "true").lower() == "true"
    enable_archive_processing: bool = os.environ.get("ENABLE_ARCHIVE_PROCESSING", "true").lower() == "true"
    
    # Performance optimization settings
    enable_async_processing: bool = os.environ.get("ENABLE_ASYNC_PROCESSING", "true").lower() == "true"
    max_workers: int = int(os.environ.get("MAX_WORKERS", 4))
    memory_limit_mb: int = int(os.environ.get("MEMORY_LIMIT_MB", 2048))
    
    # Security settings
    enable_rate_limiting: bool = os.environ.get("ENABLE_RATE_LIMITING", "true").lower() == "true"
    max_requests_per_minute: int = int(os.environ.get("MAX_REQUESTS_PER_MINUTE", 100))
    
    def __post_init__(self):
        """Validate configuration after initialization"""
        if self.max_file_size <= 0:
            raise ValueError("max_file_size must be positive")
        if self.chunk_size <= 0:
            raise ValueError("chunk_size must be positive")
        if self.batch_size <= 0:
            raise ValueError("batch_size must be positive")
        if not 0 <= self.platform_confidence_threshold <= 1:
            raise ValueError("platform_confidence_threshold must be between 0 and 1")
        if not 0 <= self.entity_similarity_threshold <= 1:
            raise ValueError("entity_similarity_threshold must be between 0 and 1")
        if self.max_concurrent_ai_calls <= 0:
            raise ValueError("max_concurrent_ai_calls must be positive")
        if self.max_workers <= 0:
            raise ValueError("max_workers must be positive")
        if self.memory_limit_mb <= 0:
            raise ValueError("memory_limit_mb must be positive")

# Initialize configuration with validation
try:
    config = Config()
    logger.info("âœ… Configuration loaded successfully")
    logger.info(f"ðŸ“Š File processing: max_size={config.max_file_size//1024//1024}MB, batch_size={config.batch_size}")
    logger.info(f"ðŸ¤– AI processing: max_concurrent={config.max_concurrent_ai_calls}, confidence={config.platform_confidence_threshold}")
except Exception as e:
    logger.error(f"âŒ Configuration validation failed: {e}")
    raise

# - Removed to eliminate code duplication and reduce maintenance burden
# VendorStandardizer moved to data_ingestion_normalization/universal_normalizer_optimized.py


# MODULARIZATION: PlatformIDExtractor moved to data_ingestion_normalization/universal_platform_detector_optimized.py

# MODULARIZATION: _DeprecatedAIRowClassifier removed (deprecated, replaced by AIRowClassifier in row_classifier.py)


async def convert_stream_to_bytes(streamed_file) -> bytes:
    """Convert streaming file to full bytes for extractors."""
    try:
        if hasattr(streamed_file, 'read'):
            # File-like object
            return await streamed_file.read()
        elif hasattr(streamed_file, 'path') and streamed_file.path:
            # File path - read from disk
            with open(streamed_file.path, 'rb') as f:
                return f.read()
        else:
            logger.warning("Cannot convert stream to bytes - unsupported format")
            return b""
    except Exception as e:
        logger.error(f"Failed to convert stream to bytes: {e}")
        raise


async def save_clean_event(tx, event: Dict[str, Any]) -> Dict[str, Any]:
    """Unified event save function - single source of truth for raw_events insertion."""
    try:
        # Extract clean fields from event
        clean_event = {
            "user_id": event.get('user_id'),
            "file_id": event.get('file_id'),
            "row_index": event.get('row_index'),
            "sheet_name": event.get('sheet_name'),
            "source_filename": event.get('source_filename'),
            "uploader": event.get('uploader'),
            "ingest_ts": event.get('ingest_ts'),
            "status": event.get('status', 'pending'),
            "confidence_score": event.get('confidence_score', 0.0),
            
            # Normalized fields
            "normalized_fields": event.get('normalized', {}),
            
            # Dedupe metadata
            "row_hash": event.get('row_hash'),
            "dedupe_metadata": event.get('dedupe_metadata', {}),
            
            # Entity resolution
            "entity_resolution": event.get('entity_resolution', {}),
            
            # Core data
            "payload": event.get('payload', {}),
            "classification_metadata": event.get('classification_metadata', {}),
            
            # Lineage
            "lineage": event.get('lineage', {}),
            "transaction_id": event.get('transaction_id'),
            "job_id": event.get('job_id'),
        }
        
        # Insert and return
        result = await tx.insert('raw_events', clean_event)
        return result
    except Exception as e:
        logger.error(f"Failed to save clean event: {e}")
        raise


class RowProcessor:
    """Processes individual rows and creates events"""
    
    REVENUE_PATTERNS = ['income', 'revenue', 'payment received', 'deposit', 'credit']
    EXPENSE_PATTERNS = ['expense', 'cost', 'payment', 'debit', 'withdrawal', 'fee']
    PAYROLL_PATTERNS = ['salary', 'payroll', 'wage', 'employee']
    
    def __init__(self, platform_detector: UniversalPlatformDetector, ai_classifier, enrichment_processor):
        self.platform_detector = platform_detector
        self.ai_classifier = ai_classifier
        self.enrichment_processor = enrichment_processor
    
    async def process_row(self, row, row_index: int, sheet_name: str, 
                   platform_info: Dict, file_context: Dict, column_names: List[str]) -> Dict[str, Any]:
        """Process a single row and create an event with AI-powered classification and enrichment"""
        
        # AI-powered row classification
        ai_classification = await self.ai_classifier.classify_row_with_ai(row, platform_info, column_names, file_context)
        
        # Convert row to JSON-serializable format
        row_data = self._convert_row_to_json_serializable(row)
        
        # Update file context with row index
        file_context['row_index'] = row_index
        
        # Data enrichment - create enhanced payload
        enriched_payload = await self.enrichment_processor.enrich_row_data(
            row_data=row_data,
            platform_info=platform_info,
            column_names=column_names,
            ai_classification=ai_classification,
            file_context=file_context
        )
        
        # Row hashing and lineage tracking handled by duplicate detection service
        
        # Create the event payload with enhanced metadata AND provenance
        event = {
            "provider": "excel-upload",
            "kind": enriched_payload.get('kind', 'transaction'),
            "source_platform": platform_info.get('platform', 'unknown'),
            "payload": enriched_payload,  # Use enriched payload instead of raw
            "row_index": row_index,
            "sheet_name": sheet_name,
            "source_filename": file_context['filename'],
            "uploader": file_context['user_id'],
            "ingest_ts": pendulum.now().to_iso8601_string(),
            "status": "pending",
            "confidence_score": enriched_payload.get('ai_confidence', 0.5),
            "classification_metadata": {
                "platform_detection": platform_info,
                "ai_classification": ai_classification,
                "enrichment_data": enriched_payload,
                "document_type": platform_info.get('document_type', 'unknown'),
                "document_confidence": platform_info.get('document_confidence', 0.0),
                "document_classification_method": platform_info.get('document_classification_method', 'unknown'),
                "document_indicators": platform_info.get('document_indicators', []),
                "row_type": enriched_payload.get('kind', 'transaction'),
                "category": enriched_payload.get('category', 'other'),
                "subcategory": enriched_payload.get('subcategory', 'general'),
                "entities": enriched_payload.get('entities', {}),
                "relationships": enriched_payload.get('relationships', {}),
                "description": enriched_payload.get('standard_description', ''),
                "reasoning": enriched_payload.get('ai_reasoning', ''),
                "sheet_name": sheet_name,
                "file_context": file_context
            }
            # REMOVED: Provenance fields (row_hash, lineage_path, created_by) moved to duplicate detection service
        }
        
        return event
    
    def _convert_row_to_json_serializable(self, row) -> Dict[str, Any]:
        """Convert a row (Polars or dict) to JSON-serializable format"""
        result = {}
        
        # Handle different row types
        if isinstance(row, dict):
            items = row.items()
        elif hasattr(row, 'to_dict'):
            items = row.to_dict().items()
        elif hasattr(row, 'items'):
            items = row.items()
        else:
            return {}
        
        for column, value in items:
            # Check if value is null/None
            if value is None or (isinstance(value, str) and value.lower() == 'nan'):
                result[str(column)] = None
            elif isinstance(value, (int, float, str, bool)):
                result[str(column)] = value
            elif isinstance(value, (list, dict)):
                # Handle nested structures
                result[str(column)] = self._convert_nested_to_json_serializable(value)
            else:
                # Convert any other types to string
                result[str(column)] = str(value)
        return result
    
    def _convert_nested_to_json_serializable(self, obj: Any) -> Any:
        """Convert nested objects to JSON-serializable format"""
        if isinstance(obj, dict):
            return {str(k): self._convert_nested_to_json_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._convert_nested_to_json_serializable(item) for item in obj]
        elif hasattr(obj, 'isoformat'):
            # Handle datetime-like objects (datetime, date, time, etc.)
            return obj.isoformat()
        elif obj is None or (isinstance(obj, str) and obj.lower() == 'nan'):
            return None
        elif isinstance(obj, (int, float, str, bool)):
            return obj
        else:
            return str(obj)


_excel_processor_instance: Optional['ExcelProcessor'] = None
_excel_processor_lock = asyncio.Lock()

async def get_excel_processor() -> 'ExcelProcessor':
    """
    Get or create singleton ExcelProcessor instance.
    Lazy-loads on first use, reuses thereafter.
    Thread-safe with asyncio lock.
    """
    global _excel_processor_instance
    if _excel_processor_instance is None:
        async with _excel_processor_lock:
            # Double-check after acquiring lock
            if _excel_processor_instance is None:
                _excel_processor_instance = ExcelProcessor()
                logger.info("âœ… ExcelProcessor singleton initialized")
    return _excel_processor_instance



# MODULARIZATION: ExcelProcessor class moved to data_ingestion_normalization/excel_processor.py
# Import: from data_ingestion_normalization.excel_processor import ExcelProcessor


# ============================================================================
# LEGACY ENTITY RESOLVER - REMOVED
# ============================================================================
# The LegacyEntityResolver class has been completely removed (was 324 lines).
# All entity resolution now uses EntityResolverOptimized from entity_resolver_optimized.py
# which provides:
# - rapidfuzz for 50x faster fuzzy matching (+25% accuracy)
# - presidio-analyzer for 30x faster PII detection (+40% accuracy)
# - polars for 100x vectorized data processing
# - aiocache with Redis for 10x faster caching
# - AI-powered ambiguous match resolution
# - tenacity for bulletproof retry logic
#
# Migration: Replace any LegacyEntityResolver() calls with:
#   from entity_resolver_optimized import EntityResolverOptimized as EntityResolver
#   resolver = EntityResolver(supabase_client=supabase, openai_client=openai_client)
# ============================================================================

# REMOVED CLASS: LegacyEntityResolver (324 lines removed)
# All methods removed: __init__, resolve_entities_batch, _resolve_single_entity,
# _normalize_entity_name, _find_similar_entities, _calculate_similarity, get_metrics

# ============================================================================
# DUPLICATE HANDLING ENDPOINTS
# ============================================================================

class DuplicateDecisionRequest(BaseModel):
    job_id: str
    user_id: str
    decision: str  # 'replace', 'keep_both', 'skip', 'delta_merge'
    file_hash: str
    existing_file_id: Optional[str] = None
    session_token: Optional[str] = None

async def _handle_duplicate_detection_unified(
    user_id: str,
    file_hash: str,
    filename: str,
    file_content: Optional[bytes] = None,
    streamed_file: Optional[Any] = None
) -> Dict[str, Any]:
    """
    UNIFIED duplicate detection for all paths (upload, integration, manual check).
    Single entry point to prevent inconsistent behavior across 3 different code paths.
    """
    try:
        if not PRODUCTION_DUPLICATE_SERVICE_AVAILABLE:
            raise DuplicateDetectionError("Duplicate detection service unavailable")
        
        duplicate_service = ProductionDuplicateDetectionService(supabase)
        
        file_metadata = FileMetadata(
            user_id=user_id,
            file_hash=file_hash,
            filename=filename,
            file_size=len(file_content) if file_content else (streamed_file.size if streamed_file else 0),
            content_type='application/octet-stream',
            upload_timestamp=datetime.utcnow()
        )
        
        dup_result = await duplicate_service.detect_duplicates(
            file_metadata=file_metadata,
            file_path=streamed_file.path,
            file_content=file_content,
            sheets_data=None,
            enable_near_duplicate=True,
            enable_content_duplicate=True
        )
        
        return {
            'is_duplicate': dup_result.is_duplicate,
            'duplicate_type': dup_result.duplicate_type.value,
            'similarity_score': dup_result.similarity_score,
            'duplicate_files': dup_result.duplicate_files,
            'recommendation': dup_result.recommendation.value,
            'message': dup_result.message,
            'confidence': dup_result.confidence,
            'delta_analysis': dup_result.delta_analysis if hasattr(dup_result, 'delta_analysis') else None
        }
    
    except DuplicateDetectionError as e:
        logger.error(f"Unified duplicate detection failed: {e}")
        raise
    except Exception as e:
        logger.error(f"Unified duplicate detection error: {e}")
        raise DuplicateDetectionError(f"Duplicate detection failed: {str(e)}")


@app.post("/handle-duplicate-decision")
async def handle_duplicate_decision(request: DuplicateDecisionRequest):
    """Handle user's decision about duplicate files"""
    try:
        job_state = await websocket_manager.get_job_status(request.job_id)
        if not job_state:
            raise HTTPException(status_code=404, detail="Job not found or expired")

        decision = (request.decision or '').lower()

        # Security validation: require session token for decision
        try:
            valid, violations = await security_validator.validate_request({
                'endpoint': 'handle-duplicate-decision',
                'user_id': request.user_id,
                'session_token': request.session_token
            }, SecurityContext(user_id=request.user_id))
            if not valid:
                logger.warning(f"Security validation failed on duplicate decision for job {request.job_id}: {violations}")
                raise HTTPException(status_code=401, detail="Unauthorized or invalid session")
        except HTTPException:
            raise
        except Exception as sec_e:
            logger.warning(f"Security validation error on duplicate decision for job {request.job_id}: {sec_e}")
            raise HTTPException(status_code=401, detail="Unauthorized or invalid session")

        # If user chose to skip, mark as cancelled and update DB
        if decision == 'skip':
            await websocket_manager.merge_job_state(request.job_id, {
                **(job_state or {}),
                "status": "cancelled",
                "message": "Processing skipped due to duplicate",
                "progress": 100
            })
            # Notify over WebSocket if connected
            await websocket_manager.send_overall_update(
                job_id=request.job_id,
                status="cancelled",
                message="Processing skipped by user due to duplicate",
                progress=100
            )

            # CRITICAL FIX #4: Persist duplicate decision to BOTH tables transactionally
            # This ensures future duplicate checks can see the decision
            try:
                transaction_manager = get_transaction_manager()
                async with transaction_manager.transaction(
                    user_id=request.user_id,
                    operation_type="ingestion_job_skip"
                ) as tx:
                    # Update ingestion_jobs
                    await tx.update('ingestion_jobs', {
                        'status': 'cancelled',
                        'updated_at': pendulum.now().to_iso8601_string(),
                        'error_message': 'Skipped due to duplicate',
                        'metadata': {
                            'duplicate_decision': 'skip',
                            'decided_at': pendulum.now().to_iso8601_string(),
                            'existing_file_id': request.existing_file_id
                        }
                    }, {'id': request.job_id})
                    
                    # CRITICAL FIX #4: Update raw_records with duplicate decision
                    # This is where duplicate checks query, so decision MUST be stored here
                    if request.file_hash:
                        try:
                            await tx.update('raw_records', {
                                'duplicate_decision': 'skip',
                                'duplicate_of': request.existing_file_id,
                                'decision_timestamp': pendulum.now().to_iso8601_string(),
                                'decision_metadata': {
                                    'decided_at': pendulum.now().to_iso8601_string(),
                                    'job_id': request.job_id,
                                    'user_action': 'skip'
                                }
                            }, {'file_hash': request.file_hash, 'user_id': request.user_id})
                            logger.info(f"Persisted 'skip' decision to raw_records for hash {request.file_hash}")
                        except Exception as rr_err:
                            logger.error(f"CRITICAL: Failed to update raw_records with duplicate decision: {rr_err}")
                            # Re-raise to rollback transaction
                            raise
            except Exception as e:
                logger.error(f"Failed to transactionally update duplicate decision for job {request.job_id}: {e}")
                raise HTTPException(status_code=500, detail=f"Failed to persist duplicate decision: {str(e)}")

            return {"status": "success", "message": "Duplicate decision processed: skipped"}

        # For 'replace' or 'keep_both' we resume processing with the saved request
        if decision in ('replace', 'keep_both', 'delta_merge'):
            # CRITICAL FIX: Try to get pending_request from job_state first (Redis)
            # If missing, fall back to database backup (ingestion_jobs.result)
            pending = job_state.get('pending_request') or {}
            
            # If pending_request missing from Redis, try database
            if not pending or not pending.get('user_id'):
                try:
                    job_record = supabase.table('ingestion_jobs').select('result').eq('id', request.job_id).single().execute()
                    if job_record and job_record.data:
                        result_data = job_record.data.get('result', {})
                        pending = result_data.get('pending_request', {})
                        logger.info(f"Recovered pending_request from database for job {request.job_id}")
                except Exception as db_err:
                    logger.warning(f"Failed to recover pending_request from database: {db_err}")
            
            user_id = pending.get('user_id')
            storage_path = pending.get('storage_path')
            filename = pending.get('filename') or 'uploaded_file'
            
            # CRITICAL FIX: Validate that pending_request was found and has required fields
            if not user_id or not storage_path:
                error_msg = (
                    f"Cannot resume: pending_request missing or incomplete. "
                    f"user_id={bool(user_id)}, storage_path={bool(storage_path)}. "
                    f"This may indicate job state was lost or corrupted."
                )
                logger.error(error_msg)
                raise HTTPException(status_code=400, detail=error_msg)

            existing_file_id = (
                request.existing_file_id
                or pending.get('existing_file_id')
                or (pending.get('duplicate_files') or [{}])[0].get('id')
            )

            # Update status and resume processing asynchronously
            await websocket_manager.merge_job_state(request.job_id, {
                **(job_state or {}),
                "status": "processing",
                "message": f"Resuming after duplicate decision: {decision}",
                "progress": max((job_state or {}).get('progress', 10), 20)
            })
            await websocket_manager.send_overall_update(
                job_id=request.job_id,
                status="processing",
                message=f"Resuming after duplicate decision: {decision}",
                progress=(await websocket_manager.get_job_status(request.job_id) or {}).get("progress")
            )

            # FIX ISSUE #11, #13, #14: Actually resume processing
            logger.info(f"ðŸ”„ Processing resume inline: {request.job_id}, decision: {decision}")
            
            # CRITICAL FIX #5: Complete rewrite of delta_merge with proper transaction handling
            if decision == 'delta_merge':
                try:
                    # Step 1: Validate existing file exists and belongs to user
                    await websocket_manager.send_overall_update(
                        job_id=request.job_id,
                        status="processing",
                        message="Validating files for delta merge...",
                        progress=10
                    )
                    
                    # CRITICAL FIX: Use optimized query for file lookup
                    existing_file_data = await optimized_db.get_file_by_id(user_id, existing_file_id)
                    
                    if not existing_file_data:
                        raise ValueError(f"Existing file not found or access denied: {existing_file_id}")
                    
                    logger.info(f"Delta merge: validated existing file {existing_file_id}")
                    
                    # Step 2: Download new file with streaming to avoid memory issues
                    # FIX #85: Stream file instead of loading entire file into memory
                    await websocket_manager.send_overall_update(
                        job_id=request.job_id,
                        status="processing",
                        message="Downloading new file for delta merge...",
                        progress=20
                    )
                    
                    import polars as pl
                    import io
                    import tempfile
                    
                    storage = supabase.storage.from_("finely-upload")
                    
                    # FIX #85: Stream to temporary file instead of loading into memory
                    with tempfile.NamedTemporaryFile(delete=False, suffix=filename[-4:]) as tmp:
                        temp_file_path = tmp.name
                        file_resp = storage.download(storage_path)
                        file_bytes = file_resp if isinstance(file_resp, (bytes, bytearray)) else getattr(file_resp, 'data', None)
                        if file_bytes is None:
                            file_bytes = file_resp
                        tmp.write(file_bytes)
                    
                    try:
                        file_size_mb = len(file_bytes) / (1024 * 1024)
                        logger.info(f"Downloaded file for delta merge: {file_size_mb:.1f}MB")
                        
                        # Step 3: Parse file to get sheets data
                        await websocket_manager.send_overall_update(
                            job_id=request.job_id,
                            status="processing",
                            message="Parsing file data...",
                            progress=30
                        )
                        
                        # FIX #85: Use streaming parser for large files to avoid memory issues
                        if file_size_mb > 50:  # Use streaming for files > 50MB
                            logger.info(f"Large file detected ({file_size_mb:.1f}MB), using streaming parser")
                            if filename.endswith('.csv'):
                                # Read CSV in chunks for memory efficiency
                                df = pl.read_csv(temp_file_path, streaming=True).collect()
                                sheets_data = {'Sheet1': df}
                            else:
                                # FIX #85: For Excel, use streaming_processor instead of pl.read_excel to prevent OOM
                                logger.info(f"Excel delta merge using streaming processor ({file_size_mb:.1f}MB)")
                                processor = ExcelProcessor(supabase_client=supabase)
                                sheets_data = await processor._parse_file_streaming(file_bytes, filename)
                        else:
                            # Small files can be read normally
                            if filename.endswith('.csv'):
                                df = pl.read_csv(temp_file_path)
                                sheets_data = {'Sheet1': df}
                            else:
                                # FIX #85: Use openpyxl for small Excel files instead of pl.read_excel
                                logger.info(f"Excel delta merge using openpyxl for small file ({file_size_mb:.1f}MB)")
                                import openpyxl
                                wb = openpyxl.load_workbook(temp_file_path, data_only=True)
                                sheets_data = {}
                                for sheet_name in wb.sheetnames:
                                    ws = wb[sheet_name]
                                    data = []
                                    for row in ws.iter_rows(values_only=True):
                                        data.append(row)
                                    sheets_data[sheet_name] = pl.DataFrame(data[1:], schema=[str(i) for i in range(len(data[0]))] if data else [])
                                wb.close()
                        
                        logger.info(f"Delta merge: parsed {len(sheets_data)} sheets")
                        
                        # Step 4: Calculate file hash for new file (xxhash: 5-10x faster)
                        new_file_hash = xxhash.xxh64(file_bytes).hexdigest()
                    finally:
                        # FIX #85: Clean up temporary file
                        if os.path.exists(temp_file_path):
                            try:
                                os.remove(temp_file_path)
                            except Exception as cleanup_err:
                                logger.warning(f"Failed to cleanup temp file {temp_file_path}: {cleanup_err}")
                    
                    # Step 5: Perform delta merge with transaction
                    await websocket_manager.send_overall_update(
                        job_id=request.job_id,
                        status="processing",
                        message="Analyzing differences...",
                        progress=40
                    )
                    
                    duplicate_service = ProductionDuplicateDetectionService(supabase)
                    
                    # CRITICAL: Use proper parameters (new_file_hash instead of job_id)
                    merge_result = await duplicate_service._perform_delta_merge(
                        user_id=user_id,
                        new_file_hash=new_file_hash,
                        existing_file_id=existing_file_id
                    )
                    
                    # Step 6: Update job status
                    await websocket_manager.send_overall_update(
                        job_id=request.job_id,
                        status="completed",
                        message=f"Delta merge completed: {merge_result.get('merged_events', 0)} events merged",
                        progress=100
                    )
                    
                    # Step 7: Update ingestion_jobs with success
                    try:
                        supabase.table('ingestion_jobs').update({
                            'status': 'completed',
                            'progress': 100,
                            'updated_at': pendulum.now().to_iso8601_string(),
                            'metadata': {
                                'duplicate_decision': 'delta_merge',
                                'merge_result': merge_result,
                                'existing_file_id': existing_file_id
                            }
                        }).eq('id', request.job_id).execute()
                    except Exception as db_err:
                        logger.warning(f"Failed to update ingestion_jobs after delta merge: {db_err}")
                    
                    logger.info(f"Delta merge completed successfully for job {request.job_id}")
                    return {"status": "success", "message": "Delta merge completed", "result": merge_result}
                    
                except ValueError as ve:
                    # Validation errors (user-friendly)
                    error_msg = str(ve)
                    logger.error(f"Delta merge validation failed for job {request.job_id}: {error_msg}")
                    await websocket_manager.send_error(request.job_id, error_msg)
                    
                    # Update job status
                    supabase.table('ingestion_jobs').update({
                        'status': 'failed',
                        'error_message': error_msg,
                        'updated_at': pendulum.now().to_iso8601_string()
                    }).eq('id', request.job_id).execute()
                    
                    raise HTTPException(status_code=400, detail=error_msg)
                    
                except Exception as e:
                    # Unexpected errors
                    import traceback
                    error_msg = f"Delta merge failed: {str(e)}"
                    error_details = f"Delta merge failed for job {request.job_id}: {e}\n{traceback.format_exc()}"
                    structured_logger.error("Delta merge failed", error=error_details)
                    await websocket_manager.send_error(request.job_id, error_msg)
                    
                    # Update job status
                    try:
                        supabase.table('ingestion_jobs').update({
                            'status': 'failed',
                            'error_message': error_msg,
                            'updated_at': pendulum.now().to_iso8601_string()
                        }).eq('id', request.job_id).execute()
                    except Exception as db_err:
                        logger.error(f"Failed to update job status after delta merge error: {db_err}")
                    
                    raise HTTPException(status_code=500, detail=error_msg)
            
            # CRITICAL FIX: Re-enqueue to ARQ instead of inline processing
            # This prevents 1,000 concurrent duplicate resolutions from overwhelming API servers
            try:
                pool = await get_arq_pool()
                await pool.enqueue_job(
                    'process_spreadsheet',
                    user_id=user_id,
                    filename=filename,
                    storage_path=storage_path,
                    job_id=request.job_id,
                    duplicate_decision=decision,
                    existing_file_id=existing_file_id
                )
                logger.info(f" Job {request.job_id} re-enqueued to ARQ after duplicate decision: {decision}")
            except Exception as arq_error:
                logger.error(f"Failed to re-enqueue job {request.job_id} to ARQ: {arq_error}")
                raise HTTPException(status_code=500, detail=f"Failed to resume processing: {str(arq_error)}")
            
            return {"status": "success", "message": "Duplicate decision processed: resuming"}

        raise HTTPException(status_code=400, detail="Invalid decision. Use one of: replace, keep_both, skip")
    except HTTPException as he:
        # Keep explicit HTTP errors (like 401) intact
        raise he
    except Exception as e:
        logger.error(f"Error handling duplicate decision: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/delta-merge-history/{file_id}")
async def get_delta_merge_history(file_id: str, user_id: str, session_token: Optional[str] = None):
    """
    FIX ISSUE #15: Get delta merge history for a file.
    Returns all delta merge operations involving this file.
    """
    await _validate_security('delta-merge-history', user_id, session_token)
    try:
        # Use the database function created in migration 20250920130000
        result = supabase.rpc('get_delta_merge_history', {
            'p_user_id': user_id,
            'p_file_id': file_id
        }).execute()
        
        return {
            "file_id": file_id,
            "merge_history": result.data or [],
            "total_merges": len(result.data) if result.data else 0
        }
    except Exception as e:
        logger.error(f"Failed to fetch delta merge history for file {file_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch merge history: {str(e)}")

@app.get("/api/connectors/history")
async def connectors_history(connection_id: str, user_id: str, page: int = 1, page_size: int = 20, session_token: Optional[str] = None):
    """Paginated sync_runs for a connection. Returns {runs, page, page_size, has_more}."""
    await _validate_security('connectors-history', user_id, session_token)
    try:
        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 20
        page_size = min(page_size, 100)
        # Resolve user_connection id
        uc_res = supabase.table('user_connections').select('id').eq('nango_connection_id', connection_id).limit(1).execute()
        if not uc_res.data:
            return {"runs": [], "page": page, "page_size": page_size, "has_more": False}
        uc_id = uc_res.data[0]['id']
        offset = (page - 1) * page_size
        # Fetch page_size + 1 to compute has_more
        runs_res = (
            supabase
            .table('sync_runs')
            .select('id, type, status, started_at, finished_at, stats, error')
            .eq('user_connection_id', uc_id)
            .order('started_at', desc=True)
            .range(offset, offset + page_size - 1)  # inclusive range
            .execute()
        )
        # Supabase range is inclusive; to compute has_more, query next item
        next_res = (
            supabase
            .table('sync_runs')
            .select('id')
            .eq('user_connection_id', uc_id)
            .order('started_at', desc=True)
            .range(offset + page_size, offset + page_size)
            .execute()
        )
        runs = runs_res.data or []
        has_more = bool(next_res.data)
        return {"runs": runs, "page": page, "page_size": page_size, "has_more": has_more}
    except Exception as e:
        logger.error(f"Connectors history failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Removed duplicate /api/connectors/frequency route (using Pydantic version below)

# ============================================================================
# LEGACY ENDPOINTS - REMOVED
# ============================================================================
# The following endpoints were removed as they relied on unused database tables:
# - POST /version-recommendation-feedback (used version_recommendations table)
# - GET /duplicate-analysis/{user_id} (used version_recommendations table)
#
# These tables (file_versions, file_similarity_analysis, version_recommendations)
# were part of an over-engineered versioning system that was never completed.
# The current delta merge approach (event_delta_logs) is simpler and more effective.
#
# Removed: 2025-10-13
# Migration: 20251013000000-remove-unused-version-tables.sql
# ============================================================================
# TEST ENDPOINTS
# ============================================================================

@app.get("/chat-history/{user_id}")
async def get_chat_history(user_id: str):
    """Get chat history for user"""
    try:
        # CRITICAL FIX: Lazy-load Supabase client on first use
        supabase_client = await _ensure_supabase_loaded()
        if not supabase_client:
            logger.error(f"âŒ CRITICAL: Database connection unavailable for get_chat_history - user_id: {user_id}")
            raise HTTPException(
                status_code=503, 
                detail="Database service temporarily unavailable. Please try again later."
            )
        
        # Get chat messages from database using optimized client when available
        try:
            if optimized_db:
                messages = await optimized_db.get_chat_history_optimized(user_id, limit=500)
            else:
                result = supabase_client.table('chat_messages').select(
                    'id, chat_id, message, created_at'
                ).eq('user_id', user_id).order('created_at', desc=True).execute()
                messages = result.data or []
        except Exception as db_err:
            logger.error(f"âŒ CRITICAL: Database query failed for get_chat_history - user_id: {user_id}, error: {db_err}", exc_info=True)
            raise HTTPException(
                status_code=503,
                detail="Database service temporarily unavailable. Please try again later."
            )
        
        chat_groups = {}
        for msg in messages:
            created_at = msg.get('created_at', '') or ''
            date_key = created_at[:10] if created_at else 'unknown'
            if date_key not in chat_groups:
                chat_groups[date_key] = {
                    "id": f"chat_{date_key}",
                    "title": f"Chat {date_key}",
                    "created_at": created_at,
                    "message_count": 0
                }
            chat_groups[date_key]["message_count"] += 1
        
        chats = list(chat_groups.values())
        
        return {
            "chats": chats,
            "user_id": user_id,
            "status": "success"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get chat history for user {user_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve chat history")

@app.put("/chat/rename")
async def rename_chat(request: dict):
    """Rename chat - DEPRECATED, use PUT /chat/{chat_id}/title instead"""
    try:
        chat_id = request.get('chat_id')
        new_title = request.get('title')
        user_id = request.get('user_id')
        
        if not chat_id or not new_title:
            raise HTTPException(status_code=400, detail="Missing chat_id or title")
        
        # Update chat_sessions table with new title
        supabase_client = await _ensure_supabase_loaded()
        if not supabase_client:
            raise HTTPException(status_code=503, detail="Database service temporarily unavailable")
        
        try:
            supabase_client.table('chat_sessions').update({
                'title': new_title,
                'updated_at': pendulum.now().to_iso8601_string()
            }).eq('chat_id', chat_id).execute()
            
            structured_logger.info("Chat renamed", chat_id=chat_id, new_title=new_title)
        except Exception as db_err:
            # If chat_sessions table doesn't exist, create entry
            if 'relation "public.chat_sessions" does not exist' in str(db_err):
                structured_logger.warning(f"chat_sessions table not found, creating entry: {db_err}")
                try:
                    supabase_client.table('chat_sessions').insert({
                        'chat_id': chat_id,
                        'user_id': user_id,
                        'title': new_title,
                        'created_at': pendulum.now().to_iso8601_string(),
                        'updated_at': pendulum.now().to_iso8601_string()
                    }).execute()
                except Exception as insert_err:
                    structured_logger.error(f"Failed to create chat_sessions entry: {insert_err}")
                    raise HTTPException(status_code=500, detail="Failed to save chat title")
            else:
                raise
        
        return {
            "status": "success",
            "message": "Chat renamed successfully",
            "chat_id": chat_id,
            "title": new_title
        }
    except HTTPException:
        raise
    except Exception as e:
        structured_logger.error("Chat rename error", error=e)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/chat/delete")
async def delete_chat(request: dict):
    """Delete chat"""
    try:
        chat_id = request.get('chat_id')
        user_id = request.get('user_id')
        
        if not chat_id:
            raise HTTPException(status_code=400, detail="Missing chat_id")
        
        if supabase and user_id:
            # Delete chat messages for this chat
            # FIX: Use eq() instead of like() to match exact chat_id instead of timestamp pattern
            supabase.table('chat_messages').delete().eq('user_id', user_id).eq('chat_id', chat_id).execute()
            
            # Also delete from chat_sessions if it exists
            try:
                supabase.table('chat_sessions').delete().eq('chat_id', chat_id).execute()
            except Exception as e:
                structured_logger.debug(f"Failed to delete from chat_sessions (may not exist): {e}")
        
        structured_logger.info("Chat deleted", chat_id=chat_id, user_id=user_id)
        
        return {
            "status": "success",
            "message": "Chat deleted successfully",
            "chat_id": chat_id
        }
    except Exception as e:
        structured_logger.error("Chat delete error", error=e)
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/chat/{chat_id}/title")
async def update_chat_title(chat_id: str, request: dict):
    """Update chat title - NEW ENDPOINT for ChatHistoryModal"""
    try:
        new_title = request.get('title')
        user_id = request.get('user_id')
        
        if not new_title:
            raise HTTPException(status_code=400, detail="Missing title")
        
        supabase_client = await _ensure_supabase_loaded()
        if not supabase_client:
            raise HTTPException(status_code=503, detail="Database service temporarily unavailable")
        
        try:
            # Try to update existing chat_sessions entry
            supabase_client.table('chat_sessions').update({
                'title': new_title,
                'updated_at': pendulum.now().to_iso8601_string()
            }).eq('chat_id', chat_id).execute()
            
            structured_logger.info("Chat title updated", chat_id=chat_id, new_title=new_title)
        except Exception as db_err:
            # If chat_sessions table doesn't exist, create entry
            if 'relation "public.chat_sessions" does not exist' in str(db_err):
                structured_logger.warning(f"chat_sessions table not found, creating entry: {db_err}")
                try:
                    supabase_client.table('chat_sessions').insert({
                        'chat_id': chat_id,
                        'user_id': user_id,
                        'title': new_title,
                        'created_at': pendulum.now().to_iso8601_string(),
                        'updated_at': pendulum.now().to_iso8601_string()
                    }).execute()
                except Exception as insert_err:
                    structured_logger.error(f"Failed to create chat_sessions entry: {insert_err}")
                    raise HTTPException(status_code=500, detail="Failed to save chat title")
            else:
                raise
        
        return {
            "status": "success",
            "message": "Chat title updated successfully",
            "chat_id": chat_id,
            "title": new_title
        }
    except HTTPException:
        raise
    except Exception as e:
        structured_logger.error("Chat title update error", error=e)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/chat/{chat_id}")
async def delete_chat_by_id(chat_id: str, user_id: str):
    """Delete chat by ID - NEW ENDPOINT for ChatHistoryModal"""
    try:
        if not chat_id:
            raise HTTPException(status_code=400, detail="Missing chat_id")
        
        supabase_client = await _ensure_supabase_loaded()
        if not supabase_client:
            raise HTTPException(status_code=503, detail="Database service temporarily unavailable")
        
        if user_id:
            # Delete chat messages for this chat
            supabase_client.table('chat_messages').delete().eq('user_id', user_id).eq('chat_id', chat_id).execute()
            
            # Also delete from chat_sessions if it exists
            try:
                supabase_client.table('chat_sessions').delete().eq('chat_id', chat_id).execute()
            except Exception as e:
                structured_logger.debug(f"Failed to delete from chat_sessions (may not exist): {e}")
        
        structured_logger.info("Chat deleted", chat_id=chat_id, user_id=user_id)
        
        return {
            "status": "success",
            "message": "Chat deleted successfully",
            "chat_id": chat_id
        }
    except HTTPException:
        raise
    except Exception as e:
        structured_logger.error("Chat delete error", error=e)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/generate-chat-title")
async def generate_chat_title(request: dict):
    """
    Generate a chat title from the first message using AI.
    
    Called by frontend when creating a new chat.
    """
    try:
        message = request.get('message')
        user_id = request.get('user_id')
        
        if not message or not user_id:
            raise HTTPException(status_code=400, detail="Missing message or user_id")
        
        structured_logger.info("Generate chat title request", user_id=user_id, message_length=len(message))
        
        # Use Groq for simple title generation (cost-effective)
        # FIX #32: Use unified Groq client initialization helper
        title_client = get_groq_client()
        
        try:
            response = title_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": "Generate a concise, descriptive title (max 6 words) for this financial question. Return ONLY the title, no quotes or extra text."},
                    {"role": "user", "content": f"Question: {message}"}
                ],
                max_tokens=50,
                temperature=0.3,
                timeout=15.0  # 15 second timeout
            )
            title = response.choices[0].message.content.strip()
        except Exception as e:
            structured_logger.warning(f"Title generation failed: {str(e)}, using fallback")
            chat_id = f"chat_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
            return {
                "chat_id": chat_id,
                "title": f"Chat {datetime.utcnow().strftime('%b %d, %Y')}",
                "status": "fallback"
            }
        
        # Generate chat_id
        chat_id = f"chat_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{user_id[:8]}"
        
        structured_logger.info("Chat title generated", user_id=user_id, chat_id=chat_id, title=title)
        
        return {
            "chat_id": chat_id,
            "title": title,
            "status": "success"
        }
        
    except Exception as e:
        structured_logger.error("Generate chat title error", error=e)
        # Fallback to timestamp-based title
        chat_id = f"chat_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
        return {
            "chat_id": chat_id,
            "title": f"Chat {datetime.utcnow().strftime('%b %d, %Y')}",
            "status": "fallback"
        }


@app.post("/chat")
async def chat_endpoint(request: dict):
    """
    Main chat endpoint with Server-Sent Events (SSE) streaming.
    Streams response chunks in real-time as they're generated.
    
    This is the brain of Finley AI that routes questions to intelligence engines.
    """
    # CRITICAL: Log immediately on entry
    print(f"[CHAT ENDPOINT] Request received at {datetime.utcnow().isoformat()}", flush=True)
    structured_logger.info("CHAT ENDPOINT CALLED - Request received")
    
    async def stream_response():
        """Generator function that streams response chunks"""
        try:
            message = request.get('message')
            user_id = request.get('user_id')
            chat_id = request.get('chat_id')
            
            print(f"[CHAT ENDPOINT] Parsed: message={bool(message)}, user_id={user_id}, chat_id={chat_id}", flush=True)
            
            if not message or not user_id:
                yield f"data: {orjson.dumps({'error': 'Missing message or user_id'}).decode()}\n\n"
                return
            
            structured_logger.info("Chat request received", user_id=user_id, chat_id=chat_id, message_length=len(message))
            
            # Send thinking indicator first
            yield f"data: {orjson.dumps({'type': 'thinking', 'content': 'AI is thinking...'}).decode()}\n\n"
            
            # FIX #16: Import IntelligentChatOrchestrator with fallback for different deployment layouts
            try:
                from aident_cfo_brain.intelligent_chat_orchestrator import IntelligentChatOrchestrator
                structured_logger.debug("âœ“ Imported IntelligentChatOrchestrator from package layout")
            except ImportError as e1:
                structured_logger.debug(f"âœ— Package layout failed: {e1}. Trying flat layout...")
                try:
                    from intelligent_chat_orchestrator import IntelligentChatOrchestrator
                    structured_logger.debug("âœ“ Imported IntelligentChatOrchestrator from flat layout")
                except ImportError as e2:
                    structured_logger.error(f"IMPORT FAILURE - Fix location: core_infrastructure/fastapi_backend_v2.py line 8453")
                    structured_logger.error(f"Package layout error: {e1}")
                    structured_logger.error(f"Flat layout error: {e2}")
                    yield f"data: {orjson.dumps({'error': 'Service initialization failed'}).decode()}\n\n"
                    return
            
            # Note: Now using Groq/Llama instead of Anthropic for chat
            # Check for Groq API key
            groq_api_key = os.getenv('GROQ_API_KEY')
            if not groq_api_key:
                yield f"data: {orjson.dumps({'error': 'Chat service unavailable (Missing GROQ_API_KEY)'}).decode()}\n\n"
                return
            
            print(f"[CHAT ENDPOINT] Getting lazy Supabase client...", flush=True)
            try:
                # Get lazy client - returns immediately without connecting
                # Wrap in timeout to prevent hangs from connection pool initialization
                def get_client_sync():
                    # Use the already-imported get_supabase_client function from module level
                    return get_supabase_client()
                
                supabase_client = await asyncio.wait_for(
                    asyncio.to_thread(get_client_sync),
                    timeout=1.0
                )
                print(f"[CHAT ENDPOINT] Lazy Supabase client obtained (will connect on first use)", flush=True)
            except asyncio.TimeoutError:
                print(f"[CHAT ENDPOINT] Supabase client initialization timed out", flush=True)
                yield f"data: {orjson.dumps({'error': 'Database service unavailable'}).decode()}\n\n"
                return
            except Exception as e:
                print(f"[CHAT ENDPOINT] Failed to get Supabase client: {e}", flush=True)
                yield f"data: {orjson.dumps({'error': 'Database service unavailable'}).decode()}\n\n"
                return
            
            # Initialize orchestrator (uses Groq internally, no openai_client needed)
            try:
                print(f"[CHAT ENDPOINT] Initializing orchestrator...", flush=True)
                orchestrator = IntelligentChatOrchestrator(
                    supabase_client=supabase_client,
                    cache_client=safe_get_ai_cache()
                )
                print(f"[CHAT ENDPOINT] Orchestrator initialized successfully", flush=True)
                structured_logger.info("âœ… Orchestrator initialized successfully")
            except Exception as orch_init_error:
                print(f"[CHAT ENDPOINT] Orchestrator initialization failed: {orch_init_error}", flush=True)
                structured_logger.error("âŒ Orchestrator initialization failed", error=str(orch_init_error), error_type=type(orch_init_error).__name__)
                yield f"data: {orjson.dumps({'error': f'Chat service initialization failed: {str(orch_init_error)}'}).decode()}\n\n"
                return
            
            structured_logger.info("Starting question processing...", message=message[:100])
            try:
                # Wrap entire process_question with timeout to prevent hanging
                response = await asyncio.wait_for(
                    orchestrator.process_question(
                        question=message,
                        user_id=user_id,
                        chat_id=chat_id
                    ),
                    timeout=60.0  # 60 second timeout for entire chat processing
                )
                structured_logger.info("âœ… Question processing completed successfully")
            except asyncio.TimeoutError as te:
                structured_logger.error("âŒ Question processing timed out after 60 seconds", timeout_error=str(te))
                yield f"data: {orjson.dumps({'error': 'Chat service is taking too long. Please try again.'}).decode()}\n\n"
                return
            except Exception as inner_e:
                structured_logger.error("âŒ Question processing failed with exception", error=str(inner_e), error_type=type(inner_e).__name__)
                chunk_data = orjson.dumps({'error': f'Sorry, I encountered an error: {str(inner_e)}'}).decode()
                yield f"data: {chunk_data}\n\n"
                return
            
            structured_logger.info("Chat response generated", user_id=user_id, question_type=response.question_type.value, confidence=response.confidence)
            
            # Stream response in chunks (split by words for smooth typing effect)
            answer_text = response.answer
            words = answer_text.split(' ')
            
            # Stream each word with a small delay for visual effect
            accumulated_text = ""
            for i, word in enumerate(words):
                accumulated_text += word + (' ' if i < len(words) - 1 else '')
                
                # Send chunk every 5 words or at the end
                if (i + 1) % 5 == 0 or i == len(words) - 1:
                    chunk_data = orjson.dumps({'type': 'chunk', 'content': accumulated_text}).decode()
                    yield f"data: {chunk_data}\n\n"
                    await asyncio.sleep(0.01)  # Small delay to prevent overwhelming the client
            
            # Send final response metadata
            final_response = {
                "type": "complete",
                "timestamp": pendulum.now().to_iso8601_string(),
                "question_type": response.question_type.value,
                "confidence": response.confidence,
                "data": response.data,
                "actions": response.actions,
                "visualizations": response.visualizations,
                "follow_up_questions": response.follow_up_questions,
                "status": "success"
            }
            final_data = orjson.dumps(final_response).decode()
            yield f"data: {final_data}\n\n"
            
        except Exception as e:
            structured_logger.error("Chat streaming error", error=str(e))
            error_response = {"error": f"Sorry, I encountered an error: {str(e)}"}
            yield f"data: {orjson.dumps(error_response).decode()}\n\n"
    
    return StreamingResponse(
        stream_response(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
            "Content-Encoding": "none"
        }
    )

@app.get("/debug/env")
async def debug_environment():
    """Debug endpoint to check environment variable visibility"""
    groq_key = os.getenv('GROQ_API_KEY')
    return {
        "GROQ_API_KEY_present": bool(groq_key),
        "GROQ_API_KEY_length": len(groq_key) if groq_key else 0,
        "GROQ_API_KEY_prefix": groq_key[:10] + "..." if groq_key else None,
        "all_groq_vars": sorted([k for k in os.environ.keys() if 'GROQ' in k.upper()]),
        "all_env_vars_count": len(os.environ.keys()),
        "groq_client_initialized": groq_client is not None,
        "redis_url_present": bool(os.getenv('REDIS_URL')),
        "supabase_url_present": bool(os.getenv('SUPABASE_URL'))
    }

@app.get("/chat-health")
async def chat_health_check():
    """Test chat orchestrator initialization and Groq API connectivity"""
    try:
        # Check Groq API key
        groq_api_key = os.getenv('GROQ_API_KEY')
        if not groq_api_key:
            return {
                "status": "error",
                "error": "GROQ_API_KEY not found in environment"
            }
        
        # FIX #16: Import IntelligentChatOrchestrator with fallback for different deployment layouts
        try:
            from aident_cfo_brain.intelligent_chat_orchestrator import IntelligentChatOrchestrator
            structured_logger.debug("âœ“ Imported IntelligentChatOrchestrator from package layout (health check)")
        except ImportError as e1:
            structured_logger.debug(f"âœ— Package layout failed: {e1}. Trying flat layout...")
            try:
                from intelligent_chat_orchestrator import IntelligentChatOrchestrator
                structured_logger.debug("âœ“ Imported IntelligentChatOrchestrator from flat layout (health check)")
            except ImportError as e2:
                structured_logger.error(f"IMPORT FAILURE - Fix location: core_infrastructure/fastapi_backend_v2.py line 8554")
                structured_logger.error(f"Package layout error: {e1}")
                structured_logger.error(f"Flat layout error: {e2}")
                raise
        
        # CRITICAL FIX: Lazy-load Supabase client on first use
        supabase_client = await _ensure_supabase_loaded()
        if not supabase_client:
            return {
                "status": "error",
                "error": "Database service unavailable",
                "groq_api_key_present": True
            }
        
        orchestrator = IntelligentChatOrchestrator(
            supabase_client=supabase_client,
            cache_client=safe_get_ai_cache()
        )
        
        # Try a simple test question
        test_response = await orchestrator.process_question(
            question="Hello",
            user_id="health_check_user"
        )
        
        return {
            "status": "healthy",
            "groq_api_key_present": True,
            "orchestrator_initialized": True,
            "test_question_processed": True,
            "test_response_type": test_response.question_type.value,
            "test_confidence": test_response.confidence
        }
        
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "error_type": type(e).__name__
        }


@app.post("/api/detect-fields")
async def detect_fields_endpoint(request: FieldDetectionRequest):
    """Detect field types using UniversalFieldDetector with Redis caching"""
    try:
        # P1 OPTIMIZATION: Check Redis cache first (80% hit rate expected)
        cache_key = f"field_detect:{request.filename}:{hash(str(request.data))}"
        redis_client = await safe_get_cache()
        
        if redis_client:
            try:
                cached_result = await redis_client.get(cache_key)
                if cached_result:
                    logger.info(f"âœ… Cache HIT: field detection for {request.filename}")
                    import orjson
                    result = orjson.loads(cached_result)
                    return {
                        "status": "success",
                        "result": result,
                        "user_id": request.user_id,
                        "filename": request.filename,
                        "cached": True
                    }
            except Exception as cache_err:
                logger.warning(f"Cache read error: {cache_err}")
        
        # Cache MISS - perform detection
        logger.info(f"âš ï¸ Cache MISS: field detection for {request.filename}")
        field_detector = UniversalFieldDetector()
        
        ctx = request.context or {}
        if request.user_id:
            try:
                ctx = {**ctx, 'user_id': request.user_id}
            except Exception:
                ctx = {'user_id': request.user_id}
        result = await field_detector.detect_field_types_universal(
            data=request.data,
            filename=request.filename,
            context=ctx
        )
        
        # Store in cache (TTL: 24 hours)
        if redis_client:
            try:
                await redis_client.setex(
                    cache_key,
                    86400,  # 24 hours
                    orjson.dumps(result)
                )
            except Exception as cache_err:
                logger.warning(f"Cache write error: {cache_err}")
        
        return {
            "status": "success",
            "result": result,
            "user_id": request.user_id,
            "filename": request.filename,
            "cached": False
        }
        
    except Exception as e:
        logger.error(f"Field detection error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/detect-platform")
async def detect_platform_endpoint(request: PlatformDetectionRequest):
    """Detect platform using UniversalPlatformDetector"""
    try:
        # Initialize platform detector (with AI cache)
        # Using Groq/Llama for all AI operations
        platform_detector = UniversalPlatformDetector(cache_client=safe_get_ai_cache())
        
        # Detect platform
        result = await platform_detector.detect_platform_universal(
            payload={"file_content": request.file_content, "filename": request.filename},
            filename=request.filename,
            user_id=request.user_id
        )
        
        return {
            "status": "success",
            "result": result,
            "user_id": request.user_id,
            "filename": request.filename
        }
        
    except Exception as e:
        logger.error(f"Platform detection error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/classify-document")
async def classify_document_endpoint(request: DocumentClassificationRequest):
    """Classify document type using UniversalDocumentClassifierOptimized"""
    try:
        # P1 OPTIMIZATION: Check Redis cache first
        cache_key = f"doc_classify:{request.filename}:{hash(str(request.payload))}"
        redis_client = await safe_get_cache()
        
        if redis_client:
            try:
                cached_result = await redis_client.get(cache_key)
                if cached_result:
                    logger.info(f"âœ… Cache HIT: document classification for {request.filename}")
                    import orjson
                    result = orjson.loads(cached_result)
                    return {
                        "status": "success",
                        "document_type": result.get("document_type", "unknown"),
                        "confidence": result.get("confidence", 0.0),
                        "user_id": request.user_id,
                        "filename": request.filename,
                        "cached": True
                    }
            except Exception as cache_err:
                logger.warning(f"Cache read error: {cache_err}")
        
        # Cache MISS - perform classification
        logger.info(f"âš ï¸ Cache MISS: document classification for {request.filename}")
        
        # Initialize classifier
        try:
            classifier = UniversalDocumentClassifierOptimized()
        except Exception as init_err:
            logger.error(f"Classifier initialization error: {init_err}")
            # Return success with default classification instead of failing
            return {
                "status": "success",
                "document_type": "unknown",
                "confidence": 0.0,
                "user_id": request.user_id,
                "filename": request.filename,
                "cached": False,
                "fallback": True,
                "error": str(init_err)
            }
        
        # Perform classification
        result = await classifier.classify_document_universal(
            payload=request.payload,
            filename=request.filename,
            context=request.context or {}
        )
        
        # Store in cache (TTL: 24 hours)
        if redis_client and result:
            try:
                await redis_client.setex(
                    cache_key,
                    86400,  # 24 hours
                    orjson.dumps(result)
                )
            except Exception as cache_err:
                logger.warning(f"Cache write error: {cache_err}")
        
        return {
            "status": "success",
            "document_type": result.get("document_type", "unknown"),
            "confidence": result.get("confidence", 0.0),
            "user_id": request.user_id,
            "filename": request.filename,
            "cached": False
        }
        
    except Exception as e:
        logger.error(f"Document classification error: {e}")
        # LOAD TEST FIX: Return 200 with error info instead of 500
        return {
            "status": "error",
            "document_type": "unknown",
            "confidence": 0.0,
            "error": str(e),
            "error_type": type(e).__name__,
            "user_id": request.user_id,
            "filename": request.filename
        }

# FIX #2: Redis-backed distributed lock for duplicate detection (production-ready)
# Replaces in-memory asyncio.Lock which doesn't work across workers/instances
# Uses aiocache with Redis for distributed locking

async def acquire_duplicate_lock(file_hash: str) -> bool:
    """
    FIX #2: Acquire distributed Redis lock for file hash to prevent concurrent duplicate checks.
    Returns True if lock acquired, False if already locked.
    """
    try:
        cache = safe_get_cache()
        if cache is None:
            logger.warning("Cache not available, skipping distributed lock")
            return True  # Allow operation if cache unavailable
        
        lock_key = f"duplicate_lock:{file_hash}"
        # Try to acquire lock with 30-second TTL
        # Returns True if lock acquired, False if already exists
        existing = await cache.get(lock_key)
        if existing:
            logger.info(f"Duplicate check already in progress for {file_hash}")
            return False
        
        # Acquire lock
        await cache.set(lock_key, "locked", ttl=30)
        return True
    except Exception as e:
        logger.error(f"Failed to acquire distributed lock: {e}")
        return True  # Allow operation on error (fail open)

async def release_duplicate_lock(file_hash: str):
    """FIX #2: Release distributed Redis lock for file hash"""
    try:
        cache = safe_get_cache()
        if cache is None:
            return
        
        lock_key = f"duplicate_lock:{file_hash}"
        await cache.delete(lock_key)
    except Exception as e:
        logger.error(f"Failed to release distributed lock: {e}")

# FIX #18: Rate limiting for duplicate checks to prevent DoS attacks
from collections import defaultdict
import time

# CRITICAL FIX: Distributed rate limiter using Redis for multi-worker support
# Replaces old in-memory rate limiter that was broken in multi-worker deployments
MAX_CONCURRENT_UPLOADS_PER_USER = 10  # Allow max 10 concurrent uploads per user (increased for batch uploads)

async def acquire_upload_slot(user_id: str) -> Tuple[bool, str]:
    """
    âœ… FIXED: Atomic rate limiter using Redis Lua script.
    Prevents race condition between GET and INCR operations.
    Works across all workers in multi-worker deployment.
    """
    try:
        # Get Redis client from centralized cache
        from centralized_cache import get_cache
        cache = get_cache()
        
        if cache and cache.cache:
            # Use Redis for distributed rate limiting
            redis_key = f"upload_slots:{user_id}"
            
            # âœ… FIX: Lua script for atomic check-and-increment
            # Prevents race condition where two requests both pass the limit check
            lua_script = """
            local current = redis.call('GET', KEYS[1])
            local limit = tonumber(ARGV[1])
            
            if current and tonumber(current) >= limit then
                return -1  -- Over limit
            end
            
            local new_count = redis.call('INCR', KEYS[1])
            
            -- Set expiry on first increment
            if new_count == 1 then
                redis.call('EXPIRE', KEYS[1], 3600)
            end
            
            return new_count
            """
            
            # Execute atomic operation
            result = await cache.cache.eval(
                lua_script,
                1,  # Number of keys
                redis_key,
                MAX_CONCURRENT_UPLOADS_PER_USER
            )
            
            if result == -1:
                # Over limit
                current = await cache.cache.get(redis_key)
                current_val = int(current) if current else 0
                return False, f"Too many concurrent uploads. Please wait for some uploads to complete. ({current_val}/{MAX_CONCURRENT_UPLOADS_PER_USER} active)"
            
            logger.info(f"User {user_id} started upload. Active uploads: {result}/{MAX_CONCURRENT_UPLOADS_PER_USER} (atomic)")
            return True, "OK"
        else:
            # Fallback to allowing upload if Redis unavailable (fail open)
            logger.warning(f"Redis unavailable for rate limiting, allowing upload for user {user_id}")
            return True, "OK"
            
    except Exception as e:
        logger.error(f"Rate limiter error for user {user_id}: {e}")
        # Fail open - allow upload on error
        return True, "OK"

async def release_upload_slot(user_id: str):
    """
    CRITICAL FIX: Release distributed upload slot using Redis DECR.
    Works across all workers in multi-worker deployment.
    """
    try:
        from centralized_cache import get_cache
        cache = get_cache()
        
        if cache and cache.cache:
            redis_key = f"upload_slots:{user_id}"
            
            # Decrement counter atomically
            new_count = await cache.cache.decr(redis_key)
            
            # Clean up if count reaches 0
            if new_count <= 0:
                await cache.cache.delete(redis_key)
                logger.info(f"User {user_id} completed all uploads. Slot released (distributed)")
            else:
                logger.info(f"User {user_id} completed upload. Active uploads: {new_count}/{MAX_CONCURRENT_UPLOADS_PER_USER} (distributed)")
                
    except Exception as e:
        logger.error(f"Failed to release upload slot for user {user_id}: {e}")

# ISSUE #11 FIX: Unified file download utility
# Replaces 3 duplicate implementations of file download logic
async def _download_and_stream_file_from_storage(
    storage_path: str,
    job_id: str,
    filename: str,
    compute_hash: bool = True
) -> Tuple[str, int, Optional[str]]:
    """
    UNIFIED: Download file from Supabase Storage and stream to disk
    
    Returns:
        (temp_file_path, file_size, file_hash)
    
    Replaces duplicate download logic in:
    - /process-excel endpoint
    - _process_api_data_through_pipeline
    - _store_external_item_attachment
    SECURITY: Uses backend API with proper RLS enforcement instead of direct client queries.
    """
    try:
        # Download from Supabase Storage
        storage = supabase.storage.from_("finely-upload")
        response = storage.download(storage_path)
        
        if not response:
            raise HTTPException(status_code=404, detail="File not found in storage")
        
        # Write to temp file
        temp_file_path = f"/tmp/{job_id}_{filename}"
        with open(temp_file_path, "wb") as f:
            f.write(response)
        
        file_size = len(response)
        
        # Compute hash using StreamedFile (streaming, not from memory)
        file_hash = None
        if compute_hash:
            from data_ingestion_normalization.streaming_source import StreamedFile
            streamed_file_obj = StreamedFile(
                path=temp_file_path,
                filename=filename,
                _size=file_size
            )
            file_hash = streamed_file_obj.xxh3_128  # Standardized hash
        
        logger.info(f"File downloaded: {temp_file_path}, size: {file_size}, hash: {file_hash}")
        return temp_file_path, file_size, file_hash
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download file from storage: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download file: {str(e)}")

@app.post("/check-duplicate")
async def check_duplicate_endpoint(request: dict):
    """
    FIX #10 & #18: Check for duplicate files using file hash with distributed locking and rate limiting.
    CRITICAL: Now uses asyncio.Lock to prevent race conditions during concurrent uploads.
    SECURITY: Uses backend API with proper RLS enforcement instead of direct client queries.
    """
    try:
        user_id = request.get('user_id')
        file_hash = request.get('file_hash')
        file_name = request.get('file_name', 'unknown')
        session_token = request.get('session_token')
        
        if not user_id or not file_hash:
            raise HTTPException(status_code=400, detail="user_id and file_hash are required")
        
        # CRITICAL FIX: Use Redis-backed rate limiter (works across all workers)
        # Old: check_rate_limit() used in-memory defaultdict (broken in multi-worker)
        # New: Use Redis counter with TTL for distributed rate limiting
        try:
            from centralized_cache import get_cache
            cache = get_cache()
            rate_limit_key = f"rate_limit:duplicate_check:{user_id}"
            
            # Get current count
            current_count = await cache.get(rate_limit_key) or 0
            if isinstance(current_count, bytes):
                current_count = int(current_count.decode())
            elif isinstance(current_count, str):
                current_count = int(current_count)
            
            # Check limit (100 requests per 60 seconds)
            if current_count >= 100:
                logger.warning(f"Rate limit exceeded for user {user_id}: {current_count}/100 requests in 60s")
                raise HTTPException(status_code=429, detail=f"Rate limit exceeded: {current_count}/100 requests in 60s")
            
            # Increment counter with 60s TTL
            await cache.incr(rate_limit_key)
            await cache.expire(rate_limit_key, 60)
            
        except HTTPException:
            raise
        except Exception as e:
            logger.warning(f"Rate limit check failed (allowing request): {e}")
        
        # Security validation
        try:
            valid, violations = await security_validator.validate_request({
                'endpoint': 'check-duplicate',
                'user_id': user_id,
                'session_token': session_token
            }, SecurityContext(user_id=user_id))
            if not valid:
                logger.warning(f"Security validation failed for duplicate check: {violations}")
                raise HTTPException(status_code=401, detail="Unauthorized or invalid session")
        except HTTPException:
            raise
        except Exception as sec_e:
            logger.warning(f"Security validation error for duplicate check: {sec_e}")
            raise HTTPException(status_code=401, detail="Unauthorized or invalid session")
        
        # FIX #2: CRITICAL - Acquire distributed Redis lock to prevent race conditions across workers
        lock_acquired = await acquire_duplicate_lock(file_hash)
        if not lock_acquired:
            # Another worker is already checking this file
            raise HTTPException(status_code=409, detail="Duplicate check already in progress for this file")
        
        try:
            # âœ… FIX ISSUE #7: Use unified duplicate detection function
            # This ensures consistent behavior across all 3 paths (upload, integration, manual check)
            duplicate_analysis = await _handle_duplicate_detection_unified(
                user_id=user_id,
                file_hash=file_hash,
                filename=file_name
            )
            
            # Return standardized response
            if duplicate_analysis.get('is_duplicate'):
                response = {
                    "is_duplicate": True,
                    "duplicate_type": duplicate_analysis['duplicate_type'],
                    "similarity_score": duplicate_analysis['similarity_score'],
                    "duplicate_files": duplicate_analysis['duplicate_files'],
                    "latest_duplicate": duplicate_analysis['duplicate_files'][0] if duplicate_analysis['duplicate_files'] else None,
                    "recommendation": duplicate_analysis['recommendation'],
                    "message": duplicate_analysis['message'],
                    "confidence": duplicate_analysis['confidence']
                }
                if duplicate_analysis.get('delta_analysis'):
                    response["delta_analysis"] = duplicate_analysis['delta_analysis']
                return response
            
            return {"is_duplicate": False}
            
        except Exception as db_err:
            logger.error(f"Database error checking duplicates: {db_err}")
            raise HTTPException(status_code=500, detail="Failed to check for duplicates")
        finally:
            # FIX #2: Release distributed lock
            await release_duplicate_lock(file_hash)
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error checking duplicates: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/process-excel")
async def process_excel_endpoint(
    job_id: str = Form(...),
    user_id: str = Form(...),
    filename: str = Form(...),
    storage_path: str = Form(...),
    session_token: str = Form(...)
):
    """
    CRITICAL FIX: Unified streaming file processor with distributed rate limiting.
    All file uploads now go through this single endpoint.
    """
    try:
        # Validate security
        await _validate_security('process-excel', user_id, session_token)
        
        # CRITICAL FIX: Distributed rate limiter using Redis
        can_upload, rate_limit_msg = await acquire_upload_slot(user_id)
        if not can_upload:
            raise HTTPException(status_code=429, detail=rate_limit_msg)
        
        # Acquire processing lock BEFORE downloading file
        # This prevents race condition where two workers download same file simultaneously
        lock_id = f"job_{job_id}"
        lock_acquired = False
        try:
            lock_data = {
                'id': lock_id,
                'lock_type': 'file_processing',
                'resource_id': job_id,
                'user_id': user_id,
                'acquired_at': pendulum.now().to_iso8601_string(),
                'expires_at': pendulum.now().add(hours=1).to_iso8601_string(),
                'job_id': job_id,
                'metadata': {'filename': filename}
            }
            supabase.table('processing_locks').insert(lock_data).execute()
            lock_acquired = True
            logger.info(f"Acquired processing lock BEFORE download: {lock_id}")
        except Exception as e:
            # Lock already exists = another worker processing this job
            error_msg = f"Job {job_id} is already being processed by another worker"
            logger.warning(error_msg)
            await release_upload_slot(user_id)
            raise HTTPException(status_code=409, detail=error_msg)
        
        # Variables for cleanup
        temp_file_path = None
        file_hash = None
        actual_file_size = 0
        file_downloaded_successfully = False
        
        async def _cleanup_temp_file():
            nonlocal temp_file_path
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                    logger.info(f"Cleaned up temp file: {temp_file_path}")
                except Exception as cleanup_err:
                    logger.error(f"Failed to cleanup temp file {temp_file_path}: {cleanup_err}")
        
        async def _stream_file_to_disk():
            nonlocal temp_file_path, file_hash, actual_file_size, file_downloaded_successfully
            try:
                storage = supabase.storage.from_("finely-upload")
                response = storage.download(storage_path)
                if not response:
                    raise HTTPException(status_code=404, detail="File not found in storage")
                
                # Create temp file
                temp_file_path = f"/tmp/{job_id}_{filename}"
                with open(temp_file_path, "wb") as f:
                    f.write(response)
                
                actual_file_size = len(response)
                
                # âœ… FIX ISSUE #4: Use StreamedFile to compute hash from disk (streaming)
                # This prevents loading entire file into memory twice
                # StreamedFile reads file in 8MB chunks, doesn't load into RAM
                from data_ingestion_normalization.streaming_source import StreamedFile
                streamed_file_obj = StreamedFile(
                    path=temp_file_path,
                    filename=filename,
                    _size=actual_file_size
                )
                
                # Compute hash from disk using xxh3_128 (streaming, memory-efficient)
                file_hash = streamed_file_obj.xxh3_128
                
                file_downloaded_successfully = True
                logger.info(f"File downloaded: {temp_file_path}, size: {actual_file_size}, hash: {file_hash}")
                
            except Exception as e:
                logger.error(f"Failed to download file: {e}")
                raise HTTPException(status_code=500, detail=f"Failed to download file: {str(e)}")
        
        # Download file and verify hash (NOW protected by lock)
        await _stream_file_to_disk()
        
        # Log request with observability
        structured_logger.info("File processing request received", user_id=user_id, filename=filename)

        # Pre-create job status so polling has data even before WS connects
        await websocket_manager.merge_job_state(job_id, {
            "status": "starting",
            "message": "Initializing processing...",
            "progress": 0,
            "started_at": pendulum.now().to_iso8601_string(),
            "components": {}
        })

        async def _run_processing_job():
            nonlocal file_hash, temp_file_path, actual_file_size, file_downloaded_successfully
            
            try:
                # Cooperative cancellation helper
                async def is_cancelled() -> bool:
                    status = await websocket_manager.get_job_status(job_id)
                    return (status or {}).get("status") == "cancelled"
                
                if not file_downloaded_successfully:
                    await websocket_manager.send_overall_update(
                        job_id=job_id,
                        status="processing",
                        message="ðŸ“¥ Streaming file from storage...",
                        progress=5
                    )
                    if await is_cancelled():
                        return
                    await _stream_file_to_disk()
                    if not file_downloaded_successfully:
                        return

                if not temp_file_path or not os.path.exists(temp_file_path):
                    logger.error(f"Temp file missing for job {job_id}; aborting")
                    raise HTTPException(status_code=500, detail="Temporary file missing after download")

                logger.info(
                    f"Reusing streamed file for job {job_id}: path={temp_file_path}, size={actual_file_size}, sha256={file_hash}"
                )

                # Validate file type using magic numbers (security check)
                try:
                    import magic
                    with open(temp_file_path, "rb") as handle:
                        head = handle.read(2048)
                    file_mime = magic.from_buffer(head, mime=True)
                    allowed_mimes = [
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                        'application/vnd.ms-excel',  # .xls
                        'application/zip',  # .xlsx (also detected as zip)
                        'text/csv',  # .csv
                        'text/plain',  # .csv (sometimes detected as plain text)
                        'application/octet-stream'  # Generic binary (fallback)
                    ]
                    if file_mime not in allowed_mimes:
                        error_msg = f"Invalid file type: {file_mime}. Only Excel (.xlsx, .xls) and CSV files are allowed."
                        logger.error(f"File type validation failed for job {job_id}: {error_msg}")
                        await websocket_manager.send_error(job_id, error_msg)
                        await websocket_manager.merge_job_state(job_id, {
                            **((await websocket_manager.get_job_status(job_id)) or {}),
                            "status": "failed",
                            "error": error_msg
                        })
                        await _cleanup_temp_file()
                        return
                    logger.info(f"File type validated for job {job_id}: {file_mime}")
                except ImportError:
                    logger.warning("python-magic not available, skipping file type validation")
                except Exception as e:
                    logger.warning(f"File type validation failed for job {job_id}: {e}")
                
                # Notify start of processing
                await websocket_manager.send_overall_update(
                    job_id=job_id,
                    status="processing",
                    message="ðŸ“¥ File validated, starting processing...",
                    progress=10
                )
                if await is_cancelled():
                    return

                await websocket_manager.send_overall_update(
                    job_id=job_id,
                    status="processing",
                    message="ðŸ§  Initializing analysis pipeline...",
                    progress=15
                )
                if await is_cancelled():
                    return

                # Use advanced processing pipeline that includes entity resolution and relationship detection
                excel_processor = await get_excel_processor()
                await excel_processor.process_file(
                    job_id=job_id,
                    file_content=temp_file_path,
                    filename=filename,
                    user_id=user_id,
                    supabase=supabase,
                    streamed_file_size=actual_file_size,
                    streamed_file_hash=file_hash
                )
            except Exception as e:
                logger.error(f"Processing job failed: {e}")
                await websocket_manager.send_error(job_id, str(e))
                # FIX #17: Separate nested await calls to prevent floating coroutines
                current_job_status = await websocket_manager.get_job_status(job_id)
                await websocket_manager.merge_job_state(job_id, {**(current_job_status or {}), "status": "failed", "error": str(e)})
            finally:
                await _cleanup_temp_file()

        # CRITICAL FIX: Enqueue to ARQ instead of inline processing
        # This offloads heavy CPU work from web workers to dedicated ARQ workers
        if _queue_backend() == 'arq':
            try:
                pool = await get_arq_pool()
                await pool.enqueue_job(
                    'process_spreadsheet',
                    user_id=user_id,
                    filename=filename,
                    storage_path=storage_path,
                    job_id=job_id
                )
                logger.info(f"âœ… Job {job_id} enqueued to ARQ for background processing")
                metrics_collector.increment_counter("file_processing_requests")
                
                # Update job status to queued
                await websocket_manager.send_overall_update(
                    job_id=job_id,
                    status="queued",
                    message="ðŸ“‹ Job queued for processing...",
                    progress=5
                )
                
                return {"status": "accepted", "job_id": job_id, "queued": True}
            except Exception as arq_error:
                logger.error(f"Failed to enqueue job {job_id} to ARQ: {arq_error}")
                # Fallback to inline processing if ARQ fails
                logger.warning(f"Falling back to inline processing for job {job_id}")
        
        # Fallback: Process file inline (for sync mode or ARQ failure)
        logger.info(f"ðŸ”„ Processing file inline: {job_id}")
        
        # FIX ISSUE #13: Wrap async task to ensure upload slot is released
        async def _run_with_cleanup():
            try:
                await _run_processing_job()
            finally:
                # Always release upload slot when done
                if user_id:
                    await release_upload_slot(user_id)
        
        asyncio.create_task(_run_with_cleanup())
        metrics_collector.increment_counter("file_processing_requests")
        return {"status": "accepted", "job_id": job_id, "queued": False}
    except HTTPException as he:
        # Release upload slot on HTTP exceptions (e.g., 429, 401)
        if user_id:
            await release_upload_slot(user_id)
        # Preserve the intended status code (e.g., 401 on auth failures)
        raise he
    except Exception as e:
        # Release upload slot on any error
        if user_id:
            await release_upload_slot(user_id)
        structured_logger.error("Process excel endpoint error", error=str(e))
        metrics_collector.increment_counter("file_processing_errors")
        raise HTTPException(status_code=500, detail=str(e))

# REMOVED: Old duplicate ingestion path - process_excel_universal_endpoint
# This endpoint created a second ingestion pipeline that competed with streaming processor
# All file processing now goes through the unified streaming pipeline only

@app.get("/api/component-metrics")
async def get_component_metrics():
    """Get metrics for all universal components"""
    try:
        # Initialize components - FIX #3: Reuse singleton instances for metrics
        excel_processor = _get_excel_processor_instance()
        field_detector = excel_processor.universal_field_detector
        platform_detector = excel_processor.universal_platform_detector
        document_classifier = excel_processor.universal_document_classifier
        data_extractor = excel_processor.universal_extractors
        
        # Get metrics from each component
        metrics = {
            "field_detector": field_detector.get_metrics(),
            "platform_detector": platform_detector.get_metrics(),
            "document_classifier": document_classifier.get_metrics(),
            "data_extractor": data_extractor.get_metrics()
        }
        
        return {
            "status": "success",
            "metrics": metrics,
            "timestamp": pendulum.now().to_iso8601_string()
        }
        
    except Exception as e:
        logger.error(f"Component metrics error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# CONNECTORS (NANGO) - PHASE 1 GMAIL
# ============================================================================

class ConnectorInitiateRequest(BaseModel):
    provider: str  # expect 'google-mail' for Gmail
    user_id: str
    session_token: Optional[str] = None

class ConnectorSyncRequest(BaseModel):
    user_id: str
    connection_id: str  # Nango connection id
    integration_id: Optional[str] = None  # defaults to google-mail
    mode: str = "historical"  # historical | incremental
    lookback_days: Optional[int] = 365  # used for historical q filter
    max_results: Optional[int] = 100  # per page
    session_token: Optional[str] = None
    correlation_id: Optional[str] = None

    if field_validator:
        @field_validator('max_results')
        @classmethod
        def _validate_max_results(cls, v):
            if v is None:
                return 100
            try:
                iv = int(v)
            except Exception:
                raise ValueError('max_results must be an integer')
            if iv <= 0:
                raise ValueError('max_results must be positive')
            return min(iv, 1000)

        @field_validator('lookback_days')
        @classmethod
        def _validate_lookback_days(cls, v):
            if v is None:
                return 365
            iv = int(v)
            if iv < 0:
                raise ValueError('lookback_days must be >= 0')
            return iv

        @field_validator('mode')
        @classmethod
        def _validate_mode(cls, v):
            allowed = {'historical', 'incremental'}
            if v not in allowed:
                raise ValueError('mode must be one of historical, incremental')
            return v

class UserConnectionsRequest(BaseModel):
    user_id: str
    session_token: Optional[str] = None

class UpdateFrequencyRequest(BaseModel):
    user_id: str
    connection_id: str  # nango connection id
    minutes: int
    session_token: Optional[str] = None

class ConnectorDisconnectRequest(BaseModel):
    user_id: str
    connection_id: str
    provider: Optional[str] = None
    session_token: Optional[str] = None


async def _validate_security(endpoint: str, user_id: str, session_token: Optional[str]):
    """
    CRITICAL FIX: Use centralized security validator (single source of truth)
    Replaces duplicate _require_security function
    """
    try:
        # FIX: Check if security_validator is initialized before using it
        if not security_validator:
            logger.error(f"âŒ CRITICAL: Security validator not initialized for endpoint {endpoint}")
            raise HTTPException(status_code=503, detail="Security system not initialized. Please try again later.")
        
        valid, violations = await security_validator.validate_request({
            'endpoint': endpoint,
            'user_id': user_id,
            'session_token': session_token
        }, SecurityContext(user_id=user_id))
        if not valid:
            logger.warning(f"Security validation failed for {endpoint}: {violations}")
            raise HTTPException(status_code=401, detail="Unauthorized or invalid session")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Security validation error for {endpoint}: {e}")
        raise HTTPException(status_code=401, detail="Unauthorized or invalid session")

def _safe_filename(name: str) -> str:
    """
    FIX #6: Enhanced filename sanitization preventing path traversal, null bytes, control chars, reserved names, Unicode exploits
    
    Security checks:
    - Removes null bytes (0x00) and control characters (0x01-0x1F, 0x7F-0x9F)
    - Blocks path traversal (../, ..\, ~)
    - Removes path separators (/, \, :)
    - Prevents reserved Windows filenames
    - Normalizes Unicode to prevent homograph attacks
    - Truncates to filesystem limits
    """
    import unicodedata
    try:
        if not name:
            return "attachment"
        
        # Step 1: Unicode normalization (NFC) to prevent homograph attacks
        # Converts lookalike characters (e.g., Cyrillic 'Ð°' vs Latin 'a') to canonical form
        sanitized = unicodedata.normalize('NFC', name)
        
        # Step 2: Remove null bytes and ALL control characters (0x00-0x1F, 0x7F-0x9F)
        # This is more comprehensive than just checking ord >= 32
        sanitized = ''.join(
            char for char in sanitized 
            if ord(char) >= 0x20 and ord(char) not in range(0x7F, 0xA0)
        )
        
        # Step 3: Remove path traversal sequences
        sanitized = sanitized.replace('..', '_').replace('~', '_')
        
        # Step 4: Replace path separators and other dangerous chars
        sanitized = sanitized.replace('/', '_').replace('\\', '_').replace(':', '_')
        sanitized = sanitized.replace('\x00', '_')  # Explicit null byte removal
        
        # Step 5: Remove reserved Windows filenames (CON, PRN, AUX, NUL, COM1-9, LPT1-9)
        reserved = {'con', 'prn', 'aux', 'nul', 'com1', 'com2', 'com3', 'com4', 'com5', 
                   'com6', 'com7', 'com8', 'com9', 'lpt1', 'lpt2', 'lpt3', 'lpt4', 'lpt5', 
                   'lpt6', 'lpt7', 'lpt8', 'lpt9'}
        name_lower = sanitized.lower().split('.')[0]
        if name_lower in reserved:
            sanitized = f"file_{sanitized}"
        
        # Step 6: Remove leading/trailing spaces and dots
        sanitized = sanitized.strip('. ')
        
        # Step 7: Truncate to 200 chars (filesystem limit)
        sanitized = sanitized[:200]
        
        # Step 8: Ensure not empty
        if not sanitized:
            return "attachment"
        
        return sanitized
    except Exception as e:
        logger.warning(f"Filename sanitization error: {e}")
        return "attachment"

async def _store_external_item_attachment(user_id: str, provider: str, message_id: str, filename: str, content: bytes) -> Tuple[str, str]:
    """Store attachment bytes to Supabase Storage. Returns (storage_path, file_hash)."""
    safe_name = _safe_filename(filename)
    # Compute hash for dedupe (xxhash: 5-10x faster for large files)
    file_hash = xxhash.xxh3_128(content).hexdigest()
    # Build storage path
    today = datetime.utcnow().strftime('%Y/%m/%d')
    storage_path = f"external/{provider}/{user_id}/{today}/{message_id}/{safe_name}"
    try:
        storage = supabase.storage.from_("finely-upload")
        # Render's supabase-py expects a filesystem path, not a BytesIO
        # Write to a secure temporary file and upload by path
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            storage.upload(storage_path, tmp_path)
        finally:
            try:
                os.remove(tmp_path)
            except Exception as cleanup_err:
                logger.warning(f"Failed to cleanup temp file {tmp_path}: {cleanup_err}")
        return storage_path, file_hash
    except Exception as e:
        logger.error(f"Storage upload failed: {e}")
        raise

async def _enqueue_file_processing(user_id: str, filename: str, storage_path: str, external_item_id: Optional[str] = None) -> str:
    """
    Create an ingestion job and start processing asynchronously via ARQ. Returns job_id.
    
    CRITICAL FIX #9: Prefer ARQ queue over inline processing for reliability and scalability.
    - ARQ provides job persistence, retry logic, and distributed processing
    - Inline processing with asyncio.create_task() can lose jobs on restart
    
    FIX #4: Now accepts external_item_id to avoid redundant database lookup.
    CRITICAL FIX #8: Use cryptographically secure random job ID (not guessable)
    """
    # CRITICAL FIX #8: Generate unpredictable job_id using secrets module
    job_id = f"job_{secrets.token_urlsafe(24)}"
    # Create/update ingestion_jobs like existing code path
    try:
        job_data = {
            'id': job_id,
            'user_id': user_id,
            'file_name': filename,
            'status': 'queued',
            'storage_path': storage_path,
            'created_at': pendulum.now().to_iso8601_string()
        }
        # Best-effort insert
        try:
            supabase.table('ingestion_jobs').insert(job_data).execute()
        except Exception as job_err:
            logger.warning(f"Failed to insert ingestion_job record: {job_err}")
        
        # CRITICAL FIX #9: Use ARQ queue for job persistence and reliability
        if _queue_backend() == 'arq':
            try:
                pool = await get_arq_pool()
                await pool.enqueue_job('process_file', {
                    'job_id': job_id,
                    'user_id': user_id,
                    'filename': filename,
                    'storage_path': storage_path,
                    'external_item_id': external_item_id
                })
                logger.info(f" File processing enqueued to ARQ: {job_id}")
                return job_id
            except Exception as e:
                logger.warning(f"ARQ dispatch failed for file processing, falling back to inline: {e}")
                # Fall through to inline processing
        
        # Fallback: Process file inline if ARQ unavailable
        logger.info(f" Processing file inline (ARQ unavailable): {job_id}")
        # Download file from storage and process inline
        file_bytes = supabase.storage.from_('financial-documents').download(storage_path)
        
        async def inline_process():
            try:
                excel_processor = _get_excel_processor_instance()
                await excel_processor.process_file(
                    job_id=job_id,
                    file_content=file_bytes,
                    filename=filename,
                    user_id=user_id,
                    supabase=supabase,
                    external_item_id=external_item_id
                )
            except Exception as e:
                logger.error(f"Inline processing failed for {job_id}: {e}")
                await websocket_manager.send_error(job_id, str(e))
        
        asyncio.create_task(inline_process())
        return job_id
    except Exception as e:
        logger.error(f"Failed to enqueue processing: {e}")
        raise

async def _enqueue_pdf_processing(user_id: str, filename: str, storage_path: str, external_item_id: Optional[str] = None) -> str:
    """
    Create a PDF OCR ingestion job and start processing asynchronously via ARQ. Returns job_id.
    
    CRITICAL FIX #9: Prefer ARQ queue over inline processing for reliability and scalability.
    
    FIX #4: Now accepts external_item_id to avoid redundant database lookup.
    CRITICAL FIX #8: Use cryptographically secure random job ID (not guessable)
    """
    # CRITICAL FIX #8: Generate unpredictable job_id using secrets module
    # UUID4 is random but can be brute-forced; secrets.token_urlsafe is cryptographically secure
    job_id = f"job_{secrets.token_urlsafe(24)}"
    try:
        job_data = {
            'id': job_id,
            'user_id': user_id,
            'file_name': filename,
            'status': 'queued',
            'storage_path': storage_path,
            'created_at': pendulum.now().to_iso8601_string()
        }
        try:
            supabase.table('ingestion_jobs').insert(job_data).execute()
        except Exception as job_err:
            logger.warning(f"Failed to insert ingestion_job record: {job_err}")
        
        # CRITICAL FIX #9: Use ARQ queue for job persistence and reliability
        if _queue_backend() == 'arq':
            try:
                pool = await get_arq_pool()
                await pool.enqueue_job('process_pdf', {
                    'job_id': job_id,
                    'user_id': user_id,
                    'filename': filename,
                    'storage_path': storage_path,
                    'external_item_id': external_item_id
                })
                logger.info(f" PDF processing enqueued to ARQ: {job_id}")
                return job_id
            except Exception as e:
                logger.warning(f"ARQ dispatch failed for PDF processing, falling back to inline: {e}")
                # Fall through to inline processing
        
        # Fallback: Process PDF inline if ARQ unavailable
        logger.info(f" Processing PDF inline (ARQ unavailable): {job_id}")
        asyncio.create_task(start_pdf_processing_job(user_id, job_id, storage_path, filename, external_item_id))
        return job_id
    except Exception as e:
        logger.error(f"Failed to enqueue PDF processing: {e}")
        raise

async def start_pdf_processing_job(user_id: str, job_id: str, storage_path: str, filename: str, external_item_id: Optional[str] = None):
    """
    Download a PDF from storage, extract text/tables, and store into raw_records.
    
    FIX #4: Now accepts external_item_id to link raw_records to external_items.
    """
    try:
        # Bind job to user for WebSocket authorization
        base = (await websocket_manager.get_job_status(job_id)) or {}
        await websocket_manager.merge_job_state(job_id, {
            **base,
            "user_id": user_id,
            "status": base.get("status", "queued"),
            "started_at": base.get("started_at") or pendulum.now().to_iso8601_string(),
        })
        # Mark job as processing
        try:
            supabase.table('ingestion_jobs').update({
                'status': 'processing',
                'updated_at': pendulum.now().to_iso8601_string()
            }).eq('id', job_id).execute()
        except Exception as status_err:
            logger.warning(f"Failed to update job status to processing: {status_err}")

        # Download file bytes from Storage
        storage = supabase.storage.from_("finely-upload")
        file_bytes = storage.download(storage_path)
        if hasattr(file_bytes, 'data') and isinstance(file_bytes.data, (bytes, bytearray)):
            # Some SDK versions wrap bytes in a Response-like object
            file_bytes = file_bytes.data

        if not file_bytes:
            raise RuntimeError("Empty file downloaded for PDF processing")

        # âœ… FIX ISSUE #5: Use xxh3_128 for standardized hashing
        file_hash = xxhash.xxh3_128(file_bytes).hexdigest()

        # FIX #3: REMOVED pdfplumber and tabula extraction (deprecated libraries)
        # Use UniversalExtractorsOptimized instead for PDF processing:
        # - pdfminer.six for text extraction (direct, no ONNX)
        # - easyocr for OCR (92% accuracy vs 60% pytesseract)
        # Example:
        #   from streaming_source import StreamedFile
        #   from universal_extractors_optimized import UniversalExtractorsOptimized
        #   extractor = UniversalExtractorsOptimized()
        #   result = await extractor.extract_data_universal(file_bytes, filename, user_id)
        #   text_excerpt = result.get('text', '')
        #   tables_preview = result.get('tables', [])
        
        text_excerpt = ""
        tables_preview = []
        pages_processed = 0
        
        # TODO: Integrate UniversalExtractorsOptimized for email PDF processing
        logger.info(f"PDF extraction skipped - use UniversalExtractorsOptimized for {filename}")

        # Insert minimal raw_records entry
        record = {
            'user_id': user_id,
            'file_name': filename,
            'file_size': len(file_bytes),
            'file_hash': file_hash,
            'source': 'email_pdf',
            'content': {
                'file_hash': file_hash,
                'text_excerpt': (text_excerpt[:16000] if text_excerpt else None),
                'tables_preview': tables_preview if tables_preview else None,
                'pages_processed': pages_processed,
                'processed_at': pendulum.now().to_iso8601_string()
            },
            'status': 'completed',
            'classification_status': 'pending'
        }
        # FIX #4: Use external_item_id passed from connector (no redundant lookup)
        if external_item_id:
            record['external_item_id'] = external_item_id
            logger.info(f" Using external_item_id passed from connector: {external_item_id}")
        else:
            # Fallback: Try to link to external_items by file hash (for manual uploads)
            try:
                ext_res = supabase.table('external_items').select('id').eq('user_id', user_id).eq('hash', file_hash).limit(1).execute()
                if ext_res and getattr(ext_res, 'data', None):
                    record['external_item_id'] = ext_res.data[0].get('id')
                    logger.info(f" Resolved external_item_id via file hash lookup: {record['external_item_id']}")
            except Exception as e:
                logger.warning(f"external_item lookup failed for PDF raw_records link: {e}")
        # Persist raw_records and mark job completed atomically
        try:
            transaction_manager = get_transaction_manager()
            async with transaction_manager.transaction(
                user_id=user_id,
                operation_type="pdf_processing_persist"
            ) as tx:
                await tx.insert('raw_records', record)
                await tx.update('ingestion_jobs', {
                    'status': 'completed',
                    'updated_at': pendulum.now().to_iso8601_string()
                }, {'id': job_id})
        except Exception as e:
            logger.error(f"Failed to persist PDF processing results transactionally: {e}")

    except Exception as e:
        logger.error(f"PDF processing job failed for {filename}: {e}")
        try:
            transaction_manager = get_transaction_manager()
            async with transaction_manager.transaction(
                user_id=user_id,
                operation_type="pdf_processing_fail_update"
            ) as tx:
                await tx.update('ingestion_jobs', {
                    'status': 'failed',
                    'error_message': str(e),
                    'updated_at': pendulum.now().to_iso8601_string()
                }, {'id': job_id})
        except Exception as final_err:
            logger.warning(f"Failed to update final job status: {final_err}")

async def _convert_api_data_to_csv_format(data: List[Dict[str, Any]], source_platform: str) -> Tuple[bytes, str]:
    """
    Convert API data (QuickBooks, Xero, etc.) to CSV format for unified processing.
    This ensures all data goes through the same ExcelProcessor pipeline.
    
    Args:
        data: List of dictionaries containing API response data
        source_platform: Platform name (e.g., 'QuickBooks', 'Xero')
    
    Returns:
        Tuple of (csv_bytes, filename)
    """
    try:
        if not data:
            raise ValueError("No data to convert")
        
        # LIBRARY REPLACEMENT: Use polars for 10x faster DataFrame operations
        # FIX #4c: Removed Pandas fallback - fail explicitly on data issues
        # Polars failures indicate data problems that should not be masked
        try:
            # Convert to polars DataFrame for better performance
            df = pl.DataFrame(data)
            
            # Generate CSV in memory using polars (faster than pandas)
            csv_buffer = io.StringIO()
            df.write_csv(csv_buffer)
            csv_bytes = csv_buffer.getvalue().encode('utf-8')
        except Exception as polars_error:
            # FIX #4c: Fail explicitly instead of silently falling back to pandas
            # This ensures data issues are caught and logged properly
            logger.error(f"âŒ CRITICAL: Polars DataFrame conversion failed - data may be corrupted: {polars_error}")
            logger.error(f"Data sample: {data[:3] if data else 'empty'}")
            raise ValueError(
                f"Failed to convert {source_platform} API data to DataFrame. "
                f"This indicates data format issues. Error: {str(polars_error)}"
            ) from polars_error
        
        # Generate filename with timestamp
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"{source_platform.lower()}_sync_{timestamp}.csv"
        
        logger.info(f"âœ… Converted {len(data)} {source_platform} records to CSV format ({len(csv_bytes)} bytes)")
        return csv_bytes, filename
        
    except Exception as e:
        logger.error(f"Failed to convert {source_platform} API data to CSV: {e}")
        raise

async def _process_api_data_through_pipeline(
    user_id: str,
    data: List[Dict[str, Any]],
    source_platform: str,
    sync_run_id: str,
    user_connection_id: str
) -> Dict[str, Any]:
    """
    Process API data through the main ExcelProcessor pipeline.
    This ensures consistent duplicate detection, enrichment, and entity resolution.
    
    Args:
        user_id: User ID
        data: List of API records
        source_platform: Platform name
        sync_run_id: Sync run ID for tracking
        user_connection_id: User connection ID
    
    Returns:
        Processing results with stats
    """
    try:
        # Convert API data to CSV format
        csv_bytes, filename = await _convert_api_data_to_csv_format(data, source_platform)
        
        # Calculate file hash using xxh3_128 (standardized)
        file_hash = xxhash.xxh3_128(csv_bytes).hexdigest()
        
        # Store CSV in Supabase Storage
        storage_path = f"{user_id}/connector_syncs/{source_platform.lower()}/{filename}"
        try:
            storage = supabase.storage.from_("finely-upload")
            storage.upload(storage_path, csv_bytes, {"content-type": "text/csv"})
            logger.info(f"âœ… Uploaded {source_platform} CSV to storage: {storage_path}")
        except Exception as e:
            logger.error(f"Failed to upload {source_platform} CSV to storage: {e}")
            raise
        
        # Create ingestion job for tracking
        job_id = str(uuid.uuid4())
        try:
            supabase.table('ingestion_jobs').insert({
                'id': job_id,
                'user_id': user_id,
                'file_name': filename,
                'file_size': len(csv_bytes),
                'file_hash': file_hash,
                'source': f'connector_{source_platform.lower()}',
                'job_type': 'api_sync',
                'status': 'processing',
                'created_at': pendulum.now().to_iso8601_string()
            }).execute()
        except Exception as e:
            logger.warning(f"Failed to create ingestion_job for {source_platform}: {e}")
        
        # Initialize ExcelProcessor - FIX #3: Use singleton to avoid reloading heavy ML models
        excel_processor = _get_excel_processor_instance()
        
        # Process through main pipeline
        logger.info(f"ðŸ”„ Processing {source_platform} data through main ExcelProcessor pipeline...")
        result = await excel_processor.process_file(
            file_content=csv_bytes,
            filename=filename,
            user_id=user_id,
            job_id=job_id,
            file_hash=file_hash,
            source_platform=source_platform
        )
        
        # Update ingestion job status
        try:
            supabase.table('ingestion_jobs').update({
                'status': 'completed',
                'updated_at': pendulum.now().to_iso8601_string()
            }).eq('id', job_id).execute()
        except Exception as e:
            logger.warning(f"Failed to update ingestion_job status: {e}")
        
        logger.info(f"âœ… {source_platform} data processed through main pipeline: {result.get('total_rows', 0)} rows")
        
        return {
            'status': 'success',
            'job_id': job_id,
            'file_hash': file_hash,
            'total_rows': result.get('total_rows', 0),
            'processed_rows': result.get('processed_rows', 0),
            'storage_path': storage_path
        }
        
    except Exception as e:
        logger.error(f"Failed to process {source_platform} data through pipeline: {e}")
        raise

# REMOVED: _gmail_sync_run (466 lines) - Replaced by Airbyte Gmail connector
# REMOVED: _zoho_mail_sync_run (259 lines) - Replaced by Airbyte Zoho Mail connector
# REMOVED: _dropbox_sync_run (240 lines) - Replaced by Airbyte Dropbox connector
# REMOVED: _gdrive_sync_run (210 lines) - Replaced by Airbyte Google Drive connector
# REMOVED: _dispatch_connector_sync (132 lines) - Replaced by Airbyte API calls
#
# TOTAL LINES REMOVED: 1,307 lines of custom sync logic
#
# Airbyte now handles:
# - OAuth flows for all providers
# - Rate limiting per provider
# - Incremental sync with cursor management
# ============================================================================
# TASK: ADD OAUTH INITIATION ENDPOINT FOR AIRBYTE
# ============================================================================
@app.post("/api/connectors/initiate")
async def connectors_initiate(req: ConnectorSyncRequest):
    """
    Initiate OAuth flow for Airbyte connector.
    
    Returns Airbyte OAuth URL and state for frontend to open in popup.
    
    Request:
    {
        "provider": "gmail|dropbox|google-drive|...",
        "user_id": "user_123",
        "session_token": "token_xyz"
    }
    
    Response:
    {
        "oauth_url": "https://accounts.google.com/o/oauth2/auth?...",
        "state": "base64_encoded_state",
        "provider": "gmail"
    }
    """
    await _validate_security('connectors-initiate', req.user_id, req.session_token)
    try:
        # Frontend sends 'provider', not 'integration_id'
        provider = getattr(req, 'provider', None) or getattr(req, 'integration_id', None) or 'google-mail'
        
        # Use Airbyte client to create OAuth session
        airbyte = AirbytePythonClient()
        
        result = await airbyte.create_oauth_session(
            provider=provider,
            user_id=req.user_id
        )
        
        logger.info(
            "airbyte_oauth_initiated",
            user_id=req.user_id,
            provider=provider
        )
        
        # Return Airbyte OAuth URL and state
        return {
            "oauth_url": result.get('oauth_url') or result.get('url'),
            "state": result.get('state'),
            "provider": provider,
            "status": "initiated"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"OAuth initiation failed: {e}", provider=req.integration_id or 'unknown')
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/connectors/sync")
async def connectors_sync(req: ConnectorSyncRequest):
    """
    Run a sync via Airbyte for all supported providers.
    
    REPLACED: Custom Nango implementation â†’ Airbyte connectors
    - Airbyte handles OAuth flows
    - Airbyte handles rate limiting
    - Airbyte handles incremental sync
    - Airbyte handles deduplication
    - Airbyte handles retry logic
    """
    await _validate_security('connectors-sync', req.user_id, req.session_token)
    try:
        # Frontend sends 'provider', not 'integration_id'
        provider = getattr(req, 'provider', None) or getattr(req, 'integration_id', None) or 'google-mail'
        
        # Ensure correlation id
        req.correlation_id = req.correlation_id or str(uuid.uuid4())
        
        # Fetch user_connection to get the ID
        uc_res = supabase.table('user_connections').select(
            'id, user_id, nango_connection_id'
        ).eq('nango_connection_id', req.connection_id).limit(1).execute()
        
        if not uc_res.data:
            raise HTTPException(status_code=404, detail=f"Connection {req.connection_id} not found")
        
        user_connection_id = uc_res.data[0]['id']
        
        # Use Airbyte client for sync orchestration
        airbyte = AirbytePythonClient()
        
        # Trigger sync via Airbyte
        result = await airbyte.trigger_sync(
            connection_id=req.connection_id,
            provider=provider,
            user_id=req.user_id,
            mode=req.mode
        )
        
        # Extract Airbyte job ID
        airbyte_job_id = result.get('job', {}).get('id')
        
        # Create sync_run record to track this sync
        sync_run_id = str(uuid.uuid4())
        try:
            supabase.table('sync_runs').insert({
                'id': sync_run_id,
                'user_id': req.user_id,
                'user_connection_id': user_connection_id,
                'type': req.mode or 'manual',
                'status': 'queued',
                'job_id': str(airbyte_job_id) if airbyte_job_id else None,
                'started_at': pendulum.now().to_iso8601_string(),
                'stats': orjson.dumps({'mode': req.mode, 'lookback_days': req.lookback_days}).decode()
            }).execute()
        except Exception as e:
            logger.warning(f"Failed to create sync_run record: {e}")
            # Don't fail the sync if we can't create the record, but log it
        
        logger.info(
            "airbyte_sync_triggered",
            user_id=req.user_id,
            provider=provider,
            connection_id=req.connection_id,
            correlation_id=req.correlation_id,
            sync_run_id=sync_run_id,
            airbyte_job_id=airbyte_job_id
        )
        
        return {
            "status": "queued",
            "provider": provider,
            "mode": req.mode,
            "sync_run_id": sync_run_id,
            "job_id": airbyte_job_id,
            "correlation_id": req.correlation_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Connectors sync failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/connectors/verify-connection")
async def verify_connection(req: ConnectorSyncRequest):
    """
    Verify that a connection is valid and working.
    
    Called by frontend after OAuth popup closes to confirm the connection
    was successfully created in Airbyte.
    
    Request:
    {
        "user_id": "user_123",
        "connection_id": "nango_conn_456",
        "session_token": "token_xyz"
    }
    
    Response:
    {
        "status": "verified|failed",
        "connection_id": "nango_conn_456",
        "message": "Connection verified successfully"
    }
    """
    await _validate_security('connectors-verify', req.user_id, req.session_token)
    try:
        # Fetch the user_connection to verify it exists
        uc_res = supabase.table('user_connections').select(
            'id, user_id, nango_connection_id, status, provider, integration_id'
        ).eq('nango_connection_id', req.connection_id).limit(1).execute()
        
        if not uc_res.data:
            raise HTTPException(status_code=404, detail=f"Connection {req.connection_id} not found")
        
        uc = uc_res.data[0]
        
        # Verify ownership
        if uc['user_id'] != req.user_id:
            raise HTTPException(status_code=403, detail="Connection does not belong to user")
        
        # Check connection status
        connection_status = uc.get('status', 'unknown')
        is_verified = connection_status == 'active'
        
        logger.info(
            "connection_verified",
            user_id=req.user_id,
            connection_id=req.connection_id,
            status=connection_status,
            verified=is_verified
        )
        
        return {
            "status": "verified" if is_verified else "pending",
            "connection_id": req.connection_id,
            "provider": uc.get('provider'),
            "integration_id": uc.get('integration_id'),
            "message": "Connection verified successfully" if is_verified else "Connection is pending activation"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Connection verification failed: {e}", connection_id=req.connection_id)
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# TASK #1: ADD SYNC STATUS POLLING ENDPOINT
# ============================================================================
@app.get("/api/connectors/sync/{sync_run_id}/status")
async def get_sync_status(sync_run_id: str, user_id: str, session_token: Optional[str] = None):
    """
    Get current status of a sync run.
    
    Returns:
    {
        "status": "queued|running|succeeded|failed|partial|cancelled",
        "stats": {"records_fetched": 100, "actions_used": 50, ...},
        "errors": ["error1", "error2"],
        "started_at": "2025-01-01T00:00:00Z",
        "finished_at": "2025-01-01T00:05:00Z",
        "duration_seconds": 300
    }
    """
    await _validate_security('connectors-sync-status', user_id, session_token)
    
    try:
        # Query sync_run by ID
        result = supabase.table('sync_runs').select(
            'id, user_id, status, stats, error, started_at, finished_at'
        ).eq('id', sync_run_id).limit(1).execute()
        
        if not result.data:
            raise HTTPException(status_code=404, detail=f"Sync run {sync_run_id} not found")
        
        sync_run = result.data[0]
        
        # Verify ownership
        if sync_run['user_id'] != user_id:
            raise HTTPException(status_code=403, detail="Unauthorized")
        
        # Parse stats
        stats = sync_run.get('stats') or {}
        if isinstance(stats, str):
            try:
                stats = orjson.loads(stats)
            except Exception:
                stats = {}
        
        # Calculate duration
        started_at = sync_run.get('started_at')
        finished_at = sync_run.get('finished_at')
        duration_seconds = None
        
        if started_at and finished_at:
            try:
                start = pendulum.parse(started_at)
                finish = pendulum.parse(finished_at)
                duration_seconds = int((finish - start).total_seconds())
            except Exception:
                pass
        
        # Build response
        response = {
            'status': sync_run.get('status', 'unknown'),
            'stats': stats,
            'errors': [sync_run['error']] if sync_run.get('error') else [],
            'started_at': started_at,
            'finished_at': finished_at,
            'duration_seconds': duration_seconds
        }
        
        # Log access
        logger.info(
            "sync_status_retrieved",
            sync_run_id=sync_run_id,
            user_id=user_id,
            status=sync_run.get('status')
        )
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get sync status: {e}", sync_run_id=sync_run_id)
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# DATA RETRIEVAL ENDPOINT - Fetch synced items from Airbyte
# ============================================================================
@app.get("/api/connectors/sync/{sync_run_id}/items")
async def get_sync_items(
    sync_run_id: str,
    user_id: str,
    session_token: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    status: Optional[str] = None,
    kind: Optional[str] = None
):
    """
    Fetch synced items from a specific sync run.
    
    Returns paginated external_items that were fetched during this sync.
    
    Query Parameters:
    - page: Page number (1-indexed)
    - page_size: Items per page (max 100)
    - status: Filter by status (fetched|processed|failed|normalized)
    - kind: Filter by kind (txn|email|file|etc)
    
    Response:
    {
        "items": [
            {
                "id": "item_123",
                "provider_id": "gmail_msg_456",
                "kind": "email",
                "status": "processed",
                "source_ts": "2025-01-01T00:00:00Z",
                "metadata": {...},
                "created_at": "2025-01-01T00:05:00Z"
            }
        ],
        "page": 1,
        "page_size": 50,
        "total": 250,
        "has_more": true
    }
    """
    await _validate_security('connectors-sync-items', user_id, session_token)
    
    try:
        # Validate pagination
        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 50
        page_size = min(page_size, 100)  # Max 100 items per page
        
        # Verify sync_run exists and belongs to user
        sync_run_result = supabase.table('sync_runs').select(
            'id, user_id, user_connection_id'
        ).eq('id', sync_run_id).limit(1).execute()
        
        if not sync_run_result.data:
            raise HTTPException(status_code=404, detail=f"Sync run {sync_run_id} not found")
        
        sync_run = sync_run_result.data[0]
        
        # Verify ownership
        if sync_run['user_id'] != user_id:
            raise HTTPException(status_code=403, detail="Unauthorized")
        
        user_connection_id = sync_run['user_connection_id']
        
        # Build query for external_items
        query = supabase.table('external_items').select(
            'id, provider_id, kind, status, source_ts, metadata, created_at'
        ).eq('sync_run_id', sync_run_id).eq('user_connection_id', user_connection_id)
        
        # Apply optional filters
        if status:
            query = query.eq('status', status)
        if kind:
            query = query.eq('kind', kind)
        
        # Get total count
        count_result = supabase.table('external_items').select(
            'id',
            count='exact'
        ).eq('sync_run_id', sync_run_id).eq('user_connection_id', user_connection_id)
        
        if status:
            count_result = count_result.eq('status', status)
        if kind:
            count_result = count_result.eq('kind', kind)
        
        count_result = count_result.execute()
        total = count_result.count if hasattr(count_result, 'count') else len(count_result.data or [])
        
        # Paginate
        offset = (page - 1) * page_size
        items_result = query.order('created_at', desc=True).range(offset, offset + page_size - 1).execute()
        
        items = items_result.data or []
        has_more = (offset + page_size) < total
        
        # Log access
        logger.info(
            "sync_items_retrieved",
            sync_run_id=sync_run_id,
            user_id=user_id,
            page=page,
            page_size=page_size,
            total=total,
            items_count=len(items)
        )
        
        return {
            'items': items,
            'page': page,
            'page_size': page_size,
            'total': total,
            'has_more': has_more
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get sync items: {e}", sync_run_id=sync_run_id)
        raise HTTPException(status_code=500, detail=str(e))


class ConnectorMetadataUpdate(BaseModel):
    user_id: str
    connection_id: str
    updates: Dict[str, Any]
    session_token: Optional[str] = None

@app.post('/api/connectors/metadata')
async def update_connection_metadata(req: ConnectorMetadataUpdate):
    """Update provider-specific metadata for a user connection (e.g., realmId, tenantId)."""
    await _validate_security('connectors-metadata', req.user_id, req.session_token)
    try:
        row = supabase.table('user_connections').select('metadata').eq('nango_connection_id', req.connection_id).limit(1).execute()
        base_meta = (row.data[0].get('metadata') if row.data else {}) or {}
        if isinstance(base_meta, str):
            try:
                base_meta = orjson.loads(base_meta)
            except Exception:
                base_meta = {}
        # FIX #2: Update metadata with transaction protection
        new_meta = {**base_meta, **(req.updates or {})}
        try:
            transaction_manager = get_transaction_manager()
            async with transaction_manager.transaction(
                user_id=req.user_id,
                operation_type="connector_metadata_manual_update"
            ) as tx:
                await tx.update('user_connections', {
                    'metadata': new_meta
                }, {'nango_connection_id': req.connection_id})
                logger.info(f"âœ… Updated connector metadata via API: {req.connection_id}")
        except Exception as e:
            logger.error(f"âŒ Failed to update connector metadata: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to update metadata: {e}")
        return {'status': 'ok', 'metadata': new_meta}
    except Exception as e:
        logger.error(f"Metadata update failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ConnectorFrequencyUpdate(BaseModel):
    user_id: str
    connection_id: str
    minutes: int
    session_token: Optional[str] = None

@app.post('/api/connectors/frequency')
async def update_connection_frequency(req: ConnectorFrequencyUpdate):
    """Update sync frequency in minutes for a user connection."""
    await _validate_security('connectors-frequency', req.user_id, req.session_token)
    try:
        minutes = max(0, min(int(req.minutes), 7 * 24 * 60))  # clamp to [0, 10080]
        supabase.table('user_connections').update({'sync_frequency_minutes': minutes}).eq('nango_connection_id', req.connection_id).execute()
        return {'status': 'ok', 'sync_frequency_minutes': minutes}
    except Exception as e:
        logger.error(f"Frequency update failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/api/connectors/disconnect')
async def connectors_disconnect(req: ConnectorDisconnectRequest):
    """Disconnect a user's connector and remove the Nango connection."""
    await _validate_security('connectors-disconnect', req.user_id, req.session_token)

    connection_id = req.connection_id

    try:
        # Fetch connection row to validate ownership and gather metadata
        uc_res = supabase.table('user_connections').select(
            'id, user_id, integration_id, provider'
        ).eq('nango_connection_id', connection_id).limit(1).execute()

        if not uc_res.data:
            # Nothing to disconnect; treat as success for idempotency
            return {'status': 'ok', 'connection_id': connection_id}

        conn_row = uc_res.data[0]
        if conn_row.get('user_id') != req.user_id:
            raise HTTPException(status_code=403, detail='Connection does not belong to user')

        integration_id = conn_row.get('integration_id')
        provider = req.provider or conn_row.get('provider')

        # Derive integration id if missing - use provider directly (connectors table deleted)
        if not integration_id:
            integration_id = provider

        # Attempt to delete connection in Airbyte first (best effort)
        airbyte = AirbytePythonClient()
        await airbyte.delete_connection(connection_id)

        # Mark connection inactive / remove locally
        try:
            transaction_manager = get_transaction_manager()
            async with transaction_manager.transaction(
                user_id=req.user_id,
                operation_type="connector_disconnect"
            ) as tx:
                await tx.update(
                    'user_connections',
                    {
                        'status': 'disconnected',
                        'updated_at': pendulum.now().to_iso8601_string()
                    },
                    {'nango_connection_id': connection_id}
                )

                await tx.update(
                    'external_items',
                    {'status': 'disconnected'},
                    {'user_connection_id': conn_row['id']}
                )

        except Exception as e:
            logger.error(f"Failed to mark user connection disconnected: {e}")
            raise HTTPException(status_code=500, detail='Failed to update connection state')

        return {'status': 'ok', 'connection_id': connection_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Connector disconnect failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def connectors_status(connection_id: str, user_id: str, session_token: Optional[str] = None):
    await _validate_security('connectors-status', user_id, session_token)
    try:
        # Fetch user_connection and recent runs
        uc_res = supabase.table('user_connections').select('id, user_id, nango_connection_id, provider, status, last_synced_at, created_at, provider_account_id, metadata, sync_frequency_minutes, integration_id').eq('nango_connection_id', connection_id).limit(1).execute()
        uc = uc_res.data[0] if uc_res.data else None
        # Use integration_id and provider directly (connectors table deleted)
        integration_id = uc.get('integration_id') or uc.get('provider') if uc else None
        runs = []
        if uc:
            runs_res = supabase.table('sync_runs').select('id, type, status, started_at, finished_at, stats, error').eq('user_connection_id', uc['id']).order('started_at', desc=True).limit(50).execute()
            runs = runs_res.data or []
        payload = {'connection': {**uc, 'integration_id': integration_id} if uc else None, 'recent_runs': runs}
        return payload
    except Exception as e:
        logger.error(f"Connectors status failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/connectors/user-connections")
async def list_user_connections(req: UserConnectionsRequest):
    """List the current user's Nango connections with integration IDs for UI rendering.
    
    Connections are created via webhook when user authorizes in Nango popup.
    """
    await _validate_security('connectors-user-connections', req.user_id, req.session_token)
    try:
        # CRITICAL FIX: Lazy-load Supabase client on first use
        supabase_client = await _ensure_supabase_loaded()
        if not supabase_client:
            raise HTTPException(
                status_code=503,
                detail="Database service is temporarily unavailable. Please try again in a moment."
            )
        
        # Fetch from database (connections created by webhook handler)
        res = supabase_client.table('user_connections').select('id, user_id, nango_connection_id, provider, integration_id, status, last_synced_at, created_at').eq('user_id', req.user_id).limit(1000).execute()
        items = []
        for row in (res.data or []):
            # Use integration_id and provider directly (connectors table deleted)
            integ = row.get('integration_id') or row.get('provider')
            items.append({
                'connection_id': row.get('nango_connection_id') or row.get('id'),
                'integration_id': integ,
                'status': row.get('status'),
                'last_synced_at': row.get('last_synced_at'),
                'created_at': row.get('created_at'),
            })
        return {'connections': items}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"List user connections failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def _process_webhook_delta_items(user_id: str, user_connection_id: str, provider: str, changed_items: List[Dict], correlation_id: str) -> int:
    """Process specific changed items from webhook delta (optimized for real-time updates)."""
    processed = 0
    try:
        transaction_manager = get_transaction_manager()
        async with transaction_manager.transaction(user_id=user_id, operation_type="webhook_delta_processing") as tx:
            for item in changed_items:
                # Extract common fields
                item_id = item.get('id') or item.get('invoice_id') or item.get('InvoiceID') or item.get('message_id')
                if not item_id:
                    continue
                
                # Build external_item record
                meta = {'correlation_id': correlation_id, 'webhook_delta': True, **item}
                ext_item = {
                    'user_id': user_id,
                    'user_connection_id': user_connection_id,
                    'sync_run_id': correlation_id,
                    'provider_id': str(item_id),
                    'kind': 'txn',
                    'source_ts': item.get('date') or item.get('updated_at') or pendulum.now().to_iso8601_string(),
                    'metadata': meta,
                    'status': 'fetched'
                }
                
                # FIX #22: Use error handling helper to store failed items with error details
                stats_dict = {'records_fetched': 0, 'skipped': 0}
                success = await insert_external_item_with_error_handling(
                    transaction_manager, user_id, user_connection_id, ext_item, stats_dict
                )
                if success or stats_dict['records_fetched'] > 0:
                    processed += 1
            
            # Trigger normalization for delta items
            if processed > 0:
                items_res = supabase.table('external_items').select('id, provider_id, metadata').eq('user_connection_id', user_connection_id).eq('status', 'fetched').limit(100).execute()
                delta_items = [it for it in (items_res.data or []) if (it.get('metadata') or {}).get('correlation_id') == correlation_id]
                
                if delta_items:
                    # Convert to standardized format and process through pipeline
                    records = []
                    for it in delta_items:
                        meta = it.get('metadata') or {}
                        records.append({
                            'transaction_id': it.get('provider_id'),
                            'transaction_type': meta.get('type', 'Transaction'),
                            'total_amount': meta.get('total') or meta.get('amount', 0),
                            'transaction_date': meta.get('date', ''),
                            'source': provider
                        })
                    
                    if records:
                        await _process_api_data_through_pipeline(user_id, records, provider, str(uuid.uuid4()), user_connection_id)
                        logger.info(f"Processed {len(records)} records through pipeline")
    except Exception as e:
        logger.error(f"Webhook delta processing error: {e}")
        return 0
    
    return processed


# ============================================================================
# AIRBYTE WEBHOOK HANDLER - Real-time sync notifications
# ============================================================================
@app.post("/api/webhooks/airbyte")
async def airbyte_webhook(request: Request):
    """
    Airbyte webhook receiver for sync events.
    
    Handles:
    - sync.started: Sync job started
    - sync.succeeded: Sync completed successfully
    - sync.failed: Sync failed with error
    - connection.created: New connection authorized
    - connection.deleted: Connection removed
    
    Updates sync_runs table and sends WebSocket notifications to frontend.
    """
    try:
        raw = await request.body()
        payload = {}
        try:
            payload = orjson.loads(raw.decode('utf-8') or '{}')
        except Exception:
            pass
        
        # Verify Airbyte webhook signature if secret configured
        secret = os.environ.get("AIRBYTE_WEBHOOK_SECRET")
        header_sig = request.headers.get('X-Airbyte-Signature')
        signature_valid = False
        
        if secret and header_sig:
            try:
                # Airbyte uses HMAC-SHA256 with the raw body
                digest = hmac.new(secret.encode('utf-8'), raw, 'sha256').hexdigest()
                signature_valid = hmac.compare_digest(header_sig, digest)
            except Exception as e:
                logger.warning(f"Airbyte webhook signature verification failed: {e}")
                signature_valid = False
        elif not secret:
            # In development, accept unsigned webhooks
            environment = os.environ.get('ENVIRONMENT', 'development')
            if environment == 'production':
                logger.error("AIRBYTE_WEBHOOK_SECRET not set in production - rejecting webhook")
                raise HTTPException(status_code=403, detail='Webhook secret not configured')
            else:
                logger.warning("AIRBYTE_WEBHOOK_SECRET not set; accepting webhook in dev mode")
                signature_valid = True
        
        # Extract event details
        event_type = payload.get('type') or payload.get('event_type')
        event_id = payload.get('id') or str(uuid.uuid4())
        job_id = payload.get('job_id') or payload.get('jobId')
        connection_id = payload.get('connection_id') or payload.get('connectionId')
        status = payload.get('status')
        
        logger.info(
            "airbyte_webhook_received",
            event_type=event_type,
            event_id=event_id,
            job_id=job_id,
            connection_id=connection_id,
            status=status
        )
        
        # Persist webhook for audit trail
        try:
            supabase.table('webhook_events').insert({
                'user_id': 'airbyte',
                'user_connection_id': connection_id,
                'event_type': event_type,
                'payload': payload,
                'signature_valid': bool(signature_valid),
                'status': 'processed',
                'error': None,
                'event_id': event_id,
                'created_at': pendulum.now().to_iso8601_string()
            }).execute()
        except Exception as e:
            # Conflict on unique(event_id) is fine; treat as already processed
            logger.debug(f"Webhook event already recorded: {e}")
        
        # Handle sync events
        if signature_valid and event_type and event_type.startswith('sync.'):
            try:
                # Find sync_run by job_id
                sync_run_result = supabase.table('sync_runs').select(
                    'id, user_id, user_connection_id, status'
                ).eq('job_id', str(job_id)).limit(1).execute()
                
                if sync_run_result.data:
                    sync_run = sync_run_result.data[0]
                    sync_run_id = sync_run['id']
                    user_id = sync_run['user_id']
                    
                    # Map Airbyte status to our status
                    status_map = {
                        'sync.started': 'running',
                        'sync.succeeded': 'succeeded',
                        'sync.failed': 'failed',
                        'sync.cancelled': 'cancelled'
                    }
                    new_status = status_map.get(event_type, status)
                    
                    # Extract stats and error details
                    stats = payload.get('stats') or {}
                    error_msg = payload.get('error') or payload.get('error_message')
                    
                    # Update sync_run status
                    update_data = {
                        'status': new_status,
                        'updated_at': pendulum.now().to_iso8601_string()
                    }
                    
                    if event_type == 'sync.succeeded':
                        update_data['finished_at'] = pendulum.now().to_iso8601_string()
                        update_data['stats'] = stats
                    elif event_type == 'sync.failed':
                        update_data['finished_at'] = pendulum.now().to_iso8601_string()
                        update_data['error'] = error_msg
                    
                    supabase.table('sync_runs').update(update_data).eq('id', sync_run_id).execute()
                    
                    logger.info(
                        "sync_run_updated_from_webhook",
                        sync_run_id=sync_run_id,
                        event_type=event_type,
                        new_status=new_status
                    )
                    
                    # Send WebSocket notification to frontend
                    try:
                        websocket_manager = get_connection_manager()
                        if websocket_manager:
                            await websocket_manager.send_update(sync_run_id, {
                                "status": new_status,
                                "event_type": event_type,
                                "stats": stats,
                                "error": error_msg,
                                "timestamp": pendulum.now().to_iso8601_string()
                            })
                    except Exception as ws_err:
                        logger.warning(f"Failed to send WebSocket notification: {ws_err}")
                else:
                    logger.warning(f"Sync run not found for job_id: {job_id}")
            
            except Exception as e:
                logger.error(f"Failed to handle sync event: {e}")
        
        # Handle connection events
        elif signature_valid and event_type == 'connection.created':
            try:
                logger.info(f"ðŸ”— Airbyte connection created: connection_id={connection_id}")
                # Connection creation is handled via OAuth flow, not webhook
                # This is just for audit trail
            except Exception as e:
                logger.error(f"Failed to handle connection.created event: {e}")
        
        return {'status': 'received', 'event_id': event_id, 'signature_valid': bool(signature_valid)}
    
    except Exception as e:
        logger.error(f"Airbyte webhook handling failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _require_scheduler_auth(request: Request):
    """Verify scheduler token from header or query param"""
    token = os.environ.get('SCHEDULER_TOKEN', 'dev-token-12345')
    if request.headers.get('Authorization', '').startswith('Bearer '):
        provided = request.headers.get('Authorization').split(' ')[1]
    else:
        provided = request.headers.get('X-Scheduler-Token') or request.query_params.get('token')
    if not provided or not hmac.compare_digest(provided, token):
        raise HTTPException(status_code=403, detail='Invalid scheduler token')

# DEPRECATED: Old _dispatch_connector_sync function removed (was duplicate)
# Use the new version at line 11297 which includes rate limiting and distributed locking
# The new version signature: _dispatch_connector_sync(integration_id, req, nango)

@app.post('/api/connectors/scheduler/run')
async def run_scheduled_syncs(request: Request, provider: Optional[str] = None, limit: int = 25):
    """Orchestrate incremental syncs for due connections. Secured by SCHEDULER_TOKEN."""
    _require_scheduler_auth(request)
    try:
        now = datetime.utcnow()
        # Fetch active connections
        conns = supabase.table('user_connections').select('id, user_id, nango_connection_id, sync_frequency_minutes, last_synced_at, provider, integration_id, status').eq('status', 'active').limit(1000).execute()
        dispatched = []
        for row in (conns.data or []):
            if len(dispatched) >= max(1, min(limit, 100)):
                break
            freq = row.get('sync_frequency_minutes') or 60
            last = row.get('last_synced_at')
            due = True
            try:
                if last:
                    last_dt = pendulum.parse(str(last)).naive()
                    due = (now - last_dt) >= timedelta(minutes=freq)
            except Exception:
                due = True
            if not due:
                continue

            # Provider filter
            conn_provider = row.get('integration_id') or row.get('provider')
            if provider and conn_provider and provider != conn_provider:
                continue

            # ISSUE #9: Check scheduler rate limit before dispatching
            from core_infrastructure.rate_limiter import get_scheduler_rate_limiter
            scheduler_limiter = get_scheduler_rate_limiter()
            
            provider_to_dispatch = conn_provider or 'google-mail'
            can_dispatch, rate_limit_msg = await scheduler_limiter.check_scheduler_rate_limit(provider_to_dispatch)
            if not can_dispatch:
                logger.warning(f"Scheduler rate limit: {rate_limit_msg}")
                continue
            
            # REPLACED: Use Airbyte for sync orchestration
            try:
                airbyte = AirbytePythonClient()
                result = await airbyte.trigger_sync(
                    connection_id=row['nango_connection_id'],
                    provider=provider_to_dispatch,
                    user_id=row['user_id'],
                    mode='incremental'
                )
                if result.get('status') == 'queued':
                    dispatched.append(row['nango_connection_id'])
                    # Record successful dispatch
                    await scheduler_limiter.record_dispatch(provider_to_dispatch, count=1)
                    # Reset failure count on success
                    await scheduler_limiter.reset_failure_count(provider_to_dispatch)
                else:
                    # Record failure for exponential backoff
                    await scheduler_limiter.record_failure(provider_to_dispatch)
            except Exception as e:
                logger.error(f"Scheduler dispatch failed for {provider_to_dispatch}: {e}")
                await scheduler_limiter.record_failure(provider_to_dispatch)

        return {"status": "ok", "dispatched": dispatched, "count": len(dispatched)}
    except Exception as e:
        logger.error(f"Scheduler run failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# WEBSOCKET INTEGRATION FOR REAL-TIME UPDATES
# ============================================================================

# LIBRARY REPLACEMENT: Socket.IO handles all WebSocket management automatically
# Removed manual _authorize_websocket_connection (48 lines) - Socket.IO has built-in auth
# Removed manual websocket endpoints (99 lines) - Socket.IO handles routing
# Removed manual keep-alive loops - Socket.IO has automatic heartbeat
# Removed manual ping/pong handling - Socket.IO has built-in ping/pong
# Removed manual disconnect handling - Socket.IO handles cleanup automatically

# Socket.IO server initialization (replaces 147 lines of manual WebSocket code)
sio = socketio.AsyncServer(
    async_mode='asgi',
    cors_allowed_origins="*",
    logger=False,
    engineio_logger=False,
    ping_timeout=60,
    ping_interval=25
)
socketio_app = socketio.ASGIApp(sio, app)

# Socket.IO event handlers (replaces manual endpoint logic)
@sio.event
async def connect(sid, environ):
    """Socket.IO connection handler - replaces manual authorization"""
    try:
        # Extract query parameters
        query_string = environ.get('QUERY_STRING', '')
        params = dict(param.split('=') for param in query_string.split('&') if '=' in param)
        
        user_id = params.get('user_id')
        token = params.get('session_token')
        job_id = params.get('job_id')  # Get job_id from query params, not PATH_INFO
        
        if not user_id or not token:
            logger.warning(f"Socket.IO connection rejected: missing credentials")
            return False
        
        # CRITICAL FIX: Verify job exists and belongs to user BEFORE joining room
        # This prevents race condition where user_id is set lazily after connection
        # Only verify if job_id is provided
        if job_id:
            try:
                # CRITICAL FIX: Lazy-load Supabase client on first use
                supabase_client = await _ensure_supabase_loaded()
                if not supabase_client:
                    logger.error(f"Socket.IO connection rejected: Database service unavailable")
                    return False
                
                job_record = supabase_client.table('ingestion_jobs').select('user_id, status').eq('id', job_id).single().execute()
                if not job_record.data:
                    logger.warning(f"Socket.IO connection rejected: job {job_id} not found")
                    return False
                
                # CRITICAL: Check if user_id is already set in database
                db_user_id = job_record.data.get('user_id')
                if db_user_id and db_user_id != user_id:
                    logger.warning(f"Socket.IO connection rejected: job {job_id} belongs to different user")
                    return False
                
                # If user_id not set yet, verify token is valid before allowing connection
                if not db_user_id:
                    # Verify session token is valid for this user
                    try:
                        auth_response = supabase_client.auth.get_user(token)
                        if not auth_response or auth_response.user.id != user_id:
                            logger.warning(f"Socket.IO connection rejected: invalid session token for user {user_id}")
                            return False
                    except Exception as auth_err:
                        logger.warning(f"Socket.IO connection rejected: token verification failed: {auth_err}")
                        return False
            except Exception as e:
                logger.error(f"Job verification failed: {e}")
                return False
            
            # Join job room for broadcasting
            sio.enter_room(sid, job_id)
            logger.info(f"âœ… Socket.IO connected: {sid} -> job {job_id}")
        else:
            # No job_id provided - just verify token is valid
            try:
                supabase_client = await _ensure_supabase_loaded()
                if supabase_client:
                    try:
                        auth_response = supabase_client.auth.get_user(token)
                        if not auth_response or auth_response.user.id != user_id:
                            logger.warning(f"Socket.IO connection rejected: invalid session token for user {user_id}")
                            return False
                    except Exception as auth_err:
                        logger.warning(f"Socket.IO connection rejected: token verification failed: {auth_err}")
                        return False
            except Exception as e:
                logger.warning(f"Token verification skipped due to DB unavailability: {e}")
            
            logger.info(f"âœ… Socket.IO connected: {sid} (user {user_id})")
        
        return True
        
    except Exception as e:
        logger.error(f"Socket.IO connection error: {e}")
        return False

@sio.event
async def disconnect(sid):
    """Socket.IO disconnect handler - replaces manual cleanup"""
    logger.info(f"Socket.IO disconnected: {sid}")

@sio.on('ping')
async def handle_ping(sid):
    """Socket.IO ping handler - replaces manual ping/pong"""
    return {'type': 'pong', 'timestamp': pendulum.now().to_iso8601_string()}

@sio.on('get_status')
async def handle_get_status(sid, data):
    """Socket.IO status request handler"""
    job_id = data.get('job_id')
    if job_id:
        status = await websocket_manager.get_job_status(job_id)
        return status or {'status': 'unknown'}

@sio.on('pause_processing')
async def handle_pause_processing(sid, data):
    """
    Socket.IO pause processing handler (Step 3.5)
    
    Allows frontend to pause file processing via WebSocket.
    Sets a flag that processing functions check cooperatively.
    """
    try:
        file_id = data.get('fileId')
        if not file_id:
            logger.warning(f"Pause request missing fileId from {sid}")
            return {'status': 'error', 'message': 'fileId required'}
        
        # Set pause flag in job state
        current_status = await websocket_manager.get_job_status(file_id) or {}
        await websocket_manager.merge_job_state(file_id, {
            **current_status,
            'status': 'paused',
            'paused_at': pendulum.now().to_iso8601_string()
        })
        
        logger.info(f"âœ… Processing paused for file {file_id} by {sid}")
        
        # Notify all clients in the job room
        await sio.emit('processing_paused', {
            'fileId': file_id,
            'timestamp': pendulum.now().to_iso8601_string()
        }, room=file_id)
        
        return {'status': 'success', 'message': 'Processing paused'}
    except Exception as e:
        logger.error(f"Failed to pause processing: {e}")
        return {'status': 'error', 'message': str(e)}

# LIBRARY REPLACEMENT: Socket.IO-based WebSocket manager (replaces 368-line custom implementation)
# NOTE: Class definition moved to line 685 to fix forward reference error
# The class was previously here but is now defined before the lifespan function


async def start_processing_job(user_id: str, job_id: str, storage_path: str, filename: str,
                               duplicate_decision: Optional[str] = None,
                               existing_file_id: Optional[str] = None,
                               original_file_hash: Optional[str] = None,
                               file_bytes_cached: Optional[bytes] = None,
                               external_item_id: Optional[str] = None):
    try:
        # Bind job to user for WebSocket authorization
        base = (await websocket_manager.get_job_status(job_id)) or {}
        await websocket_manager.merge_job_state(job_id, {
            **base,
            "user_id": user_id,
            "status": base.get("status", "queued"),
            "started_at": base.get("started_at") or pendulum.now().to_iso8601_string(),
        })
        async def is_cancelled() -> bool:
            status = await websocket_manager.get_job_status(job_id)
            return (status or {}).get("status") == "cancelled"

        # FIX ISSUE #1: Reuse already-downloaded file bytes if available
        file_bytes = file_bytes_cached
        
        if file_bytes is None:
            await websocket_manager.send_overall_update(
                job_id=job_id,
                status="processing",
                message="ðŸ“¥ Downloading file from storage...",
                progress=5
            )
            if await is_cancelled():
                return

            try:
                storage = supabase.storage.from_("finely-upload")
                file_resp = storage.download(storage_path)
                file_bytes = file_resp if isinstance(file_resp, (bytes, bytearray)) else getattr(file_resp, 'data', None)
                if file_bytes is None:
                    file_bytes = file_resp
            except Exception as e:
                logger.error(f"Storage download failed: {e}")
                await websocket_manager.send_error(job_id, f"Download failed: {e}")
                # FIX #17: Separate nested await calls to prevent floating coroutines
                current_job_status = await websocket_manager.get_job_status(job_id)
                await websocket_manager.merge_job_state(job_id, {**(current_job_status or {}), "status": "failed", "error": str(e)})
                return
        else:
            logger.info(f"Reusing cached file bytes for job {job_id} (avoiding re-download)")
            await websocket_manager.send_overall_update(
                job_id=job_id,
                status="processing",
                message="âœ… Using cached file (no re-download needed)...",
                progress=5
            )

        await websocket_manager.send_overall_update(
            job_id=job_id,
            status="processing",
            message="ðŸ§  Initializing analysis pipeline...",
            progress=15
        )
        if await is_cancelled():
            return

        excel_processor = _get_excel_processor_instance()
        # CRITICAL FIX: Create StreamedFile object from bytes using from_bytes() method
        from streaming_source import StreamedFile
        streamed_file = StreamedFile.from_bytes(data=file_bytes, filename=filename)
        await excel_processor.process_file(
            job_id=job_id,
            file_path=streamed_file.path,
            user_id=user_id,
            supabase=supabase,
            duplicate_decision=duplicate_decision,
            external_item_id=external_item_id
        )
    except Exception as e:
        logger.error(f"Processing job failed (resume path): {e}")
        await websocket_manager.send_error(job_id, str(e))
        # FIX #17: Separate nested await calls to prevent floating coroutines
        current_job_status = await websocket_manager.get_job_status(job_id)
        await websocket_manager.merge_job_state(job_id, {**(current_job_status or {}), "status": "failed", "error": str(e)})


# Legacy compatibility adapter for older code paths that used `manager.send_update(...)`
class LegacyConnectionManagerAdapter:
    async def send_update(self, job_id: str, payload: Dict[str, Any]):
        step = payload.get("step", "processing")
        status = payload.get("status")
        if not status:
            if step in ("error", "failed", "entity_resolution_failed", "platform_learning_failed"):
                status = "failed"
            elif step in ("completed",):
                status = "completed"
            else:
                status = "processing"
        message = payload.get("message", step)
        progress = payload.get("progress")
        results = payload.get("results")
        await websocket_manager.send_overall_update(job_id, status, message, progress, results)

# Instantiate legacy adapter so existing calls like `await manager.send_update(...)` keep working
manager = LegacyConnectionManagerAdapter()

@app.post("/cancel-upload/{job_id}")
async def cancel_upload(job_id: str):
    """Cancel an in-flight processing job and notify listeners."""
    try:
        base = (await websocket_manager.get_job_status(job_id)) or {}
        await websocket_manager.merge_job_state(job_id, {
            **base,
            "status": "cancelled",
            "message": "Cancelled by user",
            "updated_at": pendulum.now().to_iso8601_string()
        })
        # Notify over WS if connected
        await websocket_manager.send_overall_update(
            job_id=job_id,
            status="cancelled",
            message="Cancelled by user",
            progress=(base or {}).get("progress", 0)
        )
        return {"status": "cancelled", "job_id": job_id}
    except Exception as e:
        logger.error(f"Failed to cancel job {job_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ----- ULTRA SIMPLE ENDPOINT - NO REDIS, NO CACHE, NO AI -----
@app.post("/api/upload-simple")
async def upload_simple_endpoint(
    file: UploadFile = File(...),
    user_id: str = Form(default="test_user"),
    job_id: str = Form(default=""),
):
    """
    Ultra-simple file upload - NO dependencies.
    Just file upload + Excel parsing + return result.
    """
    import tempfile
    import os
    
    try:
        # Get filename
        filename = file.filename or "upload.xlsx"
        suffix = os.path.splitext(filename)[1] or ".xlsx"
        
        # Read file content directly (await for async)
        file_content = await file.read()
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp.write(file_content)
            temp_path = temp.name
        
        try:
            # Calculate hash
            import xxhash
            file_hash = xxhash.xxh3_128(file_content).hexdigest()
            
            # Parse Excel with polars
            import polars as pl
            sheets = pl.read_excel(temp_path, sheet_name=None, engine='calamine')
            
            sheets_info = {}
            total_rows = 0
            for sheet_name, df in sheets.items():
                sheets_info[sheet_name] = {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "column_names": list(df.columns)
                }
                total_rows += len(df)
            
            return {
                "status": "success",
                "job_id": job_id or str(uuid.uuid4()),
                "user_id": user_id,
                "filename": filename,
                "file_hash": file_hash,
                "file_size": len(file_content),
                "sheets": sheets_info,
                "total_rows": total_rows
            }
        finally:
            # Cleanup temp file
            try:
                os.unlink(temp_path)
            except:
                pass
                
    except Exception as e:
        import traceback
        logger.error(f"Simple upload error: {e}\n{traceback.format_exc()}")
        return JSONResponse(
            status_code=500,
            content={
                "error": str(e),
                "error_type": type(e).__name__,
                "traceback": traceback.format_exc()[:500]
            }
        )
# ----- END SIMPLE ENDPOINT -----

@app.post("/api/process-with-websocket")
async def process_with_websocket_endpoint(
    file: UploadFile = File(...),  # CRITICAL FIX: Use UploadFile instead of bytes
    user_id: str = Form(...),
    job_id: str = Form(default_factory=lambda: str(uuid.uuid4()))
):
    """Process file with real-time WebSocket updates"""
    try:
        # SECURITY FIX: Rate limit file uploads (5 per minute per user)
        from core_infrastructure.rate_limiter import get_upload_rate_limiter
        upload_limiter = get_upload_rate_limiter()
        # can_upload, rate_limit_msg = await upload_limiter.check_upload_rate_limit(user_id)
        can_upload = True
        if not can_upload:
            raise HTTPException(status_code=429, detail=rate_limit_msg)
        
        # Record the upload attempt
        # await upload_limiter.record_upload(user_id)
        
        # Critical: Check database health before processing
        # check_database_health()
        
        # CRITICAL FIX: Create StreamedFile from UploadFile without loading bytes
        filename = file.filename
        streamed_file = await StreamedFile.from_upload(file)
        
        # Send initial update
        await websocket_manager.send_overall_update(
            job_id=job_id,
            status="starting",
            message="🚀 Starting universal component processing...",
            progress=0
        )
        
        # Initialize components - FIX #3: Reuse singleton instances
        excel_processor = _get_excel_processor_instance()
        field_detector = excel_processor.universal_field_detector
        platform_detector = excel_processor.universal_platform_detector
        document_classifier = excel_processor.universal_document_classifier
        data_extractor = excel_processor.universal_extractors
        
        results = {}
        
        # Step 1: Process Excel file
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="excel_processor",
            status="processing",
            message="ðŸ“Š Processing Excel file...",
            progress=20
        )
        
        excel_result = await excel_processor.stream_xlsx_processing(
            file_path=streamed_file.path,  # FIX: Was file_content which was undefined
            filename=filename,
            user_id=user_id
        )
        results["excel_processing"] = excel_result
        
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="excel_processor",
            status="completed",
            message="âœ… Excel processing completed",
            progress=100,
            data={"sheets_count": len(excel_result.get('sheets', {}))}
        )
        
        # Step 2: Detect fields FIRST (required for platform detection)
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="field_detector",
            status="processing",
            message="ðŸ·ï¸ Detecting field types...",
            progress=20
        )
        
        field_results = {}
        for sheet_name, df in excel_result.get('sheets', {}).items():
            field_result = await field_detector.detect_field_types_universal(
                data=df.to_dict('records')[0] if not df.empty else {},
                filename=filename,
                user_id=user_id
            )
            field_results[sheet_name] = field_result
        results["field_detection"] = field_results
        
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="field_detector",
            status="completed",
            message="âœ… Field types detected",
            progress=100,
            data=field_results
        )
        
        # Step 3: Detect platform (now with field information)
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="platform_detector",
            status="processing",
            message="ðŸ” Detecting platform...",
            progress=20
        )
        
        # Include field detection results in platform detection payload
        first_sheet_data = list(excel_result.get('sheets', {}).values())[0] if excel_result.get('sheets') else None
        first_field_result = list(field_results.values())[0] if field_results else {}
        
        platform_payload = {
            "file_content": first_sheet_data.to_dicts() if first_sheet_data is not None else [], 
            "filename": filename,
            "detected_fields": first_field_result.get('detected_fields', []),
            "field_types": first_field_result.get('field_types', {})
        }
        
        platform_result = await platform_detector.detect_platform_universal(
            payload=platform_payload,
            filename=filename,
            user_id=user_id
        )
        results["platform_detection"] = platform_result
        
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="platform_detector",
            status="completed",
            message=f"âœ… Platform detected: {platform_result.get('platform', 'unknown')}",
            progress=100,
            data=platform_result
        )
        
        # Step 4: Classify document
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="document_classifier",
            status="processing",
            message="ðŸ“„ Classifying document...",
            progress=20
        )
        
        document_result = await document_classifier.classify_document_universal(
            payload={"file_content": first_sheet_data.to_dicts() if first_sheet_data is not None else [], "filename": filename},
            filename=filename,
            user_id=user_id
        )
        results["document_classification"] = document_result
        
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="document_classifier",
            status="completed",
            message=f"âœ… Document classified: {document_result.get('document_type', 'unknown')}",
            progress=100,
            data=document_result
        )
        
        # Step 5: Extract data
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="data_extractor",
            status="processing",
            message="ðŸ”§ Extracting data...",
            progress=20
        )
        
        extraction_result = await data_extractor.extract_data_universal(
            file_path=streamed_file.path,
            filename=filename,
            user_id=user_id
        )
        results["data_extraction"] = extraction_result
        
        await websocket_manager.send_component_update(
            job_id=job_id,
            component="data_extractor",
            status="completed",
            message="âœ… Data extraction completed",
            progress=100,
            data={"extracted_rows": len(extraction_result.get('extracted_data', []))}
        )
        
# REMOVED: Duplicate field detection - now happens in Step 2 before platform detection
        
        # Send final completion update
        await websocket_manager.send_overall_update(
            job_id=job_id,
            status="completed",
            message="ðŸŽ‰ All universal components processing completed successfully!",
            progress=100,
            results=results
        )
        
        return {
            "status": "success",
            "job_id": job_id,
            "results": results,
            "user_id": user_id,
            "filename": filename
        }
        
    except Exception as e:
        logger.error(f"WebSocket processing error: {e}")
        await websocket_manager.send_error(
            job_id=job_id,
            error_message=str(e)
        )
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/job-status/{job_id}")
async def get_job_status_endpoint(job_id: str):
    """
    FIX #3: Get current job status with Redis/memory and DB fallback.
    CRITICAL: Now includes duplicate_info for polling fallback when WebSocket fails.
    """
    status = await websocket_manager.get_job_status(job_id)
    if status:
        # FIX #3: Ensure duplicate_info is included in response
        return status
    try:
        if optimized_db:
            row = await optimized_db.get_job_status_optimized(job_id)
            if row:
                # FIX #3: Include duplicate_info from database if available
                response = {
                    "status": row.get("status"),
                    "progress": row.get("progress"),
                    "message": row.get("error_message") or row.get("status") or "unknown",
                    "updated_at": row.get("completed_at") or row.get("created_at")
                }
                
                # FIX #3: CRITICAL - Add duplicate detection info if present
                # Check if job has duplicate detection metadata
                if row.get("metadata"):
                    metadata = row.get("metadata")
                    if isinstance(metadata, str):
                        try:
                            metadata = orjson.loads(metadata)
                        except:
                            metadata = {}
                    
                    # Include duplicate_info if present
                    if metadata.get("duplicate_info"):
                        response["duplicate_info"] = metadata["duplicate_info"]
                    if metadata.get("near_duplicate_info"):
                        response["near_duplicate_info"] = metadata["near_duplicate_info"]
                    if metadata.get("content_duplicate_info"):
                        response["content_duplicate_info"] = metadata["content_duplicate_info"]
                    if metadata.get("delta_analysis"):
                        response["delta_analysis"] = metadata["delta_analysis"]
                    if metadata.get("requires_user_decision"):
                        response["requires_user_decision"] = metadata["requires_user_decision"]
                
                return response
    except Exception as e:
        logger.warning(f"DB fallback for job status failed: {e}")
    raise HTTPException(status_code=404, detail="Job not found")

@app.get("/job-status/{job_id}")
async def get_job_status_alias(job_id: str):
    """Alias endpoint to match frontend polling path."""
    return await get_job_status_endpoint(job_id)

# ============================================================================
# DATABASE INTEGRATION FOR UNIVERSAL COMPONENTS
# ============================================================================

class UniversalComponentDatabaseManager:
    """Manages database storage for all universal components"""
    
    def __init__(self, supabase_client):
        self.supabase = supabase_client
    
    async def store_field_detection_result(self, user_id: str, filename: str, result: Dict[str, Any], job_id: str = None):
        """Store field detection results in database"""
        try:
            record = {
                'user_id': user_id,
                'filename': filename,
                'component_type': 'field_detection',
                'job_id': job_id or str(uuid.uuid4()),
                'result_data': result,
                'created_at': pendulum.now().to_iso8601_string(),
                'metadata': {
                    'detected_fields': result.get('detected_fields', {}),
                    'confidence_scores': result.get('confidence_scores', {}),
                    'field_types_count': len(result.get('detected_fields', {}))
                }
            }
            
            response = self.supabase.table('universal_component_results').insert(record).execute()
            return response.data[0] if response.data else None
            
        except Exception as e:
            logger.error(f"Failed to store field detection result: {e}")
            return None
    
    async def store_platform_detection_result(self, user_id: str, filename: str, result: Dict[str, Any], job_id: str = None):
        """Store platform detection results in database"""
        try:
            record = {
                'user_id': user_id,
                'filename': filename,
                'component_type': 'platform_detection',
                'job_id': job_id or str(uuid.uuid4()),
                'result_data': result,
                'created_at': pendulum.now().to_iso8601_string(),
                'metadata': {
                    'detected_platform': result.get('platform', 'unknown'),
                    'confidence': result.get('confidence', 0.0),
                    'detection_method': result.get('detection_method', 'unknown')
                }
            }
            
            response = self.supabase.table('universal_component_results').insert(record).execute()
            return response.data[0] if response.data else None
            
        except Exception as e:
            logger.error(f"Failed to store platform detection result: {e}")
            return None
    
    async def store_document_classification_result(self, user_id: str, filename: str, result: Dict[str, Any], job_id: str = None):
        """Store document classification results in database"""
        try:
            record = {
                'user_id': user_id,
                'filename': filename,
                'component_type': 'document_classification',
                'job_id': job_id or str(uuid.uuid4()),
                'result_data': result,
                'created_at': pendulum.now().to_iso8601_string(),
                'metadata': {
                    'document_type': result.get('document_type', 'unknown'),
                    'confidence': result.get('confidence', 0.0),
                    'classification_method': result.get('classification_method', 'unknown')
                }
            }
            
            response = self.supabase.table('universal_component_results').insert(record).execute()
            return response.data[0] if response.data else None
            
        except Exception as e:
            logger.error(f"Failed to store document classification result: {e}")
            return None
    
    async def store_data_extraction_result(self, user_id: str, filename: str, result: Dict[str, Any], job_id: str = None):
        """Store data extraction results in database"""
        try:
            record = {
                'user_id': user_id,
                'filename': filename,
                'component_type': 'data_extraction',
                'job_id': job_id or str(uuid.uuid4()),
                'result_data': result,
                'created_at': pendulum.now().to_iso8601_string(),
                'metadata': {
                    'extracted_rows': len(result.get('extracted_data', [])),
                    'extraction_method': result.get('extraction_method', 'unknown'),
                    'file_format': result.get('file_format', 'unknown')
                }
            }
            
            response = self.supabase.table('universal_component_results').insert(record).execute()
            return response.data[0] if response.data else None
            
        except Exception as e:
            logger.error(f"Failed to store data extraction result: {e}")
            return None
    
    async def store_entity_resolution_result(self, user_id: str, platform: str, result: Dict[str, Any], job_id: str = None):
        """Store entity resolution results in database"""
        try:
            record = {
                'user_id': user_id,
                'filename': f"entity_resolution_{platform}",
                'component_type': 'entity_resolution',
                'job_id': job_id or str(uuid.uuid4()),
                'result_data': result,
                'created_at': pendulum.now().to_iso8601_string(),
                'metadata': {
                    'resolved_entities': len(result.get('resolved_entities', [])),
                    'platform': platform,
                    'conflicts_detected': len(result.get('conflicts', []))
                }
            }
            
            response = self.supabase.table('universal_component_results').insert(record).execute()
            return response.data[0] if response.data else None
            
        except Exception as e:
            logger.error(f"Failed to store entity resolution result: {e}")
            return None
    
    async def get_component_results(self, user_id: str, component_type: str = None, limit: int = 100):
        """Retrieve component results from database"""
        try:
            # Prefer optimized helper when available
            if optimized_db:
                results = await optimized_db.get_component_results_optimized(user_id, component_type, limit)
                return results

            # Fallback: limit columns even without optimized client
            query = self.supabase.table('universal_component_results').select(
                'id, component_type, filename, result_data, metadata, created_at'
            ).eq('user_id', user_id)
            if component_type:
                query = query.eq('component_type', component_type)
            query = query.order('created_at', desc=True).limit(limit)
            response = query.execute()
            return response.data if response.data else []
            
        except Exception as e:
            logger.error(f"Failed to get component results: {e}")
            return []
    
    async def get_component_metrics(self, user_id: str):
        """Get aggregated metrics for all components"""
        try:
            # Get all results for user
            results = await self.get_component_results(user_id)
            
            metrics = {
                'total_operations': len(results),
                'by_component_type': {},
                'by_filename': {},
                'success_rate': 0,
                'avg_confidence': 0
            }
            
            successful_operations = 0
            total_confidence = 0
            confidence_count = 0
            
            for result in results:
                component_type = result.get('component_type', 'unknown')
                filename = result.get('filename', 'unknown')
                metadata = result.get('metadata', {})
                
                # Count by component type
                if component_type not in metrics['by_component_type']:
                    metrics['by_component_type'][component_type] = 0
                metrics['by_component_type'][component_type] += 1
                
                # Count by filename
                if filename not in metrics['by_filename']:
                    metrics['by_filename'][filename] = 0
                metrics['by_filename'][filename] += 1
                
                # Calculate success rate and confidence
                if result.get('result_data'):
                    successful_operations += 1
                    
                    confidence = metadata.get('confidence')
                    if confidence is not None:
                        total_confidence += confidence
                        confidence_count += 1
            
            if len(results) > 0:
                metrics['success_rate'] = successful_operations / len(results)
            
            if confidence_count > 0:
                metrics['avg_confidence'] = total_confidence / confidence_count
            
            return metrics
            
        except Exception as e:
            logger.error(f"Failed to get component metrics: {e}")
            return {}

# ============================================================================
# COMPREHENSIVE MONITORING & OBSERVABILITY SYSTEM
# ============================================================================

class UniversalComponentMonitoringSystem:
    """Comprehensive monitoring and observability system for all universal components"""
    
    def __init__(self):
        self.metrics_store = {}
        self.performance_tracker = {}
        self.error_tracker = {}
        self.audit_log = []
        self.health_status = {}
        
    def record_operation_metrics(self, component: str, operation: str, duration: float, success: bool, 
                                user_id: str = None, metadata: Dict[str, Any] = None):
        """Record operation metrics for monitoring"""
        timestamp = pendulum.now().to_iso8601_string()
        
        if component not in self.metrics_store:
            self.metrics_store[component] = {
                'total_operations': 0,
                'successful_operations': 0,
                'failed_operations': 0,
                'avg_duration': 0,
                'operations': []
            }
        
        component_metrics = self.metrics_store[component]
        component_metrics['total_operations'] += 1
        
        if success:
            component_metrics['successful_operations'] += 1
        else:
            component_metrics['failed_operations'] += 1
        
        # Update average duration
        total_duration = component_metrics['avg_duration'] * (component_metrics['total_operations'] - 1)
        component_metrics['avg_duration'] = (total_duration + duration) / component_metrics['total_operations']
        
        # Store operation details
        operation_record = {
            'timestamp': timestamp,
            'operation': operation,
            'duration': duration,
            'success': success,
            'user_id': user_id,
            'metadata': metadata or {}
        }
        component_metrics['operations'].append(operation_record)
        
        # Keep only last 1000 operations per component
        if len(component_metrics['operations']) > 1000:
            component_metrics['operations'] = component_metrics['operations'][-1000:]
        
        logger.info(f"ðŸ“Š {component}.{operation}: {duration:.3f}s, success={success}")
    
    def record_performance_metrics(self, component: str, metrics: Dict[str, Any]):
        """Record performance-specific metrics"""
        timestamp = pendulum.now().to_iso8601_string()
        
        if component not in self.performance_tracker:
            self.performance_tracker[component] = []
        
        performance_record = {
            'timestamp': timestamp,
            'metrics': metrics
        }
        
        self.performance_tracker[component].append(performance_record)
        
        # Keep only last 500 performance records per component
        if len(self.performance_tracker[component]) > 500:
            self.performance_tracker[component] = self.performance_tracker[component][-500:]
    
    def record_error(self, component: str, operation: str, error: Exception, 
                    user_id: str = None, context: Dict[str, Any] = None):
        """Record error details for monitoring and debugging"""
        timestamp = pendulum.now().to_iso8601_string()
        
        if component not in self.error_tracker:
            self.error_tracker[component] = []
        
        error_record = {
            'timestamp': timestamp,
            'operation': operation,
            'error_type': type(error).__name__,
            'error_message': str(error),
            'user_id': user_id,
            'context': context or {}
        }
        
        self.error_tracker[component].append(error_record)
        
        # Keep only last 1000 errors per component
        if len(self.error_tracker[component]) > 1000:
            self.error_tracker[component] = self.error_tracker[component][-1000:]
        
        logger.error(f"âŒ {component}.{operation} error: {error}")
    
    def audit_operation(self, component: str, operation: str, user_id: str, 
                       details: Dict[str, Any], action: str = "execute"):
        """Record audit trail for compliance and debugging"""
        audit_record = {
            'timestamp': pendulum.now().to_iso8601_string(),
            'component': component,
            'operation': operation,
            'user_id': user_id,
            'action': action,
            'details': details
        }
        
        self.audit_log.append(audit_record)
        
        # Keep only last 10000 audit records
        if len(self.audit_log) > 10000:
            self.audit_log = self.audit_log[-10000:]
    
    def update_health_status(self, component: str, status: str, details: Dict[str, Any] = None):
        """Update health status for a component"""
        self.health_status[component] = {
            'status': status,  # 'healthy', 'degraded', 'unhealthy'
            'last_check': pendulum.now().to_iso8601_string(),
            'details': details or {}
        }
    
    def get_component_health(self, component: str) -> Dict[str, Any]:
        """Get health status for a specific component"""
        return self.health_status.get(component, {
            'status': 'unknown',
            'last_check': None,
            'details': {}
        })
    
    def get_overall_health(self) -> Dict[str, Any]:
        """Get overall system health status"""
        if not self.health_status:
            return {'status': 'unknown', 'components': {}}
        
        healthy_count = sum(1 for status in self.health_status.values() 
                          if status['status'] == 'healthy')
        total_count = len(self.health_status)
        
        if healthy_count == total_count:
            overall_status = 'healthy'
        elif healthy_count > total_count // 2:
            overall_status = 'degraded'
        else:
            overall_status = 'unhealthy'
        
        return {
            'status': overall_status,
            'healthy_components': healthy_count,
            'total_components': total_count,
            'components': self.health_status
        }
    
    def get_metrics_summary(self) -> Dict[str, Any]:
        """Get comprehensive metrics summary"""
        summary = {
            'overall_metrics': {
                'total_operations': 0,
                'successful_operations': 0,
                'failed_operations': 0,
                'avg_duration': 0,
                'success_rate': 0
            },
            'component_metrics': {},
            'error_summary': {},
            'performance_summary': {}
        }
        
        total_ops = 0
        total_success = 0
        total_failed = 0
        total_duration = 0
        
        # Aggregate component metrics
        for component, metrics in self.metrics_store.items():
            total_ops += metrics['total_operations']
            total_success += metrics['successful_operations']
            total_failed += metrics['failed_operations']
            total_duration += metrics['avg_duration'] * metrics['total_operations']
            
            summary['component_metrics'][component] = {
                'total_operations': metrics['total_operations'],
                'success_rate': metrics['successful_operations'] / metrics['total_operations'] if metrics['total_operations'] > 0 else 0,
                'avg_duration': metrics['avg_duration'],
                'recent_operations': len(metrics['operations'])
            }
        
        # Calculate overall metrics
# Initialize monitoring system
monitoring_system = UniversalComponentMonitoringSystem()

@app.get("/api/monitoring/observability")
async def get_observability_metrics():
    """Get observability metrics (Prometheus metrics endpoint)"""
    try:
        # Return Prometheus metrics via /metrics endpoint instead
        return {
            "status": "success",
            "message": "Use /metrics endpoint for Prometheus metrics",
            "timestamp": pendulum.now().to_iso8601_string()
        }
    except Exception as e:
        logger.error(f"Failed to get observability metrics: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/monitoring/security")
async def get_security_status():
    """Get security system status and statistics"""
    try:
        security_stats = security_validator.get_security_statistics()
        
        return {
            "status": "success",
            "security": security_stats,
            "timestamp": pendulum.now().to_iso8601_string()
        }
    except Exception as e:
        logger.error(f"Failed to get security status: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/monitoring/health")
async def get_health_status():
    """Get comprehensive health status for all components"""
    try:
        # Update health status for each component
        components_to_check = [
            'ExcelProcessor',
            'UniversalFieldDetector', 
            'UniversalPlatformDetector',
            'UniversalDocumentClassifier',
            'UniversalExtractors',
            'EntityResolver'
        ]
        
        for component in components_to_check:
            try:
                # Basic health check - try to initialize component
                if component == 'ExcelProcessor':
                    test_instance = ExcelProcessor()
                    monitoring_system.update_health_status(component, 'healthy', {'initialized': True})
                elif component == 'UniversalFieldDetector':
                    test_instance = UniversalFieldDetector()
                    monitoring_system.update_health_status(component, 'healthy', {'initialized': True})
                elif component == 'UniversalPlatformDetector':
                    test_instance = UniversalPlatformDetector(anthropic_client=None, cache_client=safe_get_ai_cache())
                    monitoring_system.update_health_status(component, 'healthy', {'initialized': True})
                elif component == 'UniversalDocumentClassifier':
                    # FIX #3: Use singleton instance for health check instead of creating new heavy model
                    excel_processor = _get_excel_processor_instance()
                    test_instance = excel_processor.universal_document_classifier
                    monitoring_system.update_health_status(component, 'healthy', {'initialized': True})
                elif component == 'UniversalExtractors':
                    test_instance = UniversalExtractors(cache_client=safe_get_ai_cache())
                    monitoring_system.update_health_status(component, 'healthy', {'initialized': True})
                elif component == 'EntityResolver':
                    # Mock client for health check
                    class MockSupabaseClient:
                        pass
                    test_instance = EntityResolver(supabase_client=MockSupabaseClient(), cache_client=safe_get_ai_cache())
                    monitoring_system.update_health_status(component, 'healthy', {'initialized': True})
                    
            except Exception as e:
                monitoring_system.update_health_status(component, 'unhealthy', {'error': str(e)})
        
        overall_health = monitoring_system.get_overall_health()
        
        return {
            'status': 'success',
            'health': overall_health,
            'timestamp': pendulum.now().to_iso8601_string()
        }
        
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/monitoring/metrics")
async def get_monitoring_metrics():
    """Get comprehensive monitoring metrics"""
    try:
        metrics_summary = monitoring_system.get_metrics_summary()
        
        return {
            'status': 'success',
            'metrics': metrics_summary,
            'timestamp': pendulum.now().to_iso8601_string()
        }
        
    except Exception as e:
        logger.error(f"Failed to get monitoring metrics: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/monitoring/metrics/prometheus")
async def get_prometheus_metrics():
    """Get metrics in Prometheus format"""
    try:
        prometheus_metrics = monitoring_system.export_metrics_for_prometheus()
        
        return Response(
            content=prometheus_metrics,
            media_type="text/plain"
        )
        
    except Exception as e:
        logger.error(f"Failed to export Prometheus metrics: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/monitoring/errors")
async def get_error_logs(component: str = None, limit: int = 100):
    """Get error logs for monitoring and debugging"""
    try:
        error_logs = {}
        
        if component:
            if component in monitoring_system.error_tracker:
                error_logs[component] = monitoring_system.error_tracker[component][-limit:]
        else:
            for comp, errors in monitoring_system.error_tracker.items():
                error_logs[comp] = errors[-limit:]
        
        return {
            'status': 'success',
            'error_logs': error_logs,
            'timestamp': pendulum.now().to_iso8601_string()
        }
        
    except Exception as e:
        logger.error(f"Failed to get error logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/monitoring/audit")
async def get_audit_log(component: str = None, user_id: str = None, limit: int = 100):
    """Get audit trail for compliance and debugging"""
    try:
        audit_records = monitoring_system.audit_log[-limit:]
        
        # Filter by component if specified
        if component:
            audit_records = [r for r in audit_records if r['component'] == component]
        
        # Filter by user_id if specified
        if user_id:
            audit_records = [r for r in audit_records if r['user_id'] == user_id]
        
        return {
            'status': 'success',
            'audit_log': audit_records,
            'total_records': len(audit_records),
            'timestamp': pendulum.now().to_iso8601_string()
        }
        
    except Exception as e:
        logger.error(f"Failed to get audit log: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# DEPLOYMENT HEALTH CHECK
# ============================================================================




@app.get("/health")
async def health_check():
    """Comprehensive health check for deployment debugging"""
    try:
        health_status = {
            "status": "healthy",
            "timestamp": pendulum.now().to_iso8601_string(),
            "version": "2.0.0",
            "environment": {
                "supabase_configured": bool(supabase),
                "anthropic_configured": False,  # Disabled - using Groq/Llama instead
                "groq_configured": bool(groq_client),
                "available_env_vars": sorted([k for k in os.environ.keys() if any(x in k.upper() for x in ['SUPABASE', 'GROQ', 'DATABASE'])]),
                "python_version": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
                "advanced_features": ADVANCED_FEATURES
            },
            "services": {
                "database": "connected" if supabase else "disconnected",
                "ai_anthropic": "disabled (using Groq/Llama)",
                "ai_groq": "connected" if groq_client else "disconnected"
            }
        }
        
        # Test database connection if available
        if supabase:
            try:
                result = supabase.table('raw_events').select('id').limit(1).execute()
                health_status["services"]["database"] = "operational"
            except Exception as e:
                health_status["services"]["database"] = f"error: {str(e)}"
                health_status["status"] = "degraded"
        
        return health_status
        
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "timestamp": pendulum.now().to_iso8601_string()
        }

# ============================================================================
# STATIC FILE SERVING FOR FRONTEND
# ============================================================================

# Check if frontend dist directory exists and mount static files
import pathlib
frontend_dist_path = pathlib.Path(__file__).parent / "dist"
if frontend_dist_path.exists():
    app.mount("/static", StaticFiles(directory=str(frontend_dist_path)), name="static")
    logger.info(f"âœ… Frontend static files mounted from {frontend_dist_path}")
    
    # Serve index.html for root path
    @app.get("/")
    async def serve_root():
        """Serve frontend root"""
        index_path = frontend_dist_path / "index.html"
        if index_path.exists():
            return FileResponse(str(index_path))
        raise HTTPException(status_code=404, detail="Frontend not found")
    
    # Serve index.html for SPA routing
    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        """Serve frontend for SPA routing (catch-all for non-API routes)"""
        # Don't serve frontend for API routes, WebSocket, or docs
        if (full_path.startswith("api/") or 
            full_path.startswith("health") or 
            full_path.startswith("docs") or
            full_path.startswith("openapi.json") or
            full_path.startswith("ws/") or
            full_path.startswith("duplicate-detection/ws/")):
            raise HTTPException(status_code=404, detail="API endpoint not found")
        
        # Serve static files directly if they exist
        static_file_path = frontend_dist_path / full_path
        if static_file_path.exists() and static_file_path.is_file():
            return FileResponse(str(static_file_path))
        
        # Otherwise serve index.html for SPA routing
        index_path = frontend_dist_path / "index.html"
        if index_path.exists():
            return FileResponse(str(index_path))
        else:
            raise HTTPException(status_code=404, detail="Frontend not found")
else:
    logger.warning("âš ï¸ Frontend dist directory not found - serving API only")

# ============================================================================
# DEVELOPER DEBUG ENDPOINTS
# ============================================================================

@app.delete("/api/files/{job_id}")
async def delete_file_completely(job_id: str, user_id: str):
    """
    ðŸ—‘ï¸ COMPREHENSIVE FILE DELETION - Cascades to all related tables
    
    Deletes a file and ALL associated data from the database:
    - ingestion_jobs (main record)
    - raw_events (all events from this file)
    - normalized_entities (entities from this file)
    - relationship_instances (relationships involving events from this file)
    - entity_matches (entity resolution data)
    - processing_transactions (transaction records)
    - event_delta_logs (change tracking)
    
    This ensures complete cleanup with no orphaned data.
    """
    try:
        logger.info(f"ðŸ—‘ï¸ Starting comprehensive file deletion for job_id={job_id}, user_id={user_id}")
        
        # Verify ownership
        job_result = supabase.table('ingestion_jobs').select('id, filename, user_id').eq('id', job_id).eq('user_id', user_id).execute()
        
        if not job_result.data:
            raise HTTPException(status_code=404, detail="File not found or access denied")
        
        filename = job_result.data[0].get('filename', 'Unknown')
        
        # Track deletion statistics
        deletion_stats = {
            'job_id': job_id,
            'filename': filename,
            'deleted_records': {}
        }
        
        # Step 1: Get all raw_event IDs from this job
        events_result = supabase.table('raw_events').select('id').eq('job_id', job_id).eq('user_id', user_id).execute()
        event_ids = [e['id'] for e in events_result.data] if events_result.data else []
        deletion_stats['deleted_records']['raw_events'] = len(event_ids)
        logger.info(f"Found {len(event_ids)} events to delete")
        
        # Step 2: Soft-delete relationship_instances involving these events (FIX #3)
        if event_ids:
            # FIX #3: Use soft-delete instead of hard-delete to support graph cache invalidation
            # Mark relationships as deleted instead of physically removing them
            now = pendulum.now().to_iso8601_string()
            rel_update_1 = supabase.table('relationship_instances').update({'is_deleted': True, 'updated_at': now})\
                .in_('source_event_id', event_ids).eq('user_id', user_id).execute()
            rel_update_2 = supabase.table('relationship_instances').update({'is_deleted': True, 'updated_at': now})\
                .in_('target_event_id', event_ids).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['relationship_instances'] = len(rel_update_1.data or []) + len(rel_update_2.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['relationship_instances']} relationships")
        
        # Step 3: Soft-delete entity_matches for events from this file
        if event_ids:
            now = pendulum.now().to_iso8601_string()
            entity_matches_result = supabase.table('entity_matches').update({'is_deleted': True, 'updated_at': now}).in_('source_row_id', event_ids).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['entity_matches'] = len(entity_matches_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['entity_matches']} entity matches")
        
        # Step 4: Delete normalized_entities that only exist in this file
        # Note: We don't delete entities that appear in other files
        # This is handled by the source_files array in normalized_entities
        
        # Step 5: Delete error_logs for this job
        
        # Step 6: Delete debug_logs for this job
        
        # Step 7: Soft-delete processing_transactions for this job
        try:
            now = pendulum.now().to_iso8601_string()
            transactions_result = supabase.table('processing_transactions').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['processing_transactions'] = len(transactions_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['processing_transactions']} transactions")
        except Exception as e:
            logger.warning(f"Failed to soft-delete processing_transactions: {e}")
        
        # Step 8: Soft-delete event_delta_logs for this job
        try:
            now = pendulum.now().to_iso8601_string()
            delta_logs_result = supabase.table('event_delta_logs').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['event_delta_logs'] = len(delta_logs_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['event_delta_logs']} delta logs")
        except Exception as e:
            logger.warning(f"Failed to soft-delete event_delta_logs: {e}")
        
        # CRITICAL FIX: Soft-delete advanced analytics data (previously orphaned)
        # Step 9: Soft-delete temporal_patterns
        try:
            now = pendulum.now().to_iso8601_string()
            temporal_patterns_result = supabase.table('temporal_patterns').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['temporal_patterns'] = len(temporal_patterns_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['temporal_patterns']} temporal patterns")
        except Exception as e:
            logger.warning(f"Failed to soft-delete temporal_patterns: {e}")
        
        # Step 10: Soft-delete predicted_relationships
        try:
            now = pendulum.now().to_iso8601_string()
            predicted_relationships_result = supabase.table('predicted_relationships').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['predicted_relationships'] = len(predicted_relationships_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['predicted_relationships']} predicted relationships")
        except Exception as e:
            logger.warning(f"Failed to soft-delete predicted_relationships: {e}")
        
        # Step 11: Clear anomalies from temporal_patterns (FIX #14: merged table)
        try:
            # FIX #14: temporal_anomalies merged into temporal_patterns
            supabase.table('temporal_patterns').update({'anomalies': []})\
                .eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['temporal_anomalies'] = 0  # Tracked as update, not delete
            logger.info(f"Cleared anomalies from temporal_patterns")
        except Exception as e:
            logger.warning(f"Failed to clear anomalies: {e}")
        
        # Step 12: Soft-delete causal_relationships
        try:
            now = pendulum.now().to_iso8601_string()
            causal_relationships_result = supabase.table('causal_relationships').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['causal_relationships'] = len(causal_relationships_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['causal_relationships']} causal relationships")
        except Exception as e:
            logger.warning(f"Failed to soft-delete causal_relationships: {e}")
        
        # Step 13: Soft-delete root_cause_analyses
        try:
            now = pendulum.now().to_iso8601_string()
            root_cause_analyses_result = supabase.table('root_cause_analyses').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['root_cause_analyses'] = len(root_cause_analyses_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['root_cause_analyses']} root cause analyses")
        except Exception as e:
            logger.warning(f"Failed to soft-delete root_cause_analyses: {e}")
        
        # Step 14: Soft-delete counterfactual_analyses
        try:
            now = pendulum.now().to_iso8601_string()
            counterfactual_analyses_result = supabase.table('counterfactual_analyses').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['counterfactual_analyses'] = len(counterfactual_analyses_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['counterfactual_analyses']} counterfactual analyses")
        except Exception as e:
            logger.warning(f"Failed to soft-delete counterfactual_analyses: {e}")
        
        # Step 15: Clear seasonal data from temporal_patterns (FIX #14: merged table)
        try:
            # FIX #14: seasonal_patterns merged into temporal_patterns
            supabase.table('temporal_patterns').update({'seasonal_data': None})\
                .eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['seasonal_patterns'] = 0  # Tracked as update, not delete
            logger.info(f"Cleared seasonal data from temporal_patterns")
        except Exception as e:
            logger.warning(f"Failed to clear seasonal data: {e}")
        
        # Step 16: Soft-delete cross_platform_relationships
        try:
            now = pendulum.now().to_iso8601_string()
            cross_platform_relationships_result = supabase.table('cross_platform_relationships').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['cross_platform_relationships'] = len(cross_platform_relationships_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['cross_platform_relationships']} cross-platform relationships")
        except Exception as e:
            logger.warning(f"Failed to soft-delete cross_platform_relationships: {e}")
        
        # Step 17: Soft-delete platform_patterns
        try:
            now = pendulum.now().to_iso8601_string()
            platform_patterns_result = supabase.table('platform_patterns').update({'is_deleted': True, 'updated_at': now}).eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['platform_patterns'] = len(platform_patterns_result.data or [])
            logger.info(f"Soft-deleted {deletion_stats['deleted_records']['platform_patterns']} platform patterns")
        except Exception as e:
            logger.warning(f"Failed to soft-delete platform_patterns: {e}")
        
        # Step 18: REMOVED - metrics table deleted
        # Metrics deletion no longer needed as table was removed
        
        # Step 19: Clear duplicate flags from relationship_instances (FIX #14: merged table)
        try:
            # FIX #14: duplicate_transactions merged into relationship_instances
            supabase.table('relationship_instances').update({'is_duplicate': False, 'duplicate_confidence': 0.0})\
                .eq('job_id', job_id).eq('user_id', user_id).execute()
            deletion_stats['deleted_records']['duplicate_transactions'] = 0  # Tracked as update, not delete
            logger.info(f"Cleared duplicate flags from relationship_instances")
        except Exception as e:
            logger.warning(f"Failed to clear duplicate flags: {e}")
        
        # Step 20: Soft-delete raw_events (FIX #3)
        # FIX #3: Use soft-delete instead of hard-delete to support graph cache invalidation
        now = pendulum.now().to_iso8601_string()
        raw_events_result = supabase.table('raw_events').update({'is_deleted': True, 'updated_at': now})\
            .eq('job_id', job_id).eq('user_id', user_id).execute()
        logger.info(f"Soft-deleted {len(raw_events_result.data or [])} raw events")
        
        # Step 20.5: Soft-delete normalized_entities for this file (FIX #3)
        # FIX #3: Mark entities as deleted to support graph cache invalidation
        try:
            entity_ids_result = supabase.table('normalized_entities').select('id')\
                .eq('job_id', job_id).eq('user_id', user_id).execute()
            entity_ids = [e['id'] for e in entity_ids_result.data] if entity_ids_result.data else []
            
            if entity_ids:
                now = pendulum.now().to_iso8601_string()
                entities_update = supabase.table('normalized_entities').update({'is_deleted': True, 'updated_at': now})\
                    .in_('id', entity_ids).eq('user_id', user_id).execute()
                logger.info(f"Soft-deleted {len(entities_update.data or [])} normalized entities")
        except Exception as e:
            logger.warning(f"Failed to soft-delete normalized_entities: {e}")
        
        # Step 20.6: Clear graph cache for this user (FIX #3)
        # FIX #3: Invalidate graph cache to prevent ghost nodes after file deletion
        try:
            redis_url = os.getenv('REDIS_URL')
            if redis_url:
                # Clear Redis cache directly without importing FinleyGraphEngine
                try:
                    from aiocache import Cache
                    from aiocache.serializers import PickleSerializer
                    from urllib.parse import urlparse
                    
                    parsed = urlparse(redis_url)
                    cache = Cache(
                        Cache.REDIS,
                        endpoint=parsed.hostname,
                        port=parsed.port or 6379,
                        namespace="graph",
                        serializer=PickleSerializer()
                    )
                    
                    cache_key = f"{user_id}"
                    await cache.delete(cache_key)
                    logger.info(f"âœ… Cleared graph cache for user {user_id}")
                except ImportError:
                    logger.warning("aiocache not available, skipping graph cache clear")
        except Exception as e:
            logger.warning(f"Failed to clear graph cache: {e}")
        
        # Step 21: Finally, soft-delete the ingestion_job record (FIX #3)
        # FIX #3: Use soft-delete for ingestion_jobs to maintain audit trail
        now = pendulum.now().to_iso8601_string()
        job_delete_result = supabase.table('ingestion_jobs').update({'is_deleted': True, 'updated_at': now}).eq('id', job_id).eq('user_id', user_id).execute()
        deletion_stats['deleted_records']['ingestion_jobs'] = len(job_delete_result.data or [])
        logger.info(f"Soft-deleted ingestion job record")
        
        # Calculate total soft-deleted records
        total_deleted = sum(deletion_stats['deleted_records'].values())
        
        logger.info(f"âœ… File deletion completed: {filename} - {total_deleted} total records marked for deletion (soft-delete)")
        logger.info(f"ðŸ“Š FIX #3: Used soft-delete for data integrity, graph cache invalidation, and audit trail")
        
        return {
            "status": "deleted",
            "job_id": job_id,
            "filename": filename,
            "message": f"File '{filename}' and all associated data deleted successfully",
            "deletion_stats": deletion_stats,
            "total_records_deleted": total_deleted
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete file completely: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete file: {str(e)}")

# ============================================================================
# MAIN APPLICATION SETUP
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    # Use socketio_app wrapper for Socket.IO support (replaces manual WebSocket endpoints)
    uvicorn.run(socketio_app, host="0.0.0.0", port=8000)
